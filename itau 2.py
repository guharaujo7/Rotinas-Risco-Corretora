import os, sys, re, random, struct, tkinter as tk, threading, time, tempfile, shutil, webbrowser, ctypes, json as _json_mod, uuid as _uuid_mod, queue, base64, sqlite3, io, csv, unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from tkinter import ttk, filedialog, messagebox, font as tkfont
from datetime import datetime, date, timedelta

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    import win32com.client as win32
    WIN32_OK = True
except Exception:
    WIN32_OK = False

try:
    import pythoncom
    PYTHONCOM_OK = True
except Exception:
    PYTHONCOM_OK = False

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_OK = True
except Exception:
    sync_playwright = None
    PLAYWRIGHT_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except Exception:
    openpyxl = None
    OPENPYXL_OK = False

try:
    import xlrd
    XLRD_OK = True
except Exception:
    xlrd = None
    XLRD_OK = False


C = {
    "bg":          "#191919",
    "surface":     "#212121",
    "surface2":    "#2a2a2a",
    "surface3":    "#333333",
    "ink":         "#e6e6e6",
    "ink_muted":   "#999999",
    "ink_faint":   "#4d4d4d",
    "accent":      "#EC7000",
    "accent_dim":  "#3d2a14",
    "accent_soft": "#2a1f12",
    "ok":          "#4ea87a",
    "ok_dim":      "#1a3a2a",
    "warn":        "#d49b45",
    "err":         "#c95f5f",
    "err_dim":     "#3d1515",
    "hair":        "#2a2a2a",
    "log_step":    "#606060",
    "log_ok":      "#4ea87a",
    "log_warn":    "#d49b45",
    "log_err":     "#c95f5f",
}

DOT_COLORS = [
    "#9b9b9b",
    "#a07450",
    "#c87941",
    "#c4a832",
    "#5a9e72",
    "#EC7000",
    "#8b72c9",
    "#c97a9e",
    "#c96060",
]
DOT_LABELS = ["Cinza","Marrom","Laranja","Amarelo","Verde","Azul","Roxo","Rosa","Vermelho"]

ICON_FILENAME = "itaulogo.png"
LOGO_FILENAME = ICON_FILENAME
APP_USER_MODEL_ID = "MesaItau.Large"

SHARED_APP_BASE_DIR = (
    r"\\BBAPROD3\fo\Diretoria de Produtos Ativos\Ativos em Reais\Risco Sacado"
    r"\Comum\AppLarge\bancodedados"
)

MESES_PT = (
    "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
)
DIAS_PT = (
    "segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
    "sexta-feira", "sábado", "domingo",
)


def format_data_pt_br(dt: datetime) -> str:
    return f"{DIAS_PT[dt.weekday()]}, {dt.day} de {MESES_PT[dt.month]} de {dt.year}"


def app_base_dir():
    return os.path.dirname(sys.executable) if getattr(sys,"frozen",False) else os.path.dirname(os.path.abspath(__file__))

def resource_path(p):
    return os.path.join(getattr(sys,"_MEIPASS",app_base_dir()), p)


def make_hairline(parent, orient="h", **kwargs):
    kw = {"bg": C["hair"]}
    kw.update(kwargs)
    if orient == "h":
        return tk.Frame(parent, height=1, **kw)
    else:
        return tk.Frame(parent, width=1, **kw)


def _make_dot(parent, color, size=10, bg=None):
    """Bolinha colorida via Canvas — estilo Notion."""
    bg = bg or parent.cget("bg")
    c = tk.Canvas(parent, width=size + 4, height=size + 4,
                  bg=bg, highlightthickness=0, bd=0)
    c.create_oval(2, 2, size + 2, size + 2, fill=color, outline="")
    return c


def _canvas_round_rect(canvas, x1, y1, x2, y2, radius, **kwargs):
    r = min(radius, (x2 - x1) / 2, (y2 - y1) / 2)
    points = [
        x1 + r, y1, x2 - r, y1,
        x2, y1, x2, y1 + r,
        x2, y2 - r, x2, y2,
        x2 - r, y2, x1 + r, y2,
        x1, y2, x1, y2 - r,
        x1, y1 + r, x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, splinesteps=36, **kwargs)


FRAME_LABELS = {
    "Home":              "Início",
    "Recebiveis":        "Recebíveis",
    "Ligacoes":          "Ligações",
}


def _icon_png_path():
    for path in (resource_path(ICON_FILENAME), os.path.join(app_base_dir(), ICON_FILENAME)):
        if os.path.isfile(path):
            return path
    return None


def _png_to_ico_bytes(png_path):
    with open(png_path, "rb") as f:
        png_data = f.read()
    if len(png_data) < 24 or png_data[:8] != b"\x89PNG\r\n\x1a\n":
        return None
    width = int.from_bytes(png_data[16:20], "big")
    height = int.from_bytes(png_data[20:24], "big")
    w_byte = 0 if width >= 256 else width
    h_byte = 0 if height >= 256 else height
    image_offset = 6 + 16
    header = struct.pack("<HHH", 0, 1, 1)
    entry = struct.pack("<BBBBHHII", w_byte, h_byte, 0, 0, 1, 32, len(png_data), image_offset)
    return header + entry + png_data


def _build_ico_from_png(png_path, ico_path):
    try:
        from PIL import Image
        img = Image.open(png_path).convert("RGBA")
        img.save(ico_path, format="ICO", sizes=[(16, 16), (32, 32), (48, 48), (256, 256)])
        return True
    except Exception:
        return False


def _ensure_ico_path():
    cached = os.path.join(app_base_dir(), "itaulogo.ico")
    png_path = _icon_png_path()
    if png_path:
        needs_build = not os.path.isfile(cached)
        if needs_build and _build_ico_from_png(png_path, cached):
            return cached
    if os.path.isfile(cached):
        return cached
    if not png_path:
        return None
    ico_bytes = _png_to_ico_bytes(png_path)
    if not ico_bytes:
        return None
    try:
        with open(cached, "wb") as f:
            f.write(ico_bytes)
        return cached
    except OSError:
        try:
            fd, tmp = tempfile.mkstemp(suffix=".ico", prefix="mesa_itau_")
            os.close(fd)
            with open(tmp, "wb") as f:
                f.write(ico_bytes)
            return tmp
        except OSError:
            return None


def apply_taskbar_presence(root):
    if sys.platform != "win32":
        return
    try:
        GWL_EXSTYLE = -20
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        SW_HIDE = 0
        SW_SHOW = 5
        root.update_idletasks()
        hwnd = _window_hwnd(root)
        if not hwnd:
            return
        style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        new_style = (style & ~WS_EX_TOOLWINDOW) | WS_EX_APPWINDOW
        if new_style != style:
            ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, new_style)
        ctypes.windll.user32.ShowWindow(hwnd, SW_HIDE)
        ctypes.windll.user32.ShowWindow(hwnd, SW_SHOW)
        root.lift()
    except Exception:
        pass


def apply_window_icon(root):
    ico_path = _ensure_ico_path()
    if not ico_path:
        return
    try:
        root.iconbitmap(default=ico_path)
    except Exception:
        pass
    if sys.platform == "win32":
        try:
            root.update_idletasks()
            hwnd = _window_hwnd(root)
            if not hwnd:
                return
            IMAGE_ICON = 1
            LR_LOADFROMFILE = 0x10
            WM_SETICON = 0x0080
            for size in (16, 32):
                hicon = ctypes.windll.user32.LoadImageW(
                    None, ico_path, IMAGE_ICON, size, size, LR_LOADFROMFILE,
                )
                if hicon:
                    which = 0 if size <= 16 else 1
                    ctypes.windll.user32.SendMessageW(hwnd, WM_SETICON, which, hicon)
        except Exception:
            pass


def apply_windows_shell(root):
    apply_taskbar_presence(root)
    apply_window_icon(root)


def apply_modern_window_chrome(root):
    if sys.platform != "win32":
        return
    try:
        hwnd = _window_hwnd(root)
        dwm = ctypes.windll.dwmapi
        dark = ctypes.c_int(1)
        round_pref = ctypes.c_int(2)
        dwm.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(dark), ctypes.sizeof(dark))
        dwm.DwmSetWindowAttribute(hwnd, 33, ctypes.byref(round_pref), ctypes.sizeof(round_pref))
    except Exception:
        pass


def _window_hwnd(root):
    root.update_idletasks()
    return ctypes.windll.user32.GetParent(root.winfo_id())


def apply_frameless_resize(root):
    if sys.platform != "win32":
        return
    try:
        hwnd = _window_hwnd(root)
        gwl_style = -16
        style = ctypes.windll.user32.GetWindowLongW(hwnd, gwl_style)
        style |= 0x00040000
        style |= 0x00020000
        style |= 0x00010000
        ctypes.windll.user32.SetWindowLongW(hwnd, gwl_style, style)
    except Exception:
        pass


def start_native_window_drag(root):
    """Inicia o arrasto nativo da janela (estilo WM_NCLBUTTONDOWN/HTCAPTION).

    Usa PostMessageW (assíncrono) em vez de SendMessageW: o SendMessageW
    entra no loop modal de mover janela do Windows de forma síncrona,
    dentro da chamada ctypes — que libera o GIL antes de chamar a API do
    Windows. Qualquer callback do Tcl/Tk disparado durante esse loop modal
    tenta reaquisitar o GIL num estado inconsistente, causando
    'Fatal Python error: PyEval_RestoreThread' e o crash da aplicação.
    PostMessageW apenas enfileira a mensagem e retorna imediatamente; o
    loop de mover é processado depois, dentro do laço normal de eventos
    do Tcl, onde o GIL é gerenciado corretamente.
    """
    if sys.platform != "win32":
        return False
    try:
        hwnd = _window_hwnd(root)
        if not hwnd:
            return False
        WM_NCLBUTTONDOWN = 0x00A1
        HTCAPTION = 2
        user32 = ctypes.windll.user32
        user32.ReleaseCapture()
        user32.PostMessageW(ctypes.c_void_p(hwnd), WM_NCLBUTTONDOWN,
                            ctypes.c_void_p(HTCAPTION), ctypes.c_void_p(0))
        return True
    except Exception:
        return False


class AppTitleBar(tk.Frame):
    BG = "#1c1c1c"
    HEIGHT = 36

    def __init__(self, parent, root):
        super().__init__(parent, bg=self.BG, height=self.HEIGHT)
        self.pack_propagate(False)
        self.root = root
        self._drag_offset = None
        self._maximized = False

        row = tk.Frame(self, bg=self.BG)
        row.pack(fill="both", expand=True)

        left = tk.Frame(row, bg=self.BG)
        left.pack(side="left", fill="y", padx=(14, 0))

        tk.Label(left, text="Mesa", bg=self.BG, fg=C["ink"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", pady=8)
        tk.Label(left, text="Itaú", bg=self.BG, fg=C["accent"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=(3, 0), pady=8)
        tk.Label(left, text="·", bg=self.BG, fg="#404040",
                 font=("Segoe UI", 9)).pack(side="left", padx=8, pady=8)
        self._module_lbl = tk.Label(left, text="Início", bg=self.BG, fg="#8a8a8a",
                                    font=("Segoe UI", 9))
        self._module_lbl.pack(side="left", pady=8)

        controls = tk.Frame(row, bg=self.BG)
        controls.pack(side="right", fill="y")

        self._btn_min = self._win_btn(controls, "─", self._minimize)
        self._btn_max = self._win_btn(controls, "□", self._toggle_maximize)
        self._btn_close = self._win_btn(controls, "✕", self.root.destroy, close=True)

        tk.Frame(self, bg="#2e2e2e", height=1).pack(fill="x", side="bottom")

        self._bind_drag(self)
        self._bind_drag(row)
        self._bind_drag(left)
        for w in left.winfo_children():
            if w is not self._module_lbl:
                self._bind_drag(w)
        self._bind_drag(self._module_lbl)
        self.bind("<Double-Button-1>", lambda _e: self._toggle_maximize())

    def _win_btn(self, parent, text, command, close=False):
        hover = "#c95f5f" if close else "#333333"
        lbl = tk.Label(parent, text=text, bg=self.BG, fg="#9a9a9a",
                       font=("Segoe UI", 9), width=4, cursor="hand2")
        lbl.pack(side="left", fill="y")
        lbl.bind("<Button-1>", lambda _e: command())
        lbl.bind("<Enter>", lambda _e, l=lbl, h=hover: l.configure(bg=h, fg="#f2f2f2"))
        lbl.bind("<Leave>", lambda _e, l=lbl: l.configure(bg=self.BG, fg="#9a9a9a"))
        return lbl

    def _bind_drag(self, widget):
        widget.bind("<ButtonPress-1>", self._start_drag, add="+")
        if sys.platform != "win32":
            widget.bind("<B1-Motion>", self._on_drag_fallback, add="+")

    def _start_drag(self, event):
        if start_native_window_drag(self.root):
            return
        if self._maximized:
            return
        self._drag_offset = (event.x_root - self.root.winfo_x(),
                             event.y_root - self.root.winfo_y())

    def _on_drag_fallback(self, event):
        if not self._drag_offset or self._maximized:
            return
        ox, oy = self._drag_offset
        self.root.geometry(f"+{event.x_root - ox}+{event.y_root - oy}")

    def _minimize(self):
        if sys.platform == "win32":
            try:
                ctypes.windll.user32.ShowWindow(_window_hwnd(self.root), 6)
                return
            except Exception:
                pass
        self.root.iconify()

    def _toggle_maximize(self):
        if sys.platform == "win32":
            try:
                hwnd = _window_hwnd(self.root)
                if self._maximized:
                    ctypes.windll.user32.ShowWindow(hwnd, 9)
                    self._maximized = False
                    self._btn_max.configure(text="□")
                else:
                    ctypes.windll.user32.ShowWindow(hwnd, 3)
                    self._maximized = True
                    self._btn_max.configure(text="❐")
                return
            except Exception:
                pass
        if self._maximized:
            self.root.state("normal")
            self._maximized = False
            self._btn_max.configure(text="□")
        else:
            self.root.state("zoomed")
            self._maximized = True
            self._btn_max.configure(text="❐")

    def set_module(self, frame_name):
        self._module_lbl.configure(text=FRAME_LABELS.get(frame_name, frame_name))


class AppStatusBar(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"], height=30)
        self.pack_propagate(False)
        self.controller = controller

        tk.Frame(self, bg="#303030", height=1).pack(fill="x")

        row = tk.Frame(self, bg=C["bg"])
        row.pack(fill="both", expand=True, padx=16)

        left = tk.Frame(row, bg=C["bg"])
        left.pack(side="left", fill="y")

        tk.Label(left, text="Mesa Itaú", bg=C["bg"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", pady=4)
        tk.Label(left, text="·", bg=C["bg"], fg="#3a3a3a",
                 font=("Segoe UI", 8)).pack(side="left", padx=6, pady=4)
        tk.Label(left, text="Large", bg=C["bg"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", pady=4)

        right = tk.Frame(row, bg=C["bg"])
        right.pack(side="right", fill="y")

        self._clock_lbl = tk.Label(right, text="", bg=C["bg"], fg=C["ink_faint"],
                                   font=("Segoe UI", 8))
        self._clock_lbl.pack(side="right", padx=(12, 0), pady=4)

        self._module_lbl = tk.Label(right, text="Início", bg=C["bg"], fg="#6b6b6b",
                                    font=("Segoe UI", 8))
        self._module_lbl.pack(side="right", pady=4)

        self._tick_clock()

    def set_module(self, frame_name):
        label = FRAME_LABELS.get(frame_name, frame_name)
        self._module_lbl.configure(text=label)

    def _tick_clock(self):
        now = datetime.now()
        self._clock_lbl.configure(text=now.strftime("%d/%m/%Y  %H:%M"))
        self.after(30_000, self._tick_clock)


def styled_label(parent, text, size=10, weight="normal", color=None, **kwargs):
    return tk.Label(parent, text=text,
                    font=("Segoe UI", size, weight),
                    fg=color or C["ink"],
                    bg=kwargs.pop("bg", C["surface"]),
                    **kwargs)


def styled_button(parent, text, command, accent=False, danger=False, small=False, **kwargs):
    if accent:
        bg  = C["accent_dim"];  fg  = C["accent"]
        abg = C["accent"];      afg = C["bg"]
    elif danger:
        bg  = C["err_dim"];     fg  = C["err"]
        abg = C["err"];         afg = C["bg"]
    else:
        bg  = C["surface2"];    fg  = C["ink_muted"]
        abg = C["surface3"];    afg = C["ink"]
    pad = (7, 3) if small else (13, 6)
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=fg, activebackground=abg, activeforeground=afg,
                    font=("Segoe UI", 8 if small else 9),
                    relief="flat", bd=0, padx=pad[0], pady=pad[1],
                    cursor="hand2", **kwargs)
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg,  fg=fg))
    return btn


def styled_button_limite(parent, text, command, variant="warn", small=False, **kwargs):
    styles = {
        "ok":  (C["ok_dim"],  C["ok"],  C["ok"],  C["bg"]),
        "warn":("#3d3520",    C["warn"], C["warn"], C["bg"]),
        "err": (C["err_dim"],  C["err"], C["err"], C["bg"]),
        "idle":(C["surface2"], C["ink_muted"], C["surface3"], C["ink"]),
    }
    bg, fg, abg, afg = styles.get(variant, styles["warn"])
    pad = (7, 3) if small else (13, 6)
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=fg, activebackground=abg, activeforeground=afg,
                    font=("Segoe UI", 8 if small else 9),
                    relief="flat", bd=0, padx=pad[0], pady=pad[1],
                    cursor="hand2", **kwargs)
    btn._limite_variant = variant
    btn._limite_bg = bg
    btn._limite_fg = fg
    btn._limite_abg = abg
    btn._limite_afg = afg
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg, fg=fg))
    return btn


def _set_limite_button(btn, text, variant):
    styles = {
        "ok":  (C["ok_dim"],  C["ok"],  C["ok"],  C["bg"]),
        "warn":("#3d3520",    C["warn"], C["warn"], C["bg"]),
        "err": (C["err_dim"],  C["err"], C["err"], C["bg"]),
        "idle":(C["surface2"], C["ink_muted"], C["surface3"], C["ink"]),
    }
    bg, fg, abg, afg = styles.get(variant, styles["warn"])
    btn.configure(text=text, bg=bg, fg=fg, activebackground=abg, activeforeground=afg)
    btn._limite_variant = variant
    btn._limite_bg = bg
    btn._limite_fg = fg
    btn._limite_abg = abg
    btn._limite_afg = afg
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg, fg=fg))


def styled_entry(parent, textvariable=None, width=20, show=None, **kwargs):
    return tk.Entry(parent, textvariable=textvariable, width=width,
                    show=show or "",
                    bg=C["surface2"], fg=C["ink"],
                    insertbackground=C["accent"],
                    relief="flat", highlightthickness=1,
                    highlightbackground=C["hair"],
                    highlightcolor=C["accent"],
                    font=("Segoe UI", 10), **kwargs)


def card_frame(parent, **kwargs):
    kw = {"bg": C["surface"], "highlightthickness": 1,
          "highlightbackground": C["hair"], "bd": 0}
    kw.update(kwargs)
    return tk.Frame(parent, **kw)


def eyebrow_label(parent, text, bg=None):
    bg = bg or C["bg"]
    return tk.Label(parent, text=text, bg=bg, fg=C["ink_faint"],
                    font=("Segoe UI", 7, "bold"))


def make_hub_option_card(parent, row, col, icon, title, sub, command, color,
                          alert=False, alert_text="● Atenção"):
    """Card de opção usado nas telas-hub (Operações Invertido, BPM, etc.),
    em grid de 3 colunas com design minimalista consistente com o app."""
    pad_x = {0: (0, 6), 1: (6, 6), 2: (6, 0)}.get(col, (6, 6))
    pad_y = (0, 0) if row == 0 else (12, 0)
    outer = tk.Frame(parent, bg=C["surface"],
                     highlightthickness=1, highlightbackground=C["hair"],
                     cursor="hand2")
    outer.grid(row=row, column=col, sticky="nsew", padx=pad_x, pady=pad_y)

    top_line = tk.Frame(outer, bg=C["hair"], height=2)
    top_line.pack(fill="x")

    body_f = tk.Frame(outer, bg=C["surface"], padx=22, pady=22)
    body_f.pack(fill="both", expand=True)

    icon_row = tk.Frame(body_f, bg=C["surface"])
    icon_row.pack(fill="x", anchor="w")
    icon_lbl = tk.Label(icon_row, text=icon, bg=C["surface"], fg=color,
                        font=("Segoe UI", 22))
    icon_lbl.pack(side="left")
    alert_lbl = None
    if alert:
        alert_lbl = tk.Label(icon_row, text=alert_text, bg=C["surface"],
                             fg=C["err"], font=("Segoe UI", 7, "bold"))
        alert_lbl.pack(side="right", anchor="n", pady=(4, 0))
    name_lbl = tk.Label(body_f, text=title, bg=C["surface"], fg=C["ink"],
                        font=("Segoe UI", 12, "bold"), anchor="w")
    name_lbl.pack(anchor="w", pady=(12, 4))
    sub_lbl = tk.Label(body_f, text=sub, bg=C["surface"], fg=C["ink_muted"],
                       font=("Segoe UI", 9), anchor="w", wraplength=200,
                       justify="left")
    sub_lbl.pack(anchor="w")
    arrow_lbl = tk.Label(body_f, text="Abrir →", bg=C["surface"], fg=C["ink_faint"],
                         font=("Segoe UI", 8, "bold"))
    arrow_lbl.pack(anchor="w", pady=(18, 0))

    widgets = [outer, top_line, body_f, icon_row, icon_lbl, name_lbl, sub_lbl, arrow_lbl]

    def _enter(_e=None):
        outer.configure(bg=C["surface2"], highlightbackground=color)
        top_line.configure(bg=color)
        for w in (body_f, icon_row, icon_lbl, name_lbl, sub_lbl):
            w.configure(bg=C["surface2"])
        if alert_lbl is not None:
            alert_lbl.configure(bg=C["surface2"])
        arrow_lbl.configure(bg=C["surface2"], fg=color)

    def _leave(_e=None):
        outer.configure(bg=C["surface"], highlightbackground=C["hair"])
        top_line.configure(bg=C["hair"])
        for w in (body_f, icon_row, icon_lbl, name_lbl, sub_lbl):
            w.configure(bg=C["surface"])
        if alert_lbl is not None:
            alert_lbl.configure(bg=C["surface"])
        arrow_lbl.configure(bg=C["surface"], fg=C["ink_faint"])

    for w in widgets:
        w.bind("<Button-1>", lambda _e: command())
        w.bind("<Enter>", _enter)
        w.bind("<Leave>", _leave)

    return outer


def section_divider(parent, text="", bg=None):
    bg = bg or C["bg"]
    row = tk.Frame(parent, bg=bg)
    if text:
        tk.Label(row, text=text, bg=bg, fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(side="left")
        spacer = tk.Frame(row, bg=C["hair"], height=1)
        spacer.pack(side="left", fill="x", expand=True, padx=(10, 0), pady=(5, 0))
    else:
        tk.Frame(row, bg=C["hair"], height=1).pack(fill="x")
    return row

class MinimalScrollbar(tk.Canvas):
    THUMB_MIN = 28

    def __init__(self, parent, command=None, bg=None, width=6, **kwargs):
        self._track_bg = bg or C["bg"]
        super().__init__(
            parent, width=width, highlightthickness=0, bd=0,
            bg=self._track_bg, cursor="arrow", **kwargs,
        )
        self._command = command
        self._first = 0.0
        self._last = 1.0
        self._thumb_fill = "#5c5c5c"
        self._thumb_hover = "#787878"
        self._drag_y = 0
        self._thumb_rect = None

        self.bind("<Configure>", self._redraw, add="+")
        self.bind("<Button-1>", self._on_press)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Enter>", lambda _e: self._paint_thumb(self._thumb_hover))
        self.bind("<Leave>", lambda _e: self._paint_thumb(self._thumb_fill))

    def set(self, first, last):
        f, l = float(first), float(last)
        if f == self._first and l == self._last:
            return
        self._first, self._last = f, l
        self._redraw()

    def _visible(self):
        return self._first > 0.001 or self._last < 0.999

    def _thumb_geometry(self):
        h = max(self.winfo_height(), 1)
        w = max(self.winfo_width(), 1)
        span = max(self._last - self._first, 0.001)
        thumb_h = max(self.THUMB_MIN, int(h * span))
        thumb_y = int(h * self._first)
        if thumb_y + thumb_h > h:
            thumb_y = max(0, h - thumb_h)
        return w, h, thumb_y, thumb_h

    def _redraw(self, _event=None):
        self.delete("all")
        self._thumb_rect = None
        if not self._visible():
            return
        w, _h, thumb_y, thumb_h = self._thumb_geometry()
        margin = 1
        x0, x1 = margin, max(margin + 2, w - margin)
        radius = (x1 - x0) / 2
        if thumb_h <= (x1 - x0):
            self.create_oval(x0, thumb_y, x1, thumb_y + (x1 - x0),
                             fill=self._thumb_fill, outline="", tags="thumb")
            self.create_oval(x0, thumb_y + thumb_h - (x1 - x0), x1, thumb_y + thumb_h,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self._thumb_rect = (x0, thumb_y, x1, thumb_y + thumb_h)
        else:
            mid_top = thumb_y + radius
            mid_bot = thumb_y + thumb_h - radius
            self.create_oval(x0, thumb_y, x1, mid_top + radius,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self.create_rectangle(x0, mid_top, x1, mid_bot,
                                  fill=self._thumb_fill, outline="", tags="thumb")
            self.create_oval(x0, mid_bot - radius, x1, thumb_y + thumb_h,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self._thumb_rect = (x0, thumb_y, x1, thumb_y + thumb_h)

    def _paint_thumb(self, color):
        for item in self.find_withtag("thumb"):
            self.itemconfig(item, fill=color)

    def _on_press(self, event):
        if not self._command or not self._visible():
            return
        w, h, thumb_y, thumb_h = self._thumb_geometry()
        if self._thumb_rect and self._thumb_rect[1] <= event.y <= self._thumb_rect[3]:
            self._drag_y = event.y - thumb_y
            return
        if event.y > thumb_y + thumb_h:
            self._command("scroll", 1, "pages")
        elif event.y < thumb_y:
            self._command("scroll", -1, "pages")

    def _on_drag(self, event):
        if not self._command or not self._visible():
            return
        w, h, _thumb_y, thumb_h = self._thumb_geometry()
        span = max(h - thumb_h, 1)
        frac = (event.y - self._drag_y) / span
        self._command("moveto", max(0.0, min(1.0, frac)))

    def _on_release(self, _event):
        self._drag_y = 0


def bind_text_mousewheel(text_widget):
    def _mw(event):
        try:
            if not text_widget.winfo_exists():
                return
        except tk.TclError:
            return
        if getattr(event, "delta", 0):
            text_widget.yview_scroll(int(-event.delta / 120), "units")
        elif event.num == 4:
            text_widget.yview_scroll(-3, "units")
        elif event.num == 5:
            text_widget.yview_scroll(3, "units")
    text_widget.bind("<MouseWheel>", _mw)
    text_widget.bind("<Button-4>", _mw)
    text_widget.bind("<Button-5>", _mw)


class ScrollableFrame(tk.Frame):
    def __init__(self, parent, bg=None, **kwargs):
        bg = bg or C["bg"]
        super().__init__(parent, bg=bg, **kwargs)
        self._bg = bg
        self._wheel_roots = []
        self._canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self._vbar = MinimalScrollbar(self, command=self._canvas.yview, bg=bg, width=6)
        self._canvas.configure(yscrollcommand=self._vbar.set)
        self._vbar.pack(side="right", fill="y", padx=(0, 2), pady=4)
        self._canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self._canvas, bg=bg)
        self._win = self._canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner)
        self._canvas.bind("<Configure>", self._on_canvas)
        self.bind("<Destroy>", self._on_destroy)
        self._wheel_roots = [self]
        self._scroll_job = None
        self.refresh_bindings()

    def _canvas_alive(self):
        try:
            return bool(self._canvas.winfo_exists())
        except tk.TclError:
            return False

    def _scroll_mousewheel(self, event):
        if not self._canvas_alive():
            return
        try:
            if getattr(event, "delta", 0):
                self._canvas.yview_scroll(int(-event.delta / 120), "units")
            elif event.num == 4:
                self._canvas.yview_scroll(-3, "units")
            elif event.num == 5:
                self._canvas.yview_scroll(3, "units")
        except tk.TclError:
            pass

    def _bind_mousewheel_tree(self, widget):
        if isinstance(widget, (tk.Text, tk.Listbox)):
            return
        widget.bind("<MouseWheel>", self._scroll_mousewheel)
        widget.bind("<Button-4>", self._scroll_mousewheel)
        widget.bind("<Button-5>", self._scroll_mousewheel)
        for child in widget.winfo_children():
            self._bind_mousewheel_tree(child)

    def refresh_bindings(self):
        if not self._canvas_alive():
            return
        for root in self._wheel_roots:
            try:
                if root.winfo_exists():
                    self._bind_mousewheel_tree(root)
            except tk.TclError:
                pass

    def link_wheel(self, container):
        if container not in self._wheel_roots:
            self._wheel_roots.append(container)
        self.refresh_bindings()

    def _on_inner(self, _event):
        if not self._canvas_alive():
            return
        if getattr(self, "_scroll_job", None):
            try:
                self.after_cancel(self._scroll_job)
            except Exception:
                pass
        self._scroll_job = self.after(16, self._sync_scrollregion)

    def _sync_scrollregion(self):
        self._scroll_job = None
        if not self._canvas_alive():
            return
        try:
            self._canvas.update_idletasks()
            self._canvas.configure(scrollregion=self._canvas.bbox("all"))
        except tk.TclError:
            pass

    def _on_canvas(self, event):
        if self._canvas_alive():
            self._canvas.itemconfigure(self._win, width=event.width)

    def _on_destroy(self, event):
        pass


class Sidebar(tk.Frame):
    NAV = [
        ("Home",              "⌂",  "Início"),
        ("Recebiveis",        "◈",  "Recebíveis"),
        ("Ligacoes",          "☎",  "Ligações"),
    ]

    def __init__(self, parent, controller, **kwargs):
        super().__init__(parent, bg=C["surface"], width=210, **kwargs)
        self.pack_propagate(False)
        self.controller = controller
        self._btns = {}
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=C["surface"])
        top.pack(fill="x", padx=18, pady=(14, 0))

        logo_row = tk.Frame(top, bg=C["surface"])
        logo_row.pack(fill="x")
        tk.Label(logo_row, text="Mesa", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Label(logo_row, text=" Itaú", bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Label(top, text="Large", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(3, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(18, 12))

        nav_outer = tk.Frame(self, bg=C["surface"])
        nav_outer.pack(fill="both", expand=True, anchor="n")

        for name, icon, label in self.NAV:
            row = tk.Frame(nav_outer, bg=C["surface"], cursor="hand2")
            row.pack(fill="x", pady=1, padx=6)

            bar = tk.Frame(row, bg=C["surface"], width=3)
            bar.pack(side="left", fill="y")

            inner = tk.Frame(row, bg=C["surface"], padx=8, pady=7)
            inner.pack(side="left", fill="x", expand=True)

            icon_lbl = tk.Label(inner, text=icon, bg=C["surface"],
                                fg=C["ink_faint"], font=("Segoe UI", 12), width=2, anchor="w")
            icon_lbl.pack(side="left")

            text_lbl = tk.Label(inner, text=label, bg=C["surface"],
                                fg=C["ink_muted"], font=("Segoe UI", 9), anchor="w")
            text_lbl.pack(side="left", padx=(6, 0))

            def _click(n=name): self.controller.show_frame(n)

            def _enter(e, r=row, inn=inner, il=icon_lbl, tl=text_lbl):
                active = getattr(self.controller, "_active_frame", None)
                n_     = ""
                for _n, _b in self._btns.items():
                    if _b["row"] is r: n_ = _n; break
                if n_ != active:
                    for w in (r, inn, il, tl):
                        try: w.configure(bg=C["surface2"])
                        except: pass

            def _leave(e, r=row, inn=inner, il=icon_lbl, tl=text_lbl):
                active = getattr(self.controller, "_active_frame", None)
                n_     = ""
                for _n, _b in self._btns.items():
                    if _b["row"] is r: n_ = _n; break
                if n_ != active:
                    for w in (r, inn, il, tl):
                        try: w.configure(bg=C["surface"])
                        except: pass

            for w in (row, inner, icon_lbl, text_lbl):
                w.bind("<Button-1>", lambda _, n=name: _click(n))
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)

            self._btns[name] = {
                "row": row, "inner": inner,
                "icon": icon_lbl, "text": text_lbl, "bar": bar
            }

    def set_active(self, name):
        for n, w in self._btns.items():
            is_active = (n == name)
            row_bg  = C["surface2"] if is_active else C["surface"]
            icon_fg = C["accent"]   if is_active else C["ink_faint"]
            text_fg = C["ink"]      if is_active else C["ink_muted"]
            bar_bg  = C["accent"]   if is_active else C["surface"]

            w["row"].configure(bg=row_bg)
            w["inner"].configure(bg=row_bg)
            w["icon"].configure(bg=row_bg, fg=icon_fg)
            w["text"].configure(bg=row_bg, fg=text_fg)
            w["bar"].configure(bg=bar_bg)



class MiniBarChart(tk.Canvas):
    """Gráfico de barras minimalista (Canvas puro), no estilo visual do app.
    Aceita uma ou duas séries sobrepostas (para Montante x Líquido)."""

    def __init__(self, parent, height=150, value_fmt=None, **kwargs):
        super().__init__(parent, height=height, bg=C["surface"],
                         highlightthickness=0, bd=0, **kwargs)
        self._labels = []
        self._series = []
        self._colors = []
        self._legend = []
        self._bar_hitboxes = []  # (x0, y0, x1, y1, label, valor, serie_idx)
        self._value_fmt = value_fmt or (lambda v: f"{v:g}")
        self._tooltip = None
        self.bind("<Configure>", lambda _e: self._redraw())
        self.bind("<Motion>", self._on_motion)
        self.bind("<Leave>", self._on_leave)

    def set_data(self, labels, series, colors, legend=None):
        """series: lista de listas de floats (mesma cardinalidade de labels)."""
        self._labels = labels
        self._series = series
        self._colors = colors
        self._legend = legend or []
        self._redraw()

    def _on_motion(self, e):
        for (x0, y0, x1, y1, label, valor, _si) in self._bar_hitboxes:
            if x0 <= e.x <= x1 and y0 <= e.y <= y1:
                self._show_tooltip(e.x_root, e.y_root, label, valor)
                return
        self._hide_tooltip()

    def _on_leave(self, _e=None):
        self._hide_tooltip()

    def _show_tooltip(self, x_root, y_root, label, valor):
        txt = f"{label}: {self._value_fmt(valor)}"
        if self._tooltip is None:
            self._tooltip = tk.Toplevel(self)
            self._tooltip.overrideredirect(True)
            try:
                self._tooltip.attributes("-topmost", True)
            except Exception:
                pass
            self._tt_lbl = tk.Label(
                self._tooltip, text=txt, bg=C["ink"], fg=C["surface"],
                font=("Segoe UI", 8), padx=6, pady=3)
            self._tt_lbl.pack()
        else:
            self._tt_lbl.configure(text=txt)
        self._tooltip.geometry(f"+{x_root + 12}+{y_root + 14}")
        self._tooltip.deiconify()

    def _hide_tooltip(self):
        if self._tooltip is not None:
            self._tooltip.withdraw()

    def destroy(self):
        if self._tooltip is not None:
            try:
                self._tooltip.destroy()
            except Exception:
                pass
        super().destroy()

    def _redraw(self, _e=None):
        self.delete("all")
        self._bar_hitboxes = []
        w = max(self.winfo_width(), 1)
        h = max(self.winfo_height(), 1)
        if not self._labels or not self._series or not self._series[0]:
            self.create_text(w / 2, h / 2, text="Sem dados no período",
                             fill=C["ink_faint"], font=("Segoe UI", 9))
            return
        n = len(self._labels)
        top_pad, bottom_pad = 14, 22
        chart_h = max(h - top_pad - bottom_pad, 10)
        vmax = max((max(s) for s in self._series if s), default=0) or 1
        left_pad = 8
        avail_w = w - left_pad * 2
        group_w = avail_w / max(n, 1)
        n_series = len(self._series)
        bar_w = max((group_w * 0.6) / max(n_series, 1), 2)

        # legenda
        if self._legend:
            lx = w - 10
            for i, (txt, color) in enumerate(reversed(list(zip(self._legend, self._colors)))):
                lx -= (len(txt) * 6 + 26)
            lx = 10
            for txt, color in zip(self._legend, self._colors):
                self.create_oval(lx, 2, lx + 8, 10, fill=color, outline="")
                self.create_text(lx + 12, 6, text=txt, fill=C["ink_muted"],
                                 font=("Segoe UI", 7), anchor="w")
                lx += len(txt) * 6 + 26

        for i, label in enumerate(self._labels):
            gx0 = left_pad + i * group_w
            for si, serie in enumerate(self._series):
                val = serie[i] if i < len(serie) else 0
                bar_h = (val / vmax) * chart_h
                bx0 = gx0 + si * bar_w + (group_w - n_series * bar_w) / 2
                y1 = top_pad + chart_h
                y0 = y1 - bar_h
                color = self._colors[si % len(self._colors)]
                self.create_rectangle(bx0, y0, bx0 + bar_w - 2, y1,
                                      fill=color, outline="")
                # área de detecção do mouse: usa a barra toda, mas garante
                # altura mínima (para valores 0 ainda serem "hover-áveis")
                hit_y0 = min(y0, y1 - 6)
                self._bar_hitboxes.append(
                    (bx0, hit_y0, bx0 + bar_w - 2, y1, label, val, si))
            # rótulo do eixo x — só mostra a cada N para não poluir
            step = max(1, n // 8)
            if i % step == 0 or i == n - 1:
                self.create_text(gx0 + group_w / 2, top_pad + chart_h + 11,
                                 text=label, fill=C["ink_faint"],
                                 font=("Segoe UI", 7), anchor="n")



def _current_username() -> str:
    try:
        return os.environ.get("USERNAME") or os.environ.get("USER") or ""
    except Exception:
        return ""


PIPE_MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
              "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]


def _pipe_month_range(year, month):
    ini = date(year, month, 1)
    if month == 12:
        fim = date(year, 12, 31)
    else:
        fim = date(year, month + 1, 1) - timedelta(days=1)
    return ini, fim


def _pipe_week_range(ref):
    ini = ref - timedelta(days=ref.weekday())  # segunda-feira
    fim = ini + timedelta(days=6)
    return ini, fim


def _pipe_decimal_to_float(raw):
    s = (raw or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pipe_toggle_button(parent, text, command, small=False):
    """Botão estilo toggle (aba / seleção) cujo hover respeita o estado
    selecionado, ao contrário de styled_button que sempre volta pra cor
    'não selecionado' ao tirar o mouse."""
    pad = (7, 3) if small else (13, 6)
    btn = tk.Button(parent, text=text, command=command,
                     bg=C["surface2"], fg=C["ink_muted"],
                     activebackground=C["accent"], activeforeground=C["bg"],
                     font=("Segoe UI", 8 if small else 9),
                     relief="flat", bd=0, padx=pad[0], pady=pad[1], cursor="hand2")
    btn._pipe_selected = False

    def on_enter(_):
        if not btn._pipe_selected:
            btn.configure(bg=C["surface3"], fg=C["ink"])

    def on_leave(_):
        if not btn._pipe_selected:
            btn.configure(bg=C["surface2"], fg=C["ink_muted"])

    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    return btn


def _pipe_set_toggle(btn, selected):
    btn._pipe_selected = selected
    if selected:
        btn.configure(bg=C["accent"], fg=C["bg"])
    else:
        btn.configure(bg=C["surface2"], fg=C["ink_muted"])



# ─── Ligações (contador diário de ligações) ────────────────────────────────
# Banco próprio (ligacoes.db), na mesma pasta de rede dos demais módulos —
# não interfere em nada do Braskem/Taxas Pré/Histórico.

LIGACOES_DB_PATH = os.path.join(SHARED_APP_BASE_DIR, "ligacoes.db")


class LigacoesData:
    _instance = None
    RETRY_SECONDS = 30

    @classmethod
    def get(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def __init__(self):
        self._available = False
        self._retry_timer = None
        self._ensure_schema()
        if not self._available:
            self._schedule_retry()

    def is_available(self):
        return self._available

    def _schedule_retry(self):
        if self._retry_timer is None:
            self._retry_timer = threading.Timer(self.RETRY_SECONDS, self._retry_tick)
            self._retry_timer.daemon = True
            self._retry_timer.start()

    def _retry_tick(self):
        self._retry_timer = None
        self._ensure_schema()
        if not self._available:
            self._schedule_retry()

    def _connect(self):
        conn = sqlite3.connect(LIGACOES_DB_PATH, timeout=20)
        conn.execute("PRAGMA journal_mode=DELETE")
        return conn

    def _ensure_schema(self):
        try:
            if not os.path.isdir(os.path.dirname(LIGACOES_DB_PATH)):
                self._available = False
                return
            conn = self._connect()
            try:
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS ligacoes_registros (
                        data TEXT PRIMARY KEY,
                        quantidade INTEGER NOT NULL DEFAULT 0,
                        updated_at TEXT NOT NULL,
                        username TEXT
                    )
                """)
                conn.commit()
            finally:
                conn.close()
            self._available = True
        except Exception as e:
            print(f"[ligacoes] _ensure_schema falhou: {e}", file=sys.stderr)
            self._available = False

    def adicionar(self, data_iso, delta):
        if not self._available or delta == 0:
            return False
        try:
            conn = self._connect()
            try:
                conn.execute("""
                    INSERT INTO ligacoes_registros (data, quantidade, updated_at, username)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(data) DO UPDATE SET
                        quantidade = MAX(quantidade + ?, 0),
                        updated_at = excluded.updated_at
                """, (data_iso, max(delta, 0), datetime.now().isoformat(timespec="seconds"),
                      _current_username(), delta))
                conn.commit()
            finally:
                conn.close()
            return True
        except Exception as e:
            print(f"[ligacoes] adicionar falhou: {e}", file=sys.stderr)
            return False

    def total_dia(self, data_iso):
        if not self._available:
            return 0
        try:
            conn = self._connect()
            try:
                cur = conn.execute(
                    "SELECT quantidade FROM ligacoes_registros WHERE data=?", (data_iso,))
                row = cur.fetchone()
                return row[0] if row else 0
            finally:
                conn.close()
        except Exception:
            return 0

    def serie(self, dt_ini, dt_fim):
        """Lista [(data_iso, quantidade), ...] com todos os dias do intervalo,
        preenchendo com 0 os dias sem registro."""
        valores = {}
        if self._available:
            try:
                conn = self._connect()
                try:
                    cur = conn.execute(
                        "SELECT data, quantidade FROM ligacoes_registros "
                        "WHERE data BETWEEN ? AND ?", (dt_ini.isoformat(), dt_fim.isoformat()))
                    valores = dict(cur.fetchall())
                finally:
                    conn.close()
            except Exception as e:
                print(f"[ligacoes] serie falhou: {e}", file=sys.stderr)
        out, d = [], dt_ini
        while d <= dt_fim:
            out.append((d.isoformat(), valores.get(d.isoformat(), 0)))
            d += timedelta(days=1)
        return out

    def total_periodo(self, dt_ini, dt_fim):
        return sum(q for _d, q in self.serie(dt_ini, dt_fim))


class LigacoesFrame(tk.Frame):
    """Contador diário de ligações — botão rápido (+1), inclusão manual de
    quantidade por dia, histórico com gráfico (semanal/mensal) e indicador
    de crescimento frente ao período equivalente anterior."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._data = LigacoesData.get()
        self._periodo = "semanal"
        self._ref_data = date.today()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24, 0))
        tk.Label(hdr, text="Ligações", bg=C["bg"], fg=C["ink"],
                 font=("Georgia", 18, "bold")).pack(side="left")
        styled_button(hdr, "← Voltar",
                      lambda: self.controller.show_frame("Home")).pack(side="right")
        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(14, 0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        body = self._sf.inner
        body.configure(bg=C["bg"])
        wrap = tk.Frame(body, bg=C["bg"])
        wrap.pack(fill="both", expand=True, padx=32, pady=(20, 24))

        contador = card_frame(wrap)
        contador.pack(fill="x")
        cbody = tk.Frame(contador, bg=C["surface"], padx=24, pady=20)
        cbody.pack(fill="x")
        esquerda = tk.Frame(cbody, bg=C["surface"])
        esquerda.pack(side="left")
        tk.Label(esquerda, text="HOJE", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        self._total_hoje_lbl = tk.Label(esquerda, text="0", bg=C["surface"],
                                         fg=C["ink"], font=("Segoe UI", 26, "bold"))
        self._total_hoje_lbl.pack(anchor="w")
        styled_button(cbody, "+1  Registrar ligação", self._registrar_uma,
                      accent=True).pack(side="right")

        manual = card_frame(wrap)
        manual.pack(fill="x", pady=(12, 0))
        mbody = tk.Frame(manual, bg=C["surface"], padx=24, pady=16)
        mbody.pack(fill="x")
        tk.Label(mbody, text="Incluir quantidade em um dia", bg=C["surface"],
                 fg=C["ink"], font=("Segoe UI", 9, "bold")).pack(anchor="w")
        row = tk.Frame(mbody, bg=C["surface"])
        row.pack(fill="x", pady=(10, 0))
        tk.Label(row, text="Data", bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))
        self._var_data_manual = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        styled_entry(row, textvariable=self._var_data_manual, width=12).pack(
            side="left", padx=(0, 20))
        tk.Label(row, text="Quantidade", bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))
        self._var_qtd_manual = tk.StringVar(value="")
        ent_qtd = styled_entry(row, textvariable=self._var_qtd_manual, width=8)
        ent_qtd.pack(side="left", padx=(0, 20))

        def only_digits_key(e):
            if e.keysym in {"Tab", "Left", "Right", "Home", "End", "BackSpace", "Delete"}:
                return
            if e.char and e.char.isdigit():
                return
            return "break"
        ent_qtd.bind("<Key>", only_digits_key)
        styled_button(row, "Adicionar", self._adicionar_manual, accent=True).pack(side="left")
        self._msg_lbl = tk.Label(mbody, text="", bg=C["surface"], fg=C["ink_muted"],
                                  font=("Segoe UI", 8))
        self._msg_lbl.pack(anchor="w", pady=(8, 0))

        filt = tk.Frame(wrap, bg=C["bg"])
        filt.pack(fill="x", pady=(20, 8))
        tk.Label(filt, text="Histórico", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self._btn_semanal = _pipe_toggle_button(filt, "Semanal",
                                                 lambda: self._set_periodo("semanal"), small=True)
        self._btn_semanal.pack(side="left", padx=(20, 0))
        self._btn_mensal = _pipe_toggle_button(filt, "Mensal",
                                                lambda: self._set_periodo("mensal"), small=True)
        self._btn_mensal.pack(side="left", padx=(6, 0))
        styled_button(filt, "◀", lambda: self._mudar_ref(-1), small=True).pack(
            side="left", padx=(20, 0))
        self._ref_lbl = tk.Label(filt, text="", bg=C["bg"], fg=C["ink_muted"],
                                  font=("Segoe UI", 9))
        self._ref_lbl.pack(side="left", padx=6)
        styled_button(filt, "▶", lambda: self._mudar_ref(1), small=True).pack(side="left")

        graf = card_frame(wrap)
        graf.pack(fill="x")
        gbody = tk.Frame(graf, bg=C["surface"], padx=16, pady=16)
        gbody.pack(fill="both", expand=True)
        self._chart = MiniBarChart(
            gbody, height=180,
            value_fmt=lambda v: f"{int(v)} ligaç{'ão' if int(v) == 1 else 'ões'}")
        self._chart.pack(fill="x")

        self._resumo_wrap = tk.Frame(wrap, bg=C["bg"])
        self._resumo_wrap.pack(fill="x", pady=(12, 0))

        self._set_periodo("semanal")

    def on_show(self):
        self._sf.refresh_bindings()
        self._total_hoje_lbl.configure(
            text=str(self._data.total_dia(date.today().isoformat())))
        self._refresh_historico()

    def _registrar_uma(self):
        self._data.adicionar(date.today().isoformat(), 1)
        self._total_hoje_lbl.configure(
            text=str(self._data.total_dia(date.today().isoformat())))
        self._refresh_historico()

    def _adicionar_manual(self):
        try:
            d = datetime.strptime(self._var_data_manual.get().strip(), "%d/%m/%Y").date()
        except ValueError:
            self._msg_lbl.configure(text="Data inválida (use dd/mm/aaaa).", fg=C["err"])
            return
        try:
            qtd = int(self._var_qtd_manual.get())
        except ValueError:
            self._msg_lbl.configure(text="Informe uma quantidade válida.", fg=C["err"])
            return
        if qtd <= 0:
            self._msg_lbl.configure(text="Quantidade deve ser maior que zero.", fg=C["err"])
            return
        self._data.adicionar(d.isoformat(), qtd)
        self._msg_lbl.configure(
            text=f"{qtd} ligação(ões) adicionada(s) em {d.strftime('%d/%m/%Y')}.", fg=C["ok"])
        self._var_qtd_manual.set("")
        if d == date.today():
            self._total_hoje_lbl.configure(
                text=str(self._data.total_dia(date.today().isoformat())))
        self._refresh_historico()

    def _set_periodo(self, p):
        self._periodo = p
        _pipe_set_toggle(self._btn_semanal, p == "semanal")
        _pipe_set_toggle(self._btn_mensal, p == "mensal")
        self._refresh_historico()

    def _mudar_ref(self, delta):
        if self._periodo == "semanal":
            self._ref_data += timedelta(weeks=delta)
        else:
            mes, ano = self._ref_data.month, self._ref_data.year
            mes += delta
            while mes > 12:
                mes -= 12; ano += 1
            while mes < 1:
                mes += 12; ano -= 1
            dia = min(self._ref_data.day, 28)
            self._ref_data = date(ano, mes, dia)
        self._refresh_historico()

    def _intervalo_atual(self):
        if self._periodo == "semanal":
            ini, fim = _pipe_week_range(self._ref_data)
            ini_ant, fim_ant = ini - timedelta(days=7), fim - timedelta(days=7)
            label = f"{ini.strftime('%d/%m')} a {fim.strftime('%d/%m/%Y')}"
        else:
            ini, fim = _pipe_month_range(self._ref_data.year, self._ref_data.month)
            mes_ant, ano_ant = self._ref_data.month - 1, self._ref_data.year
            if mes_ant < 1:
                mes_ant, ano_ant = 12, ano_ant - 1
            ini_ant, fim_ant = _pipe_month_range(ano_ant, mes_ant)
            label = f"{PIPE_MESES[self._ref_data.month - 1]}/{self._ref_data.year}"
        return ini, fim, ini_ant, fim_ant, label

    def _refresh_historico(self):
        ini, fim, ini_ant, fim_ant, label = self._intervalo_atual()
        self._ref_lbl.configure(text=label)

        serie_atual = self._data.serie(ini, fim)
        labels = [datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m") for d, _q in serie_atual]
        valores = [q for _d, q in serie_atual]
        self._chart.set_data(labels, [valores], [C["accent"]])

        total_atual = sum(valores)
        total_anterior = self._data.total_periodo(ini_ant, fim_ant)
        # Média por dia considera apenas dias úteis (seg-sex); finais de
        # semana continuam aparecendo no gráfico, mas não entram na média.
        valores_uteis = [
            q for d_iso, q in serie_atual
            if datetime.strptime(d_iso, "%Y-%m-%d").weekday() < 5
        ]
        media_dia = sum(valores_uteis) / max(len(valores_uteis), 1)

        for w in self._resumo_wrap.winfo_children():
            w.destroy()
        grid = tk.Frame(self._resumo_wrap, bg=C["bg"])
        grid.pack(fill="x")
        for c in range(3):
            grid.columnconfigure(c, weight=1, uniform="lig")

        def card(c, titulo, valor_txt, cor=None):
            outer = tk.Frame(grid, bg=C["bg"])
            outer.grid(row=0, column=c, sticky="nsew", padx=5, pady=5)
            cf = card_frame(outer); cf.pack(fill="both", expand=True)
            b = tk.Frame(cf, bg=C["surface"], padx=16, pady=14); b.pack(fill="both", expand=True)
            tk.Label(b, text=titulo, bg=C["surface"], fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold")).pack(anchor="w")
            tk.Label(b, text=valor_txt, bg=C["surface"], fg=cor or C["ink"],
                     font=("Segoe UI", 15, "bold")).pack(anchor="w", pady=(4, 0))

        card(0, "TOTAL NO PERÍODO", str(total_atual))
        card(1, "MÉDIA POR DIA", f"{media_dia:.1f}")

        if total_anterior == 0:
            cresc_txt, cor = ("—" if total_atual == 0 else "novo"), C["ink_muted"]
        else:
            cresc = (total_atual - total_anterior) / total_anterior * 100
            cor = C["ok"] if cresc >= 0 else C["err"]
            seta = "▲" if cresc >= 0 else "▼"
            cresc_txt = f"{seta} {abs(cresc):.1f}%"
        card(2, "VS. PERÍODO ANTERIOR", cresc_txt, cor)


# ─── Recebíveis (cessão de cartão — Visa/Master A/D e Adiq) ────────────────
# Primeira etapa desse fluxo novo: reconhecer o arquivo que a adquirente
# manda (pelo nome — bandeira + letra — e confirmar pelas colunas),
# organizar as notas e calcular com o Preço/Funding que o usuário digita.
# O racional do cálculo (taxa ao mês -> desconto linear -> líquido) é o
# mesmo já usado no Braskem/Invertido — ainda não validado pra esse
# produto especificamente; o usuário vai conferir e corrigir.

def _pascoa_br(ano: int) -> date:
    """Data da Páscoa (algoritmo de Gauss/Meeus) — base pros feriados
    móveis (Carnaval, Sexta-feira Santa, Corpus Christi)."""
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(ano, mes, dia)


def _feriados_nacionais_br(ano: int) -> set:
    """Feriados nacionais do Brasil — fixos + móveis (a partir da
    Páscoa). Inclui Carnaval e Corpus Christi (não são feriado nacional
    por lei em todo lugar, mas não têm expediente bancário)."""
    pascoa = _pascoa_br(ano)
    return {
        date(ano, 1, 1), pascoa - timedelta(days=48), pascoa - timedelta(days=47),
        pascoa - timedelta(days=2), pascoa + timedelta(days=60),
        date(ano, 4, 21), date(ano, 5, 1), date(ano, 9, 7), date(ano, 10, 12),
        date(ano, 11, 2), date(ano, 11, 15), date(ano, 11, 20), date(ano, 12, 25),
    }


def _eh_dia_util_br(d: date) -> bool:
    if d.weekday() >= 5:
        return False
    return d not in _feriados_nacionais_br(d.year)


def _proximo_dia_util_br(d: date) -> date:
    while not _eh_dia_util_br(d):
        d += timedelta(days=1)
    return d


def _fmt_brl(v: Decimal) -> str:
    s = f"{v:,.2f}".replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {s}"


def _fmt_cnpj_livre(v) -> str:
    d = re.sub(r"\D", "", str(v or ""))
    if len(d) != 14:
        return str(v or "—")
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


# Cabeçalhos esperados de cada formato — usados pra CONFIRMAR o palpite
# feito pelo nome do arquivo (nome pode estar errado/renomeado; colunas
# erradas são motivo de alerta, não de recusa silenciosa).
RECEBIVEIS_HEADERS_VISA_MASTER = [
    "DATA_PROCESSAMENTO", "ID_EMISSOR", "RAZAO_EMISSOR",
    "DATA_LIQUIDACAO", "DISPONIVEL_DIA", "CNPJ",
]
RECEBIVEIS_HEADERS_ADIQ = [
    "Data Liquidacao", "ICA", "Emissor", "CNPJ",
    "Valor Cessionado", "Valor Cotação", "Taxa",
]


def _recebiveis_detectar_por_nome(caminho: str):
    """Bandeira + tipo a partir do NOME do arquivo — é o único lugar
    onde essa informação aparece (as colunas de Visa A/D e Master A/D
    são idênticas entre si). Retorna (bandeira, tipo) ou (None, None)
    se não conseguir reconhecer — nesse caso a tela pede confirmação
    manual em vez de adivinhar errado."""
    nome = os.path.basename(caminho).lower()
    if "adiq" in nome:
        return ("Adiq", None)
    m = re.search(r"(visa|master)[\s\-_]*([ad])(?![a-z0-9])", nome)
    if m:
        bandeira = "Visa" if m.group(1) == "visa" else "Master"
        return (bandeira, m.group(2).upper())
    return (None, None)


def _recebiveis_ler_planilha(caminho: str):
    """Lê o arquivo inteiro (não só as 1-2 linhas de exemplo — no dia a
    dia vem muito mais nota) e devolve (bandeira, tipo, notas, avisos).
    'notas' é uma lista de dicts já normalizados; 'avisos' lista
    inconsistências (cabeçalho não bate com o nome, linha incompleta
    etc.) pra mostrar na tela em vez de falhar calado."""
    avisos = []
    if openpyxl is None:
        return (None, None, [], ["Suporte a .xlsx (openpyxl) não está disponível neste ambiente."])

    bandeira_nome, tipo_nome = _recebiveis_detectar_por_nome(caminho)

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value is not None else ""
              for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def _bate(esperado):
        return [h.strip().lower() for h in header[:len(esperado)]] == \
               [h.strip().lower() for h in esperado]

    if _bate(RECEBIVEIS_HEADERS_ADIQ):
        formato = "adiq"
        if bandeira_nome and bandeira_nome != "Adiq":
            avisos.append(f"Nome do arquivo sugere {bandeira_nome}, mas as colunas são do "
                           f"formato Adiq — segui pelas colunas.")
        bandeira, tipo = "Adiq", None
    elif _bate(RECEBIVEIS_HEADERS_VISA_MASTER):
        formato = "visa_master"
        if not bandeira_nome:
            avisos.append("Não consegui identificar bandeira/tipo pelo nome do arquivo "
                           "(esperado algo como 'visaA...', 'masterD...') — confirme manualmente.")
        bandeira, tipo = bandeira_nome, tipo_nome
    else:
        return (bandeira_nome, tipo_nome, [],
                [f"As colunas do arquivo não batem com nenhum formato conhecido "
                 f"(cabeçalho encontrado: {', '.join(h for h in header if h)})."])

    notas = []
    linhas_vazias = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            if formato == "adiq":
                data_liq, ica, emissor, cnpj, valor = row[0], row[1], row[2], row[3], row[4]
                identificador = str(ica) if ica is not None else ""
            else:
                _dt_proc, id_emissor, razao, data_liq, valor, cnpj = row[:6]
                emissor = razao
                identificador = str(id_emissor) if id_emissor is not None else ""

            if isinstance(data_liq, datetime):
                data_liq = data_liq.date()
            if not isinstance(data_liq, date) or valor in (None, ""):
                linhas_vazias += 1
                continue

            notas.append({
                "identificador": identificador,
                "razao": str(emissor or "").strip(),
                "cnpj": str(cnpj or "").strip(),
                "data_liquidacao": data_liq,
                "disponivel": _valor_livre_to_decimal(valor),
                "bandeira": bandeira or "—",
                "tipo": tipo,
                "arquivo_origem": os.path.basename(caminho),
            })
        except Exception:
            linhas_vazias += 1
            continue

    if linhas_vazias:
        avisos.append(f"{linhas_vazias} linha(s) ignorada(s) por estarem incompletas "
                       f"(sem data de liquidação ou valor).")
    if not notas:
        avisos.append("Nenhuma nota válida encontrada no arquivo.")

    return (bandeira, tipo, notas, avisos)


def _valor_livre_to_decimal(v) -> Decimal:
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(v, (int, float)):
        return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = re.sub(r"[^\d,.\-]", "", str(v or "")).strip()
    if not s:
        return Decimal("0")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return Decimal("0")


def calcular_cessao_recebivel(disponivel: Decimal, data_liquidacao: date,
                               preco_pct: Decimal, funding_pct: Decimal,
                               hoje=None) -> dict:
    """Motor de cálculo — AINDA NÃO VALIDADO pelo usuário pra esse
    produto. Segue o mesmo racional (desconto linear sobre taxa ao mês)
    já usado e conferido no Braskem/Invertido: taxa/30 x prazo. 'Preço'
    é a taxa ao mês cobrada do cliente; 'Funding' é o custo de captação
    no mesmo formato — a diferença entre os dois é o ganho da mesa.

    data_liquidacao é ajustada pro próximo dia útil (mesma regra do
    Braskem: fim de semana/feriado nacional empurra pra frente) antes de
    contar o prazo."""
    hoje = hoje or date.today()
    venc_ajustado = _proximo_dia_util_br(data_liquidacao)
    prazo_dias = max((venc_ajustado - hoje).days, 0)

    taxa_mes = preco_pct / Decimal("100")
    fator = Decimal("1") - (taxa_mes * Decimal(prazo_dias) / Decimal("30"))
    valor_liquido = (disponivel * fator).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    taxa_ref_360 = (preco_pct * Decimal("12")).quantize(Decimal("0.0001"))
    custo_caixa_periodo = (disponivel * (funding_pct / Decimal("100")) *
                            Decimal(prazo_dias) / Decimal("30")).quantize(
                                Decimal("0.01"), rounding=ROUND_HALF_UP)
    ganho_mesa_pct = (preco_pct - funding_pct).quantize(Decimal("0.0001"))

    return {
        "data_liquidacao_original": data_liquidacao,
        "data_liquidacao_ajustada": venc_ajustado,
        "prazo_dias": prazo_dias,
        "taxa_mes": preco_pct.quantize(Decimal("0.0001")),
        "taxa_ref_360": taxa_ref_360,
        "funding_pct": funding_pct.quantize(Decimal("0.0001")),
        "custo_caixa": custo_caixa_periodo,
        "ganho_mesa_pct": ganho_mesa_pct,
        "valor_disponivel": disponivel,
        "valor_liquido": valor_liquido,
    }


class RecebiveisFrame(tk.Frame):
    """Primeira etapa do fluxo de cessão de cartão (Visa/Master A/D,
    Adiq): anexar um ou mais arquivos, reconhecer automaticamente pelo
    nome + conferir pelas colunas, digitar Preço/Funding e calcular.
    Sem persistência ainda (histórico/e-mail entram depois) — é só o
    início do processo, conforme combinado."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._arquivos = []   # [{path, bandeira, tipo, notas, avisos}]
        self._resultados = None
        self._build()

    def _build(self):
        sf = ScrollableFrame(self)
        sf.pack(fill="both", expand=True)
        pad = tk.Frame(sf.inner, bg=C["bg"])
        pad.pack(fill="both", expand=True, padx=44, pady=36)

        eyebrow_label(pad, "RECEBÍVEIS · CESSÃO DE CARTÃO").pack(anchor="w")
        tk.Label(pad, text="Anexar planilhas", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 20, "bold"), anchor="w").pack(anchor="w", pady=(4, 2))
        tk.Label(pad, text="Visa A/D, Master A/D ou Adiq — o app reconhece pelo nome do "
                            "arquivo e confere pelas colunas.",
                 bg=C["bg"], fg=C["ink_muted"], font=("Segoe UI", 9),
                 anchor="w", wraplength=760, justify="left").pack(anchor="w", pady=(0, 20))

        btn_row = tk.Frame(pad, bg=C["bg"])
        btn_row.pack(fill="x", pady=(0, 16))
        styled_button(btn_row, "Anexar arquivo(s)", self._anexar, accent=True).pack(side="left")
        self._calc_btn = styled_button(btn_row, "Calcular", self._abrir_parametros)
        self._calc_btn.pack(side="left", padx=(8, 0))
        self._calc_btn.configure(state="disabled")

        self._lista_wrap = tk.Frame(pad, bg=C["bg"])
        self._lista_wrap.pack(fill="x", pady=(0, 24))

        self._resultado_wrap = tk.Frame(pad, bg=C["bg"])
        self._resultado_wrap.pack(fill="both", expand=True)

        self._render_lista()

    # ── anexar / reconhecer ────────────────────────────────────────────
    def _anexar(self):
        caminhos = filedialog.askopenfilenames(
            title="Anexar planilhas de recebíveis",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if not caminhos:
            return
        for caminho in caminhos:
            bandeira, tipo, notas, avisos = _recebiveis_ler_planilha(caminho)
            self._arquivos.append({
                "path": caminho, "bandeira": bandeira, "tipo": tipo,
                "notas": notas, "avisos": avisos,
            })
        self._resultados = None
        self._render_lista()

    def _remover_arquivo(self, idx):
        del self._arquivos[idx]
        self._resultados = None
        self._render_lista()

    def _render_lista(self):
        for w in self._lista_wrap.winfo_children():
            w.destroy()
        for w in self._resultado_wrap.winfo_children():
            w.destroy()

        if not self._arquivos:
            tk.Label(self._lista_wrap, text="Nenhum arquivo anexado ainda.",
                     bg=C["bg"], fg=C["ink_faint"], font=("Segoe UI", 9)).pack(anchor="w")
            self._calc_btn.configure(state="disabled")
            return

        total_notas = 0
        for idx, arq in enumerate(self._arquivos):
            card = card_frame(self._lista_wrap)
            card.pack(fill="x", pady=(0, 8))
            body = tk.Frame(card, bg=C["surface"], padx=14, pady=12)
            body.pack(fill="both", expand=True)

            top = tk.Frame(body, bg=C["surface"])
            top.pack(fill="x")
            nome = os.path.basename(arq["path"])
            qtd = len(arq["notas"])
            total_notas += qtd
            if arq["bandeira"] and (arq["bandeira"] == "Adiq" or arq["tipo"]):
                tag = arq["bandeira"] if arq["bandeira"] == "Adiq" else f'{arq["bandeira"]} {arq["tipo"]}'
                cor_tag = C["ok"]
            else:
                tag = "não identificado"
                cor_tag = C["err"]
            tk.Label(top, text=nome, bg=C["surface"], fg=C["ink"],
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")
            tk.Label(top, text=f"  ·  {tag}", bg=C["surface"], fg=cor_tag,
                     font=("Segoe UI", 9, "bold")).pack(side="left")
            tk.Button(top, text="remover", command=lambda i=idx: self._remover_arquivo(i),
                      bg=C["surface"], fg=C["ink_faint"], activebackground=C["surface"],
                      activeforeground=C["err"], relief="flat", bd=0, cursor="hand2",
                      font=("Segoe UI", 8)).pack(side="right")

            sub_txt = f"{qtd} nota(s) reconhecida(s)"
            tk.Label(body, text=sub_txt, bg=C["surface"], fg=C["ink_muted"],
                     font=("Segoe UI", 8)).pack(anchor="w", pady=(4, 0))

            for aviso in arq["avisos"]:
                tk.Label(body, text=f"⚠ {aviso}", bg=C["surface"], fg=C["warn"],
                         font=("Segoe UI", 8), wraplength=680, justify="left",
                         anchor="w").pack(anchor="w", pady=(4, 0))

        self._calc_btn.configure(state="normal" if total_notas else "disabled")

    # ── parâmetros (preço / funding) e cálculo ─────────────────────────
    def _abrir_parametros(self):
        todas_notas = [n for arq in self._arquivos for n in arq["notas"]]
        if not todas_notas:
            return

        dlg = tk.Toplevel(self)
        dlg.title("Calcular cessão")
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.attributes("-topmost", True)

        pad = tk.Frame(dlg, bg=C["surface"], padx=20, pady=18)
        pad.pack(fill="both", expand=True)

        tk.Label(pad, text="Preço e funding", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(pad, text=f"Aplicado às {len(todas_notas)} nota(s) anexada(s) — "
                            "taxa ao mês (%), mesmo padrão do Braskem/Invertido.",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 8),
                 wraplength=340, justify="left").pack(anchor="w", pady=(2, 14))

        def _campo(label_txt):
            tk.Label(pad, text=label_txt, bg=C["surface"], fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold")).pack(anchor="w")
            var = tk.StringVar()
            styled_entry(pad, textvariable=var, width=20).pack(anchor="w", pady=(4, 12))
            return var

        preco_var = _campo("PREÇO — TAXA AO CLIENTE (% A.M.)")
        funding_var = _campo("FUNDING — CUSTO DE CAPTAÇÃO (% A.M.)")

        err_lbl = tk.Label(pad, text="", bg=C["surface"], fg=C["err"], font=("Segoe UI", 8))
        err_lbl.pack(anchor="w")

        def _confirmar():
            try:
                preco = Decimal(preco_var.get().strip().replace(",", "."))
                funding = Decimal(funding_var.get().strip().replace(",", "."))
            except Exception:
                err_lbl.configure(text="Informe Preço e Funding como número (ex.: 1,35).")
                return
            dlg.destroy()
            self._calcular(preco, funding)

        btn_row = tk.Frame(pad, bg=C["surface"])
        btn_row.pack(fill="x", pady=(6, 0))
        styled_button(btn_row, "Cancelar", dlg.destroy).pack(side="left")
        styled_button(btn_row, "Calcular", _confirmar, accent=True).pack(side="right")

    def _calcular(self, preco_pct, funding_pct):
        todas_notas = [n for arq in self._arquivos for n in arq["notas"]]
        hoje = date.today()
        resultados = []
        for n in todas_notas:
            calc = calcular_cessao_recebivel(
                n["disponivel"], n["data_liquidacao"], preco_pct, funding_pct, hoje)
            resultados.append({**n, **calc})
        self._resultados = resultados
        self._render_resultado(preco_pct, funding_pct)

    def _render_resultado(self, preco_pct, funding_pct):
        for w in self._resultado_wrap.winfo_children():
            w.destroy()
        if not self._resultados:
            return

        make_hairline(self._resultado_wrap, bg=C["hair"]).pack(fill="x", pady=(4, 20))

        eyebrow_label(self._resultado_wrap, "RESULTADO").pack(anchor="w")
        tk.Label(self._resultado_wrap,
                 text=f"Preço {preco_pct}% a.m.  ·  Funding {funding_pct}% a.m.  ·  "
                      f"{len(self._resultados)} nota(s)",
                 bg=C["bg"], fg=C["ink_muted"], font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 14))

        total_disp = sum((r["valor_disponivel"] for r in self._resultados), Decimal("0"))
        total_liq = sum((r["valor_liquido"] for r in self._resultados), Decimal("0"))

        resumo = tk.Frame(self._resultado_wrap, bg=C["bg"])
        resumo.pack(fill="x", pady=(0, 18))
        for i, (label, val, cor) in enumerate([
            ("DISPONÍVEL TOTAL", _fmt_brl(total_disp), C["ink"]),
            ("LÍQUIDO TOTAL", _fmt_brl(total_liq), C["ok"]),
        ]):
            c = card_frame(resumo)
            c.grid(row=0, column=i, sticky="nsew", padx=(0, 10) if i == 0 else 0)
            resumo.columnconfigure(i, weight=1)
            b = tk.Frame(c, bg=C["surface"], padx=16, pady=12)
            b.pack(fill="both", expand=True)
            tk.Label(b, text=label, bg=C["surface"], fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold")).pack(anchor="w")
            tk.Label(b, text=val, bg=C["surface"], fg=cor,
                     font=("Segoe UI", 15, "bold")).pack(anchor="w", pady=(4, 0))

        for r in self._resultados:
            card = card_frame(self._resultado_wrap)
            card.pack(fill="x", pady=(0, 8))
            body = tk.Frame(card, bg=C["surface"], padx=14, pady=10)
            body.pack(fill="both", expand=True)

            top = tk.Frame(body, bg=C["surface"])
            top.pack(fill="x")
            tag = r["bandeira"] if r["bandeira"] == "Adiq" else f'{r["bandeira"]} {r["tipo"] or ""}'.strip()
            tk.Label(top, text=r["razao"] or "—", bg=C["surface"], fg=C["ink"],
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")
            tk.Label(top, text=f"  ·  {tag}", bg=C["surface"], fg=C["ink_faint"],
                     font=("Segoe UI", 8)).pack(side="left")
            tk.Label(top, text=_fmt_brl(r["valor_liquido"]), bg=C["surface"], fg=C["ok"],
                     font=("Segoe UI", 10, "bold")).pack(side="right")

            detalhe = (f'{_fmt_cnpj_livre(r["cnpj"])}  ·  disponível {_fmt_brl(r["valor_disponivel"])}'
                       f'  ·  vencimento {r["data_liquidacao_ajustada"].strftime("%d/%m/%Y")}'
                       f'  ·  prazo {r["prazo_dias"]}d  ·  taxa ref 360 {r["taxa_ref_360"]}%')
            tk.Label(body, text=detalhe, bg=C["surface"], fg=C["ink_muted"],
                     font=("Segoe UI", 8), anchor="w", wraplength=760,
                     justify="left").pack(anchor="w", pady=(4, 0))


class HomeFrame(tk.Frame):
    """Primeira versão pós-mudança de área (Middle -> Large): só o
    essencial — saudação/data e o atalho de Ligações, que continua em
    uso. O resto (módulos, rotinas, status de mercado) foi removido
    junto com as funções antigas; entra de novo aos poucos conforme o
    fluxo de trabalho da Large for definido."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._build()

    def _build(self):
        self._sf = ScrollableFrame(self)
        self._sf.pack(fill="both", expand=True)
        inner = self._sf.inner
        inner.configure(bg=C["bg"])
        inner.columnconfigure(0, weight=1)

        greet = tk.Frame(inner, bg=C["bg"])
        greet.pack(fill="x", padx=44, pady=(40, 0))

        now  = datetime.now()
        hour = now.hour
        saudacao = "Bom dia" if hour < 12 else ("Boa tarde" if hour < 18 else "Boa noite")

        eyebrow_label(greet, "MESA LARGE").pack(anchor="w")
        tk.Label(greet, text=f"{saudacao}.", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 26, "bold"), anchor="w").pack(anchor="w", pady=(6, 0))
        tk.Label(greet,
                 text=format_data_pt_br(now),
                 bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 0))

        make_hairline(inner, bg=C["hair"]).pack(fill="x", padx=44, pady=(28, 32))

        lig_row = tk.Frame(inner, bg=C["bg"])
        lig_row.pack(fill="x", padx=44, pady=(0, 40))

        rec_btn = tk.Button(
            lig_row, text="◈  Recebíveis", command=lambda: self.controller.show_frame("Recebiveis"),
            bg=C["bg"], fg=C["ink_muted"], activebackground=C["bg"], activeforeground=C["ink"],
            font=("Segoe UI", 9), relief="flat", bd=0, padx=0, cursor="hand2")
        rec_btn.pack(side="left")
        rec_btn.bind("<Enter>", lambda _e: rec_btn.configure(fg=C["ink"]))
        rec_btn.bind("<Leave>", lambda _e: rec_btn.configure(fg=C["ink_muted"]))

        tk.Label(lig_row, text="   ·   ", bg=C["bg"], fg=C["ink_faint"],
                 font=("Segoe UI", 9)).pack(side="left")

        lig_btn = tk.Button(
            lig_row, text="☎  Ligações", command=lambda: self.controller.show_frame("Ligacoes"),
            bg=C["bg"], fg=C["ink_muted"], activebackground=C["bg"], activeforeground=C["ink"],
            font=("Segoe UI", 9), relief="flat", bd=0, padx=0, cursor="hand2")
        lig_btn.pack(side="left")
        lig_btn.bind("<Enter>", lambda _e: lig_btn.configure(fg=C["ink"]))
        lig_btn.bind("<Leave>", lambda _e: lig_btn.configure(fg=C["ink_muted"]))
        self._lig_hoje_lbl = tk.Label(lig_row, text="", bg=C["bg"], fg=C["ink_faint"],
                                       font=("Segoe UI", 8))
        self._lig_hoje_lbl.pack(side="left", padx=(10, 0))
        self._refresh_ligacoes_hoje()

    def _refresh_ligacoes_hoje(self):
        try:
            n = LigacoesData.get().total_dia(date.today().isoformat())
        except Exception:
            n = 0
        if n:
            self._lig_hoje_lbl.configure(text=f"· {n} ligaç{'ão' if n == 1 else 'ões'} hoje")
        else:
            self._lig_hoje_lbl.configure(text="")

    def on_show(self):
        self._refresh_ligacoes_hoje()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mesa Itaú — Large")
        self.geometry("1060x720")
        self.minsize(860, 580)
        self.configure(bg=C["bg"])
        self._setup_ttk_styles()
        ico_path = _ensure_ico_path()
        if ico_path:
            try:
                self.iconbitmap(default=ico_path)
            except Exception:
                pass
        self.overrideredirect(True)
        self._active_frame = "Home"

        self._shell = tk.Frame(self, bg=C["bg"])
        self._shell.pack(fill="both", expand=True)

        self._titlebar = AppTitleBar(self._shell, self)
        self._titlebar.pack(side="top", fill="x")

        self._main = tk.Frame(self._shell, bg=C["bg"])
        self._main.pack(fill="both", expand=True)

        self._sidebar = Sidebar(self._main, self)
        self._sidebar.pack(side="left", fill="y")

        make_hairline(self._main, orient="v", bg=C["hair"]).pack(side="left", fill="y")

        self._content = tk.Frame(self._main, bg=C["bg"])
        self._content.pack(side="left", fill="both", expand=True)
        self._content.rowconfigure(0, weight=1)
        self._content.columnconfigure(0, weight=1)

        self.frames = {}
        for Cls, name in [
            (HomeFrame,       "Home"),
            (RecebiveisFrame, "Recebiveis"),
            (LigacoesFrame,   "Ligacoes"),
        ]:
            f = Cls(self._content, self)
            self.frames[name] = f
            f.grid(row=0, column=0, sticky="nsew")

        self._statusbar = AppStatusBar(self._shell, self)
        self._statusbar.pack(side="bottom", fill="x")

        self.show_frame("Home")
        self.after(120, self._apply_window_chrome)

    def _apply_window_chrome(self):
        apply_modern_window_chrome(self)
        apply_frameless_resize(self)
        apply_windows_shell(self)

    def _setup_ttk_styles(self):
        s = ttk.Style(self)
        try: s.theme_use("clam")
        except: pass
        s.configure("TCombobox",
                    fieldbackground=C["bg"], background=C["surface"],
                    foreground=C["ink"], selectbackground=C["surface2"],
                    selectforeground=C["ink"], borderwidth=1,
                    lightcolor=C["hair"], darkcolor=C["hair"])
        s.map("TCombobox", fieldbackground=[("readonly",C["bg"])],
              selectbackground=[("!focus",C["surface2"])])

    def _refresh_frame_scroll(self, frame):
        if isinstance(frame, ScrollableFrame):
            frame.refresh_bindings()
        for child in frame.winfo_children():
            self._refresh_frame_scroll(child)

    def show_frame(self, name):
        f = self.frames.get(name)
        if f is None:
            return
        self._active_frame = name
        self._sidebar.set_active(name)
        self._titlebar.set_module(name)
        self._statusbar.set_module(name)
        f.tkraise()
        self._refresh_frame_scroll(f)
        if hasattr(f, "on_show"):
            try:
                f.on_show()
            except Exception:
                pass


if __name__ == "__main__":
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
        except Exception:
            pass
    App().mainloop()
