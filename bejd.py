import os, sys, re, random, struct, tkinter as tk, threading, time, tempfile, shutil, webbrowser, ctypes, json as _json_mod, uuid as _uuid_mod, queue, base64
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from tkinter import ttk, filedialog, messagebox, font as tkfont
from datetime import datetime, date

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    import win32com.client as win32
    WIN32_OK = True
except Exception:
    WIN32_OK = False

try:
    import pythoncom
    PYTHONCOM_OK = True
except Exception:
    PYTHONCOM_OK = False

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_OK = True
except Exception:
    sync_playwright = None
    PLAYWRIGHT_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except Exception:
    openpyxl = None
    OPENPYXL_OK = False

RISCO_SACADO_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAK0AAACdCAYAAAGx0sh4AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAFxEA"
    "ABcRAcom8z8AADbKSURBVHhe7X0JtF5Fma2v9b2ntgKZSJAwiK12t62GBBEVu1lqq29wwDFEcm/uzTySBEggcwiBGCCurACJBDAK"
    "eciTwUBMwKCozfCaQR4SurUbiCASfJKb8YYENHl7f3xf9Xfqr/MP9/535NRae1WdOjV8tWvXcM5//nPe0OXu8OHD/8lDo9vuUMhf"
    "AG8E/nMExv2FJqvNMeOhQ4f+K/B2hI/0YBzPMY0mr+yQmM19o2Y88qyzzjp89tlnZ8A4nD8C52l9eWp8gcBfItwH/kD4/Xms8azs"
    "rfD7Egi/Gci3GidZ6JtYAPw+I0eOvJuW4XigFvBfWICmo5WkhBWSpnyreQJ4Mws599xzv8QMin6IP8pnhk8DSMOgHQuHPsV0/nzG"
    "IZId9ZYbbrjhvQgfhTCbzGb24TEgGRUMM81xOxad8gzCgxBmS9+oxf2H0wxs7lHf+MY3/o0Fa+HJgnev/MLn9l4/+us7Fg57cc81"
    "DV9B3NuA0oLpcIJNfDsKJre0oMRil+4I4Jgdi4Y9jWPyTaNyeWbnWGHUqnQkkKKC5wfQB6RjtZhSh5OSidaYTgGRFfAWnlNI5YzX"
    "cOVBwkSA6NQyg5qt8KkOWi9xkOKjCNPaQFFZx0SAdSI8KWgAcAxwLI4HAf15HsdsRU1DmhPM21DwcyyEBZOaffv2sVDy/VYgDBjN"
    "VtkxMTOiwCOuuuqqD8AnLWw2C2TTayvQO80sHYWCqef2FegdC/HQ6Ne7qyslWhA77E0AO8/A47Z1JDNpARyFnHspN4EeU3o8X13h"
    "TAiI1FAAl6N+fq0zIN7mj6rnCRb4FoBW9R8xYgQ8sZSTDUcbdX0E4+Fz5XiTZk87JLClXmY0DuG1a9cO0wJkdQCsNW/V+MpW46RM"
    "8iwQjvMBV4SjXMGSmT4LY3zLgqEvIMxZL3/qxAlZFZQ/FsTefxsLgGNHZQoGBrYsGLYdx8cDnE7TKwgiuZ/4SxS8BWFam1swj/es"
    "GfGlHQuGtexdN/6slpVf/ALycj4ptRqR5E9WhnPOOeerGs4r+M2cq5UKLk88L30ghcUOJ4QOLextCL+dYbYExxkqNJ78li+UzjIB"
    "nH/ZgUdqAZQVOSQ99KUlGl95kDABwMxSIFwf+AfhcxWRNZBAuB/U8yrCgaKKDgmlE2fMmDGChbEQxHGncyzwDuBojS+/Z4sdEyKT"
    "7IAaGxvvhM/R9/+AQwyzQD1f/TxhDhnEaqDvtm3buICy+bSQSpECCU1evdOMVAgtszmifM9X61iIh0YXrvNd1BWcRiqha7pNK6UB"
    "FCEXM649nGW4nnCgl0DPmXg5GqQBWmT9HQtXsCIZPQDnU+4HZFJx4LFHiNf0zFe/UUfHghQlTCIsCy3Qn2vk9ddf/35nDOcUnhfw"
    "GOBWog+vIfUqgVcDZnT7mGZmLcTmIHYpK2alrKjvlVdeOUQrHqQLCdmzyY9TNBto8L1huwmmt12apa/daGbSzLz6CN1NFrmNgZG/"
    "QRyZlcUJGEAD9JjpWXmGMT2m0ewZW0G5BvE6jHLxDa2Nbc0ghQNHTIVDBb9HnCwfjEeYO0Ke53FfNdiuz6oxmBuRY7kh4UU4wrx4"
    "9D1U/YBkIs1Ao8Qg2yAibJoTLcOnRrk1gJc2mL4eM/6trYtPPmzYOX+IwMfVbDQTaMKw7wbEaA2TJRrNQisaTKfHgWH4lIMyfMoz"
    "DAOUh5VfmzSYUDMEHd93333v4CDDMRmQOVXDNRkM2MXFwJaFMHjhMN5Jsqtzm+aqN9YcM2hmM/qo0aNHXwGm97NSHMusAZiGZdAB"
    "tudipQYpB6BRYZZgOJGnNkO9Y2bA7mta9x8cM2bMxQiTUZninAE2MNlAGifgMeM1fT9rIMBGs+y2TWkph4LY6swgpDS2bt1qG08u"
    "FnJtSZ1XgjcWCINLq2u/Q2HlBqEstepz0PBe7THnnnvuFBpGTJw4cRbTIL+sgADZtzm3fsx6x0KBpNE0iowz3NzcfD3S8PrB5CEr"
    "GXx2uxnIckSrhFZRf6cViDzUAOrTBh4h7AFskBkYBhKhRXWu08qNJTJGo4jOY69whStcGIwl0NNd6yKjbGYoh8433ipVA/x0xvk2"
    "D3666xyjrSKAFdMI2QtzgYAfNjgeeo5p4hWu4wxm4QozlAZwJeNyG273p6BpuAraZpxldIyxLBiwbre9g2zU4WfuL0QI8ZrWLnVo"
    "cP3YZUEKFkpQd9y0i6E0wHZevHIuB0sH56//2s8uC1AYk2IkQEaoQdmNPfvss3JZBMefQW33JTdHDBonG3mmffjhh3kFzLj273WR"
    "2QxNGckulHsQwIB58+Z9CVvGVj0m07ILI5hPw9QoDT8KaXfPnDnz6ywD4OBruxSYSTOLLgEbQGIkQPb6gaEn2LXz58//Mo1FvF1Y"
    "smE2RRnC3pjGYtM+HHElVxJAbQYzA+ArsGurvjfffPNf2eYbhj6MuIGO2Uy3AtY70nCcJ8P8UWv3+eeffxbC3LSzh9p2UcmEmoEZ"
    "5XoMfh9cFVxNA2no3LlzP4E4MR5+/zlz5pwZGcuGhgoZ1jgy/jZn7NE4tp/7/WCrzmAm0sRkVS7RMXqf1QGUGTwMA30vvPDCL6qx"
    "Nh2ljGWZctlPY2fNmsXfbo5pmT+E13f8Dax2g5lAEwsLQB8yioJCl8GnQaJj+EfNnj37C1UYG8qEfH4N6XwF6QersbxzQ5a95qUM"
    "QospdZogsAC/D1lFWEY6jmkMCyNY8JHo0s9XYyzOU5fsES4O/AXsBBoLdyLCtd9q4klNFGQwGQ7G4DAYw4LC6K7F2J0Lhj7pb9rR"
    "WH+8Z/GwJ5C2+ulME4X7V0DqBp6tYDUZe+C7E/9mz1Vf++Lea0d9df8N479GY/feMOUsPgy1Z/XwM1vvu/YdSFuTsZluA6jbHY2N"
    "jesQJ9OMoiZjAdEs0lFSxyB8vGr2OIT5AylnF56vfhpjAk3I1okuWRDXdPiyQiGORsmSW6WxLE/GAcrh1PUNhN+hmg2GAsxf83zL"
    "xJm5lgPt1ltvHcwKAbJKTVc7wIKxSGtT18CdC4RZP3hrM5SOiQHpOlYA/6iFCxeeoTOD3cciEyljMxXSB4KsaCzy8Na7bXwsX3U6"
    "jR0zaMYw6gG7vU9Z0GBW4udZP+2wocxvCL1EY7k3QDhsfHCubYaaY0bADzQ+Jfc4JvVHEKbBXM36mLE8jzhb51l5gMaJ/mms7rri"
    "VatthtIxMyDsAmGgKbvcMXHFobG2N5DbnQBHNI1gHgGPAekdDLBdMJYTt++JtrNqTgspGWhLliw5nWEC8XzAg1rmChQuXWJofH9N"
    "axIg4+1j1RwLAWhwGGi33Xbbu2wJxrEYi8H332lENUBPfA752r/xjh0L0cKM3TDQnLHCGHxuRjhnct3nWk/w8oV3yYV1Tc8y2raH"
    "reRYECADjZUgfNSoUaP+F7T3Eo7ZnSXG4txeZfKVPXv2cAsoRsKnnjm71N9QOhamhaamMc6VggsuuODrlAdXunHjxl2GNGSTjREm"
    "kYZzKUc+u96mtvoZao6FsgJUGtjFXmGTsneQPg196KGHuM7bE1Y2M9jmh0wGIwktvr5OCxd2AZsZ5NJmzZo18pyCGsg406R1d4ZJ"
    "QovtOKcVsVIzmAZx+qFxNJ4LB0e4dXfnMJlyViFgBtMgGiZwLHadkd5Z5WoIDaJhZpwZ2LVGxs6MSUGTFK5whStc4bqtS83e9YJW"
    "8fpxEQG2ZnvYet5WxOX5+nof4VEDrdEkgpsj7uBsd8cLVrshWDOsDC0v3hlmiFbTeq7zjQFiQmWrDFJ4RWT3SHg1xf29AMdyB6sS"
    "ojwsg+Wx3O63BW+vM+MVGVLZWDZcSSAZvIDnpafcrDIgXq76PKo5B7AcuVMHkPz4apF2ZJSsZnc/ZwY6mOG5pAJUm/322repqWkd"
    "r/X5+2t7wXJGjRp1I8tl+axH67Or3RKCtSld72hMBE+kH/Z271aGPmCkyj0JoN/IkSMfsZsoc+bM+R+M4zmkMQWKCgnEU4kBFq9p"
    "TM28kfhZlsdyUf4TjNf8QjB8U3D3mCLMAIUnk0QambIYwY/nUQ5NkiGkwuddqAFLly79JG+fDR8+/BUeA3IzEhAiAJuDZf6MwXg9"
    "F08vJJq3lg+w/Llz535W4+L7h11LsFWqBnhCM2QCvpFGpqgKYVEUwv3/8Ic/DBozZswqhAfx8Qs2HiTwx4D4Lq8R4Fd+wnemQNOJ"
    "DQAXNSEY5e5m+fylA2nYqeGOHcJMb2V3PsFWmVZspNpWiY3PXZAQJ4uM/qJ9wIap+q8gTUyuDF0gVhbrtManEOwCrJNl6jFy9Tcv"
    "m27ybomyDCuzY0m2CrQyNqDsgkTMmDGjAQ3aaUSyYcCfRo8evRTnRbnwOfz5N5ij+XiLkavleFWVbbAd67mMfQiLeo1c/vqJeNYr"
    "Uw8Q5mAgJtl3ZKbOujgrVCtIGg702b59+9Eg8gCJNDKx6l/34IMP8t03mQVI82SmiIjceMhWHK4ab0TYqLIRFci138ERN2j3oqEv"
    "8rmN3YuHbUOcKVlIRjg1VeTW3yanhWWMBrhIsZdJVJ9bbrnl3VQowX9COkNt0aCxsuAQemxTiKidvxRH5MbbpXqSK+9w4Z/G+GSE"
    "vh+HP55xFLGz4/o7hmQthIWx4GA0/EDO+PHjL6RaQe5ziCuZy3BsV0bxohemEz7gUCdyS+wEOHI4BZA8EjsYccf/B7nyuqTjGa/n"
    "5cc+5tP8tLX+87EWYIWllEsi+OvkyyQYq/+lGpfqeUIIBjId1B5yLU7PM53tGrQDd2F07AFh+0BcKwjcfwLi39myaNh2R+5Jhw69"
    "/M7Dh/eD5NZjDx3ai6ljD39prUiwmlG7Y2ZFieFAmHMRJ8/9kqDzzjvvUzgOc6YzSsBjwBbDDiV316Kh23YtGHI4BRJrSJ0ndi8a"
    "9i8JWzpNvaxYCL7pppvkUQviueee42IVjNLGMl+nKdfqAFgepwU+LcEnKI4DTuDbZJxyTwQ4Nfj3acW7lnjubR+x5rQgMz53euDL"
    "UEgSofEypDQtG0ySaSjj6kIuncZn7EMZYUHD/vZzWDS/vHHjxr9BHAke7Bc0HAup8PPmW7aX5daXWDorTAtmBanhLQTzcRbOvyBr"
    "h8abkba40RdyAVF9HcklSjof5YbdAuKpzEGB3NdeLBhIjeqOSc21oV3OCtUKMgQjnNlaYWr4LQkeOXLkRo3neaqIRpPozJTSUcqF"
    "Lx1o5KIeea4eGLD3O+PP2PvtkZ9qXT9tGNNEdXYOqd5ZBVpZIBg+jQpKRBx3EIdI8IQJE6bhmHMXjWcaAePgy2Vygtx4rks21MHi"
    "MzZpnYFc/TsIrw7l/gXPMY3Vpfl8Z4a6lIKOdVqZVcxG0xiSEBoDCMG8p8pGrVix4n2MR5yQTF/TyVQSXURIowF2glcS6zGSYzBe"
    "bEF5JdOVketu3MhiBZ+dzLRGqu/EziPVO62YBljD/BxHgkWRl1122YfYKJLMY40ncUIs45A+Qy6PGc80LEvLlPmaQFgWRg93Lszn"
    "CAux8DmKYnJD52l+I9ZE0/mkmmPligzBbBwQFMOGYFqYg8bxXu2fESd3ynhOIeRu2LBB/lilnSDDlvE8r+lM9QaSHkPOMS3zIBzu"
    "XdgIuvbaa+1lt0wbTztdT6w5GqEwgnMXuMbGxrtJMAj8dx4z3sBjgL8anEESmA5EHEL4SWBre8BytF7eKP8U6gvEArTTppuunQpS"
    "zoxRwwLB8FMLXAsb2tTU9D0eM97AYyDcily8ePFnoPhvjh8/fjmB8GUp2Hli3Lhx3zSMHTt2GTF//vzPuLoyixfQfYk1Z0apgRxe"
    "NJjGxwsc//gpw9PmvgQ4HXDPaY9580aKfy792O9973sfRgfdbsC++jZg40UXXcSXZvvpJCygCIfFEcfdZ/GqxqlxNDKXYDb63nvv"
    "HUyCibvvvvudiCMZAUjnb6IP3Lx583tB3J0Y4jJdEAzzx8Yrr7zyw5rWtlUy5OGHPTV8WQjhk9B4xxFIJbQp3dOpkZ5g2UEAGYKX"
    "L1/+EVPw1KlTJyOOBA14/PHHB/MVPYwnjEiEX7jkkks+belQDjsgNdS9Mo3ImExDzyDVnDPYExwUDGQWOQzpdSTPlKyE73C/c8kQ"
    "J5gHsF2DDHP4fosWqzKXTEJN7lnONcAaxUbaIkdVUWFGctjrOgiBBNMAgUj4YZgD5ZTZ84ks51zjjGBTsUwVShLJovoCGKfxYa4k"
    "EC6nzN5NZspFjTYiPNEGEmfw8b13iNfTRYR4oirB5xNokYXLcynSYmjSwhWucIUrXOEKV7jCdZ5LbcsqQbMWji5FUD2h1bx+XIoE"
    "IHVV1haUlK3V9l6XaHRMit07aCvi8ohQn5rRu5xvIGCNNkJ4Q8bgb9TUAl9GiujeR641ShsYE2rEyB0wu51YKyy/luWJ9gT3HnKt"
    "MdowI9UIFSLh8ya33MslEJZ7t9XC5bNyUvd5ew+51ghtkCc1EGpEwpdfIQw49g905CLK42+qxwT3DnLNeG2IkSoqBUyd9pMOSfJP"
    "w7QJWkbmZx/A/+TTc8k1gxVJUgH76UZ+7QUyv5EhPvOcQrlz8XkgPNYEPzyHAMTk9gxizVCHXFIBKooNF0KVoMxzCYoSEiMk87BM"
    "lo1wTC7t4FRktnVfYmlcBE9oWKSAElJJwsKFC//+rLPO2sGfyvnTeXvAMlDWzvnz539SCc486wVwzu2+qjWDHIxMIzSQqo2JSRWV"
    "8tkse5ijjsTKAyKjR4/+FutAXfbcQkq13YNYM8Qhl1DA9qS2fcqQyseSSAaJaGxs/OnmzZv5fvHMUAb8ohRg8ZomTCdbtmw5vqGh"
    "4S7rKL67X88F1eKYo4c2dg/VmgGKmFAjMxAKP6z8gC1SbKQ81wUC7iEBIPfXPNb4MDcqMtupGDzPdAhb2TLvgtTHWfbIkSPvY7ye"
    "Zx4umt1nOrDK1ZCqCUXYVBpIBeRBOTR8J4cs/8SBc/b8VsnDcATCcgHgofGZ7RrCMm9fcMEF9h+JfTzWeKZh+uQiRmhzO95ZhYoU"
    "obkbfW1IaCwQSAUGouEH2XgsNP+N5zSdkMpyABIgm3wCYQ5hm14sLjkqFi9e/A9K7KuIl6kFfjwddJ1qrUKtPEMqDMwj1ObBeCs1"
    "YOXKlR+dOnXqTMTz87ieWFlkmB8QUhHH8m00mML8CCFkt4G0Rq7UGxFrU0xqOuh81VpFWikRVAoEQgGvThrvCSVh/SdNmnQB3znD"
    "xo4fP/5yxGWI1fRBUYBvODszhYw9gNgCHMktnCcWCNMMINOBq4NldQ65VoFWVkIqIAoFcgm97rrr/hYL04O28hMMQ7GzkS4Q61/C"
    "A9gcyHqMVLMhRiAX+Wz3ISPGE4s4vzBmRgXAejy51u6OIddVYA0oUQYbAL+E0GnTpjWhQS/7PSVJnDBhwnycl/mViIiVxQV+2WHq"
    "YPFGrtiGvDIdxMQSiA9zLfyY3LgTpR6loz7OCtUKMoYDntRAKL96jO3TRiOTYBiNexKNPB1pZIsF3xavo43YOXPmfIZlIY7q99NA"
    "RkVqXrnRFIjF9PJxIxbH8ZPjqUUyJtfqqB+5rlAjtmSowadxfS+55JKPoAF/jNR5qLm5eeXzzz8vKz3AISgdAN/I5afSPbEyRBFv"
    "xGYaqaYFxzhFsBGQjkf+t0fESkcqWLdMC/Bjcm2UdAy5WpgZHNQKQ2RORVhIxSXjlUaoqvPZWbNmcSEimZmrJR4DgVwij1j4JepR"
    "04JjnKIaYmXq2bfx4iEHlgw9fBA48OPV70baQC7rBeIFs37kuoKCwayQFSNsau2zevXqDxqpuLq5h1/tYzyQuSdq0ONALsDPcAux"
    "2Mx/Ws+1i1jkJTF5xB6zb9PlJ+9eePLhPcCBX3z3PUgbXiYBv2PJdYUkDYZP4vhX/H8lqVjlH8NxWHGZBqCyubpziAk0LuQHOo1Y"
    "xMnnRFrvXjksEPvgTXzZBF9NJfMuIDsSoCy5Wn3tzgrQwsRgIAwx+DIN0GiCyuWxxtt2yeYsAeJIFuOC4pnHE6vx7SIWCHbiAoF/"
    "kXoRHf87HMvrp1q3XPUhI/bgQ7f+HeL4hz4hFxDlAilyWUf7yLXMWlAusTqn8o/Df4e4oFamA2TzbdBjGkujSaB1Tr2IZVqx09VB"
    "orgToFr5T8jjWn+65sOB2F/eSUHw3TNJclkWQDtYbrDDoGZU71zmQCwq8lOBEfvvnAoaGxvXIy4zT8H3Rvl7CTYd1ItYnjNimScQ"
    "y/IBI1be6dV67zWnBWL/76aTEc/XURm53DHILgZ+EAhQH9VaRi0kQyzAhUjmSH48iqRQuXCpYWTECrlAhliAb+iU15/Wm9jDh1tg"
    "425cxu7FotUK0vYfd/jwyyfu//m3P2rEvvKrjcMQd9Jr7/riO7/47q+9IHdXmG9ZnpabsYVQM2pzmjlDrDaYDbdhJvMsVTt58uRG"
    "jRdigCSxgKledgcdQWzrrbMGv3Lx0Mx3Lg37Fr32DUyC4VQa5t1/1+VUsRcJ21BXYjNGa2UkQIZyU1PTVSQW5PDVJbKqAzaEuoTY"
    "XeumvoukpV6ARhixqXMEyd2/ZRnf79VhxJrhNJoFlxDz9NNPD+RUQGB38G6eo0FM6wzqVGJZf8uKz35u96ozP88Pne65puEre69r"
    "+tre74wdvnfNiMkkT6aDtY3j9n53wgh+EJUfSd3z7bO/zA+m7r7qTL5Ku0OnAjOcBeZOByD1CaoWFwk/ZjwQz7Nh8QI6kljpfOSX"
    "UQX4xYsvQstbvOyllFy8uG0MIw/HJg6Wb3W1nVg6LcAbHhQBkByZDrBflI04bwfiWMgBwpYLfmpXUC9iM50PWOdJxwPc/Ntr/Qa3"
    "3rP61EDsoxs+wDg9xzRhRwDUf7tlzhUSDEeFJCjsDgBRLW9gU7Vjx4491wwDSJCQq6CKQ6cwbwcS+3Z+RHfWrFlfmTNnzpdx/NoF"
    "wo9XnRKI/ecfvE/jK+1h66dWOitEC6TholpUbASRCFHtuHHj5pFYEqXxfhETuHwdRizySsezjrlz5/49536Un76kvX/9XyPOvu6e"
    "2i4aqcEGQqtvn9PCWCjBSjKqgB9Uy6mAJGFq4CtIMqp1eTId0pHExjdhED8ovgmD+HD7EEiRam23uupKbInxNABhkhBI4u/3VC1U"
    "8kuN53kzlum7klguTAP33b18yB5sxfYC/rYhUJFUrbZ+Tgu2SlghG51ZxIA+q1atej8bQuXu2sU3GYdFgFOCkAo/0xn1JhYIoyki"
    "Vn61aH32X9+xb03Dp3etHv6PSEcbqhr+hFZbP+cKDw1QI0oWMTRiN3C4ubmZz0zRaJsSBEzPePhCLBqevG0IWCMzDVSTYpsIkiHE"
    "sh7WkSC25KcZoGtIpXMVGLGxagNRU6ZMGU9i2SA43lMwFRqpmXm5CmJZl5EXGqqw+KRNZX5MTJFaMkIMSkPHOFcRK6UBog5Ahh38"
    "QJauxHw9KX+ikVuJDiS6hFj7aYZp4NtuIm5sCkZqydyf+PmbpLJe6Wwgr57OIZXOVWaNCQ0BMqptamr6PlULgp/TOKrDpoVyxJYb"
    "ntbwGEIqIGoFwvTkiUVZ/oENmZ40PfMxP8uyzuocUs1phVY5GySNoZE0FmEhbMOGDe+kaolNmzbxPqhMCYSmkcUOx+GnGSW2RE04"
    "tmHKxufB7Mh0crlHjDRt7nSjTe4cZ5WqAezh1NbLlLidqm1oaLgRcSSSDTKIsoFALB/YYJyeC+QizLKpLCE5hp3TtJkOjojNzK2a"
    "hx1iI6FrSDWnlRuxsWpFKWwEXzOt0wGiRCkk1yDkA4FYDFsSa6oyAmTIEoijgpPQ8yULoyP2TzxGPMv1c3jXq9WcVa6G0KjcrRcb"
    "RUyfPn0k44AUsfLzzujRo7+NeC4wNhfKsEVYphCFLYAeco5pgUzZY8eO/RbLRh3PWnmAjQSbBrperXRmgBoj0wGQvH8AsuQmOMBP"
    "ARhZAh4D/bBz+BrJp7KXL1/+McTZXlMIJny+ctD0zNePr8FmmSwbW8AGnA/TAJC8ACG0mV3j1AgaE1QL2DwXhiM/fcWrMDZw7dq1"
    "/A0/kMSwHvNpmN+rstgJu5An+ebjaoEy5ClxLfMPWme8lfPTQNeTSmeGqFGiWhidXMTQ0CfZQPhbeMx4A48Bzqv9Qf5j7ACmrQdY"
    "FrBV68yoFehe04B3aowRy55PLmJLly71N8EzQ9sdc+gP2LJly0mYGhrtldLx66Y9yr12+pxzzmm4/fbb+V0a1iOkwg9bOMCrVYjV"
    "ZnW9ozGKjGrhlyxiw4cPf5UqAhGzeEzYOQXnRblJAsjtPfi8Ac07+0R49bTFMY2mY57UvMyFTBY/+J7U7je3emcGqXFBtWhIySIG"
    "NS3UIXqQxwnIdACf5MbEBlKnTZs2h+/xJvgub/h3NDc3r9O8NhIyhMK3xcpI9VNA9yOWTo0yA0lsySLGBgPhRel85bTGVUUshvSQ"
    "xsbGn3EqsfmT4ZEjR95/3XXX8REh254ZoaxXCIUvKkWYI6lnkEpnhqmRYToAvGptEfsnkgJCHsWxbIk8kM6IlTfQ4wJjDDoi82VS"
    "+H+ePHnygr17+QW6oNISQgFRKHxPqB/+3ZdUc2qgGSvTAVCi2jVr1shNcBLFjwwjXm7jGZBmwCOPPHI8P4NlZJo6cbzt0ksv/QTS"
    "yHwK2BVaGPIA68sjtGeRSmdGqsEZ1SKc2XqB2JdIFohqIbmII0kD5s2bx38Q/s6rEzg0ZsyYq5977jnOr5nFCeGw0sOvpNASQgk1"
    "v/s6Z6wZn1Qt0Oe22257l82VSqL8q4ZEWhwJnj179pnIF5MZtk4sE4gJJZl5hPYsUs05o9mAjGqVAKpLruP5jVyQuJskGkg2Fqgf"
    "RB+oKFnlASEUx6lFKR7uGTIJNbfnOGe8NYiNTJEb9rZ33HHH8Zx3+dAyjw08D2TIRLjcKt/7CPXONYQNY0PjKYHkyt0oHJM4UbDC"
    "jo3MMH/CL7cosR5Pau8h1DvXKCOXBAi5JEeJImHhdp+D3y4FdWre3rMotcW5BlqD2XhRrpIjBMM3BQdoHMlPqbPikFcTeq9zjY3J"
    "FfUqYaLEGDlkliWU0Kp7v3ON9uR6gg0k0RNJeDILQmMXkRATbCR7+HMFmeVcTArgCSuHOF9BaMqliKoGmr1wlVyKvBiatHCFK1zh"
    "Cle4whWucIUrXOEKV7jCFa5whStcXV3qblx3gppZuNebS4mhjUj9OFIrUuXWDG1a4XqLS3VyAilBdWek2hCgTS9cT3OpzlSkRED4"
    "3+k94t/zOwMpOwwp24lUWwsBd3eX6jRF3MFeBF4s/umdGPaET0ciVa/B20n4NsTtK+FAKSpcd3GpTgLijoxF6gURhMPn+Aw4Tj7r"
    "11mIbEmJOyXiuN0ZXpSywnWViztE4TssT6hBoPC9SOyJXj4mLcBx5snfzkJkQ3jSmEA4FrIXcUq8JTwphYXrLJfqBMB3khdrEKrr"
    "7IxA4ZtQ+O89e/7fI/4vQEcj1O1syogZ4YyI4dcsXqWzcB3pYtIVvlNyxQq/RKjwvUBFMIjjX60I+wdQCvYvoTz4NJXS+3JjmC3h"
    "X0i0FccZEcMXAcOvRrwZ/pTawtXLxQQ7eKESFcUKPxaqFykFEoSEeP+2gC6Hsy0IGfEZAcPPE28h3I5yMZEJeJESXqg1iRXhIFQc"
    "izD27NnTf9myZadNmjRpxpgxYy5vaGjY3NjYeFc1YNpakCojBdixYsKECecuXbr0Yy+88AL/Z+9FLAJmewA/+xbC7QgXE5YDL1CD"
    "ER+LlZ1Ts1hffPHFgZMnT54+YsSIl/z7AbobaBftQ3jnxIkTZ23dupUviyh5VwHCYdYFCuG2x8XkJOCF6ZEnUkMQKoFwObH6Zb/f"
    "eeedBy2MaPVC5bsXGAf/ccx0t44aNeoWAuEfdCasXtjxGOzaZy/gIGgvcGDGjBlNbAvbBYRZF341M67nOPSDdtfr13kyEvCkGVIC"
    "TYpUO6JEqATC7DTbr/otQHjp0tSpUyebCEwIEMudcPwKgL11RV4WQrAMLUvEQSBe6ojAQVIOJXmsPC3b2xsGGF9gAhH/bz/AGMas"
    "ex7TMA/zIiw2IJwSrnFYCNc73/gIniBDRYESJFxJL7mXCt8LVUSBcGYLAPi3gw344Q9/+F4stc/ocivLLvaOl/GcpkuJNewdCRxn"
    "bkPFwLlwj5VIpSFYBoFwEDSOUwKWQTRu3LjFZrfOuM/zK+1Mw7TMyzIAlid2wCdf8WwbCzfTX9qdvdfFDXbwAiVMpF6oZQVKkHT4"
    "mc5H2Dq6RKgE4krECvDNakfzs0vobHkXNgEBtPKd2JpWBKvlZMSKsIgUvtmSueFfK1z+5OAjcOyF23fJkiV85f4uZ/ufZs6ceRbP"
    "MZ2mz2wT4NNem23L7W+J0H/avb3P+UYqPAFETSJFOClQAuEgUgJxSaESOA5i5XsV9e118i3x+fPnl4gWcfywBvPkLrXwRajwzVaz"
    "XdqisPblIaR1+TNth58RL8LSToTl5fCwd7ez/VWI9us8RyBd0nZA7Ib/+hWub5jCN5ooK1T4ZUUKPyXQzLJJ4Jx0FiBCJfRFzCN1"
    "D7ijoaHh8IQJE1bgnLwbOE+0LINlIY3NVuU63TreYG01mBgM8Xmft4QT1olwLFz5VESOaOMtTXKbAD8lXII2+v7L9K92e890cWMA"
    "31DCd1Bup8BPiVQESiCukkhFqD//+c+P4+cxmpubv4tOfB44ZHs+w8iRI6sVbabD4Ytg4aeWV+twL8y441OI0yd5gp8RLnzhgl/k"
    "gb0Z0WKADkcaWWEAE64NvCBcRUq4cVvM1rive554E41IdYZ1gsxGSk5GqPBNpH4WKREogfNBpA8//PAx/JgYZtCrIcxt7LRYoBZH"
    "IPy7sWPHXoOO/vwTTzxxHMoT0c6bN+9LOJcRrf+ADhBmWSDMUEClTvYox5Uh5sx4iwe28ZQrWpyztybHM27ZbQ783ivehOEx+Rni"
    "lYw8sZpQMyJFfBAof/3BUvjx0aNHXwYB/gs6588qxIxITaAIv9TY2Lges+5XH3roIX5oPXPxBch+lqgkWviZZRWIOzdPsDFHefB5"
    "rBzjL2+2Fc5SokWbeSEm7UWaPPGGWRfhSuKlLdZGb2tJW1Qe3dNFxvqGeMLt+xXViDUj0mXLlp3KTzGcffbZv6Sg9CZ/ECdBcWp8"
    "KwT6w6lTp468+eab3xmXBchSifjwznaF3T04k3WwfK2j1b4BB2S2BoDvVGkjULFDlbbg4vOAz5vLIW1AWPiDfyS2MR+Hvbmi9UD6"
    "jIARbot4aVvcViLTHm1m93GRgd7wioIFcsWKfab8fFpOoAi/jD3pT6ZMmTJ23bp179IybDsRI8zcVgdQIt62iBYo15FVd55PC+Rx"
    "yboCj0BZ0c6aNWsEzmdWE36DumXp6Wtblpx2e8tFp93Ssvi023cs/fhVhx6/lR9yZDuT2wYgFm458fq2CLSZXesShsUkJ4kmCSQE"
    "CIJFWIQ0adKkaSYaCpTQJf5lCPW+8ePHz7jiiivejzxCrua3GcKW7hTkvKYX8RKIy4j3wgsv/KLVrzaYaKUulgW/S0WLenk7rITL"
    "MqKVb9UY9m1cPqRlwdAX+CH1lvlD5GPqOxYOfYpixnkO4LDnBZJtht8zxRsZ4w0tIRk+O9dvCUoEy296qUAFJB5L/fe3bNlyvKVB"
    "ek9mECmOjdQkeF7TSAczPyDixbEIl5g9e/YXUG9GtPq16EwHAmHZBLq1aJHOPq4kH1jat3nF0BLRLhr29IH71/81znM2Tn5xBShp"
    "O/yeJd7ICG9gVSTjOIgGhH8GRO9Tsg8PHz78z9OmTRuNcxUvGgiEhcg86HkbMHnC7dsNRWvlBT6B5PZAvzwbPrCC8KtYOc7G+cxn"
    "1lp/vOqUlGgP/vMP3sd0KIsip3i5reC2iSuRTBYI20QROIffZvFq0zvPRQbURDKQES2W/bk2yyrpT99yyy3v5jmC6YASwcL34hHy"
    "PFy81A8/KVyEpR7e180TLdMzH/MDXS5awNoi7XjqqaeOXr169d+uXbv2A9/5zneG3HjjjUOefPLJE3DeBDsYOK71ntWnJkX78O3v"
    "53lNyzwUu4gX5fttg/SD1l1OvN1PuFHlFUmGnytafrsKIjlgYuEXMWfOnPkNnmMaTRtmOvglZBEIs74AF58SbhAtIHZ0lWh9OoXn"
    "07jMu3sgXMK3bQ5nR/k0IPwgWBxzm3VC673XnNayYNj2rGhPeebgYxuH8LymC+LVcmTbANS6ZaDtxou1KcMLoTR0vIsq9gZlRKuN"
    "YOeKYNhQwAQTluampqZ1elfABLOD9x6ZBsiQBJ/lpAQrnesQxAs/7uySwUPRot4weLpItG3hUkQL3+6EUGScKTOChX/i/p9d+5Ed"
    "KdE+vnko0vA24Ymalnk4O7MMm3VtvyvCBZJbJvixcGNeiKq4qavzlSrKEs0GAbmz3L333juYX1v12wSI+Mn169efxHTMA5QjKCXa"
    "EhvgxzO+DB6ip4l25/JPf3nHwg/dsXPRqXfsXPzhOwUXnbZx55KP/Og1fHST4OKPbX4NH9+8c8nHfrZj/skHdi4YIqJ9zT95/66L"
    "P/aTnUtxfunpdwksj5Sh5bFs4rV6Nu1e9onpxgvtUbvYz160nhvfNt/mLp9treMC2UAQDBsHJGdbfi8fYgnPs3LmbWhouOP555/n"
    "stSWmVbq1/O9SrQvr59+0q6FJz/DmZLi62zsXYR6Fwx9cv9PLuO+OSNa2qi2Wh90W9ESZclmg7RxyeWZmD9//j9CMPtVNCJcbB1W"
    "IU24EANSBFE8sXAlTs8H0TI/0KNFu3fD0oE7Fg77xY75Q/4MHHgNJx+sjCGvxAJE3CHGp9PnYN4HX21ZfNq3D+/cRm5Cn9A+tdP3"
    "RfcRLV1UuTcql3BtYHKbQEyaNGkKBWvgzDtx4sQJTA+E2RZ+nnAz0HNBtECPnmlZN+A5pH3+Qsw+fs8LMblrgOPjgRNfeWD9qS2L"
    "Tvm939O2LPrQUwe33s0LsbCnZR7NW/FiDMjri4q8EEpD57rICE94hnQgCIcNBbxwMsIdM2bMN22boCLae955532K6ZkPaYNwAYon"
    "KVzGaXzc2d1VtITnkOWWcAh40UobEI7vHoSLMUCE2/rTNR9O3j345Z0f5HlC0/u7B229Z5viJMMLoTR0rouM8MYZ6WxATc8g7N69"
    "u19jY+MGf0cB+N2ll176QaZnPqRNCjeGnWM6HHc70dL5tIDnz3MYRAs/rBjwTbSZOwgI22wbhNu65aoPlblPG4u17B0D+JlrC/jG"
    "RfcWrLnIGG9khnRtnBDPRiMcCygI94EHHjgWon04uqPwwD333EMyS4QL34QUxOuOw0CB7+uU+ohuKNokf1q3tIU2AUcsXLjwdNj8"
    "R9j8J+BVhPdfeOGF/mdcEW/r3SuHJX/G/T8380+Qx2j6WKzkp+RHBSBwDb+m2ZXQpneti4zyBpcQDz9PuEFEOO57+eWXD4NYX1QR"
    "yYUZ/y6D8+HCTPOXCNfAOI3PFS3CXXafls6nBSpyB2REW+7ZAwJpKNxjcp89uO+G9+K8bAOQvkSsAOvJcAw/Fqtx4O0nfNsE2uzu"
    "4SLjzGjrzFzy4VNIJbMtAeH8T3REuKPAGbe5ufkipmUeoES4MTSenZxbV28ULc5x1gxPeuU+5fWz6/mzuV1ktVWsvu2Eb5NAm9u9"
    "XMLQmPzQAYDMgvClA+DnCnfixIkzbZtAsGMQN5ppmQfICDcFPcc0Nsv2ONECgTMgtCUlWrSHD4HLA+6GvZuWf+ClBcO2vTRvyKs7"
    "5p98YMe8IX9qWTTsl/vvv5G/gpGLuotVm9i9XWw0kOqAmh8Ob2pqusaEqx2zi3s5pmUe5kVYhGvAsZVrx2VFi3IrPuXFcgDfkdIe"
    "wDqx6g6MzwM+by5ftAFhaQv8qv+5gLSpfy3UIlbfzrJtJbSZPcNFxvuG+Y7I7G/hZ2YPhDOi2r59+9GNjY2bTLj00UG/XrVqFR9g"
    "zgiXQDiIldA4KZtgHqAeoi032xKei3LweaycJFe0AWHaIu0oJ1qcF6ESOC4rVoW0TVGzWFUCPdNFjfGNrKYzksJdv379uyHW33jh"
    "NjQ0/OxXv/oV923WARnxGhinKCmXqEK0oXNxzAFWyyxk8JzEvBgsv3EkPAFhcNMGtUXaUuO/cTPtQVgGInxpE4HjeDBau7ydcVt6"
    "tmDNRY3K65S8CzPrFBEXgWO+AugMdIq8AsiEixl4jabLCDcBOcd0mj6INvUQuPuPWO6+Fn5KuNbJBt/ZKfi0lt/KqzSwq3rvAdK1"
    "ZyvgbfV9KtDu7h0u0cC4k6xz8oSbFNiMGTNGsGO0g6Sj+JJhTcf0XrwZ6Hmmq0m0QNnOJhD2HW6d7mHt9u338Hk9J3mCZVuqfcNM"
    "xTbA97Or2ZgrWO3m3ufihgJeuL6zSJh0DnwRLuA7JyOysWPHXmS/mBHDhw9/Zfr06V/RdNJBBleGIVMWkPyPmHvvQYlwEQ7LKnzp"
    "eAJh/2sc2+ThRWltDojyBrHC9wM5wwnC1b7LK2M7UItgM32oXdu7XdRoI8ILVzpRyatauKNGjfpe9FPv9m9961tDmEbTZsSrsPiM"
    "aFP/xsUFTniXF8E8zIuwzNpAEC+BsN1e46wowLGIrxKiPFaOlVsiViDYX060TKfpbfCavYVgK7mo8XnCrbR/y3QYHx7HhdgD/sIM"
    "In5yw4YN8rIOhYnUQ86xDEW/5cuXfwQd/kfreP7B8pxzzmlGOruQKREvYLNXEDGB+PgiUMRXDok8UhYgZSMutJ1AnNhNQKBo/mvb"
    "JRXtTgqZaTR9mGUBG1w1CVa78fXnPAlAW4UrHYewCIh/6INYnzLhcuaFkG/nQzdM55ARLGFlAPL+A8zcP4hm7hcvvfTSM3A+vscZ"
    "Cziug7CZva3wZZXYSyCuPwcb7P2ditXav5GvkNJ8lS4iZZUDTLS+XwrR0nkiAC/apHABWSbhU7jJ5RHL+GfQYeE7CRRwc3PzKjvP"
    "tCnYeUBEsG3btkEQ7ma+UZHlmAhQ3jZc6M1etmzZ6XfddddfGTZt2vQuw8aNG0+qBK4AbYEvg68spVAnTZp0Lmz7NxtkZmtjY+NP"
    "+VI+baMMHrTNb2PKbQsKwaacJ0NRF+FOmTJlknUeQQHzgXI7z7Qx7JwizF7Tpk2bgPw7bfb2ZTKuq2GD06BxeyZOnDiVbdH2ZWZY"
    "+JUE60Wb6SPtute3i0kBvHBNtCJcQLYJ8O3CxLYJsXD78eFxm3m0Y/fNmTNH3jFrYHoPfw4IwoU/gB8I4YyGmfdRlHUAOMTyuxq0"
    "A+3jC/keGz9+/Cz+6EL7tU02u4Y9rKIQbHtdRI4RViJcJTkIF/D7Wwo3s8xjebxTO1aEC/yedxRcmpqEq7CHTjLvxsKxPPKn4DOr"
    "1ULyRGXJE1mA1ZV5XoDAsdgGxG0RoRKIy8yu8OM9bCHY9riIpDzh1nRhhn3fCRDtI1wyKVz6OH7w0UcfpVhCZxPM5/M6iGgJhL1w"
    "U6L1YuST/+Vg6TKiRVgEu2/fvkG/+c1vmE7ECj8jUsTZQGOb/YwabwPKibUQbHtdRFZdhHvFFVecCrFu1y2CLKlNTU3fx7nMLFUG"
    "IhaUWSJaAvEp0QZxtrS0HL906dIvYruyDnb8HnaUbCt4sYerfAHO/xEXgN9dsmTJp37729+yTLGTbVLEM2kQKYFwiVAJhKsRK+H7"
    "oBBsNS4izYs2KVyg0oVZv/PPP//rEGz4qZczLn9F4zmHWKySl0BZyZmWQHwQ7UsvvTT48ssv/yRfiY86nmF9rMvqNTBO4//Q3Nx8"
    "M+27//77+c9XP6tmlnscB4ESiMsIlEA4I1ICYS/UqsVKaJcUrpJLkFcX4eoLmkU06st7wpAm3iMG2Dk9b4IV0fIzTitXrjwNFz/L"
    "MVtuRXnJ1+fzWON3Yga9E3WO+dGPfvQelsGyovpTQs3MovAzv7QhnBJpPKN6oRZi7QiXINIL1zqhZuFiVlvD5djERKHxY81IE8QZ"
    "A/n4nbGj+RZCpJ0NAea+Pp/HWv4+LPn3Tp8+feqNN97If7dSoCJSQssVkRI4TgoVfu5STyAcC9SL1AvVi9QQc1yItb0uItST3S7h"
    "jh49+ioTrgkNeAX7yc0TJ06chb3kp7FcfxXhBUj3T0jDt4+XE+hBCPQ+5r366qtPRh1hRka9JnybvXNFSuA4zKbwq13qY4HmiZTw"
    "nAYo5YWrh4vI9eRXK1x/OyyI96abbnoPRZonxlQcl3iEX0Gex8aNG3fxihUrhqI8PyObMGNxmkC9SKuaTeHXutR7eO5KoBQXriNc"
    "RLbvFOu0EuFq54t4gVi4Qbx79uzpP3fu3E9g27CCYoQwnweeM2AG/QV/pOD3ZpkWeTKiNLA8RRAngXMZgRIId4RIPUdJKJ2F6yyX"
    "6ATrrFi44QcIgqKgQBQiXgLnMgImEJ8RooelccgVJ4G4pEAJhKtd8tssUqWtcF3tEp3jO9GEa+L124XMzAvftg2xiMvC0iqS4iRw"
    "XEmglUSaEmrc9gClp3Dd1SU6zXesdbgJN4hXBWMzr986mNBEfOXg02ueWm89eYF6kfo2GOJ2BigVhetpLtGZvsO9eDMCJlRQIjDC"
    "RGdAXEaM0bmQj0BcrQLNEykRtylAm124nu5SnQvEQsgTMGFCC7NxHnxahS+HqJtACW1i4XqrS3U6kBKKF5MXGRGLMEacnvDlEak6"
    "DSkbA7QphXu9uZQYHFJCMsTiK4dUfo9U3RmouYUrXNalxJJASnTVIlVeCdScwhWudpcSVD2h1RSucJ3nUkI0aJLCFa5whStc4QpX"
    "uMIVrnCF6wD3hjf8f7yMlmv5SPj+AAAAAElFTkSuQmCC"
)

C = {
    "bg":          "#191919",
    "surface":     "#212121",
    "surface2":    "#2a2a2a",
    "surface3":    "#333333",
    "ink":         "#e6e6e6",
    "ink_muted":   "#999999",
    "ink_faint":   "#4d4d4d",
    "accent":      "#EC7000",
    "accent_dim":  "#3d2a14",
    "accent_soft": "#2a1f12",
    "ok":          "#4ea87a",
    "ok_dim":      "#1a3a2a",
    "warn":        "#d49b45",
    "err":         "#c95f5f",
    "err_dim":     "#3d1515",
    "hair":        "#2a2a2a",
    "log_step":    "#606060",
    "log_ok":      "#4ea87a",
    "log_warn":    "#d49b45",
    "log_err":     "#c95f5f",
}

DOT_COLORS = [
    "#9b9b9b",
    "#a07450",
    "#c87941",
    "#c4a832",
    "#5a9e72",
    "#EC7000",
    "#8b72c9",
    "#c97a9e",
    "#c96060",
]
DOT_LABELS = ["Cinza","Marrom","Laranja","Amarelo","Verde","Azul","Roxo","Rosa","Vermelho"]

ICON_FILENAME = "itaulogo.png"
LOGO_FILENAME = ICON_FILENAME
APP_USER_MODEL_ID = "MesaItau.RiscoSacado"

MESES_PT = (
    "", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
)
DIAS_PT = (
    "segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
    "sexta-feira", "sábado", "domingo",
)


def format_data_pt_br(dt: datetime) -> str:
    return f"{DIAS_PT[dt.weekday()]}, {dt.day} de {MESES_PT[dt.month]} de {dt.year}"


def _parse_brl(raw: str):
    s = (raw or "").strip().replace("R$","").replace("r$","").replace(" ","")
    s = re.sub(r"[^\d,.\-]","",s)
    if not s or s in {".",",","-","-.","-.","-.","-."}:
        return None
    ld, lc = s.rfind("."), s.rfind(",")
    si = max(ld, lc)
    try:
        if si == -1:
            d = Decimal(re.sub(r"[^\d\-]","",s))
        else:
            ip = re.sub(r"[^\d\-]","", s[:si])
            dp = re.sub(r"[^\d]","", s[si+1:])
            if not dp: return None
            d = Decimal((ip if ip not in {"","-"} else "0")+"."+dp)
    except (InvalidOperation, ValueError):
        return None
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def _fmt_brl(d: Decimal) -> str:
    d2 = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    sign = "-" if d2 < 0 else ""
    s = f"{abs(d2):.2f}"; i, f = s.split(".")
    return f"R$ {sign}{'{:,}'.format(int(i)).replace(',','.')},{f}"

def _fmt_brl_from_raw(raw: str) -> str:
    d = _parse_brl(raw)
    return _fmt_brl(d) if d is not None else (raw or "").strip()

def _fmt_brl_plain_web(raw: str) -> str:
    d = _parse_brl(raw)
    if d is None: return ""
    s = f"{abs(d):.2f}"; i, f = s.split(".")
    sign = "-" if d < 0 else ""
    return f"{sign}{'{:,}'.format(int(i)).replace(',','.')},{f}"


def _fmt_date_short(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val or "").strip()
    return s


def _valor_to_decimal(val):
    if val is None:
        return Decimal("0")
    if isinstance(val, Decimal):
        return val.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(val, (int, float)):
        return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    parsed = _parse_brl(str(val))
    return parsed if parsed is not None else Decimal("0")


def _fmt_valor_cell(val) -> str:
    if val is None:
        return "—"
    if isinstance(val, (int, float, Decimal)):
        try:
            return _fmt_brl(Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        except Exception:
            pass
    return _fmt_brl_from_raw(str(val))


def _normalize_sacado_key(name: str) -> str:
    return RE_SPACES.sub(" ", (name or "").strip()).upper()


def _find_invertido_header(ws) -> tuple:
    """Localiza a linha de cabeçalho pelo texto 'Nome Sacado' (formato flexível)."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 30:
            break
        cells = [str(c or "").strip().lower() for c in (row or [])]
        joined = " ".join(cells)
        if "nome sacado" in joined or "nome_sacado" in joined:
            col_map = {}
            for j, cell in enumerate(cells):
                if "doc sacado" in cell or cell == "doc sacado":
                    col_map["doc_sacado"] = j
                elif "doc cedente" in cell or cell == "doc cedente":
                    col_map["doc_cedente"] = j
                elif "nome sacado" in cell or cell == "nome sacado":
                    col_map["nome"] = j
                elif cell in ("número", "numero", "nº", "nf"):
                    col_map["nf"] = j
                elif cell == "valor":
                    col_map["valor"] = j
                elif "inclus" in cell:
                    col_map["inclusao"] = j
                elif "vencimento" in cell:
                    col_map["vencimento"] = j
                elif cell == "prazo":
                    col_map["prazo"] = j
            if "nome" not in col_map:
                col_map["nome"] = 1
            col_map.setdefault("doc_sacado", 0)
            col_map.setdefault("doc_cedente", 2)
            col_map.setdefault("nf", 4)
            col_map.setdefault("valor", 5)
            col_map.setdefault("inclusao", 6)
            col_map.setdefault("vencimento", 7)
            col_map.setdefault("prazo", 8)
            return i, col_map
    return 2, {"doc_sacado": 0, "doc_cedente": 2, "nome": 1, "nf": 4, "valor": 5, "inclusao": 6, "vencimento": 7, "prazo": 8}


def _parse_invertido_xlsx(path: str) -> list:
    if not OPENPYXL_OK:
        raise RuntimeError("Biblioteca openpyxl não disponível. Instale com: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row, cols = _find_invertido_header(ws)
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row:
                continue
            if not row:
                continue
            nome_idx = cols["nome"]
            if nome_idx >= len(row) or row[nome_idx] is None:
                continue
            nome = str(row[nome_idx] or "").strip()
            if not nome:
                continue

            def _cell(key):
                idx = cols.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            rows.append({
                "uid":         len(rows),
                "doc_sacado":  only_digits(str(_cell("doc_sacado") or "")),
                "doc_cedente": only_digits(str(_cell("doc_cedente") or "")),
                "nome_sacado": nome,
                "nf":          str(_cell("nf") or "").strip(),
                "valor_raw":   _valor_to_decimal(_cell("valor")),
                "valor":       _fmt_valor_cell(_cell("valor")),
                "data_inclusao": _fmt_date_short(_cell("inclusao")),
                "data_vencimento": _fmt_date_short(_cell("vencimento")),
                "prazo":       str(_cell("prazo") if _cell("prazo") is not None else "").strip(),
            })
        return rows
    finally:
        wb.close()


def _group_invertido_ops(ops: list) -> list:
    groups = {}
    for op in ops:
        key = _normalize_sacado_key(op["nome_sacado"])
        if key not in groups:
            groups[key] = {
                "nome_sacado": op["nome_sacado"].strip(),
                "doc_sacado":  only_digits(op.get("doc_sacado") or ""),
                "doc_cedente": only_digits(op.get("doc_cedente") or ""),
                "notas": [],
                "total": Decimal("0"),
            }
        if not groups[key]["doc_sacado"] and op.get("doc_sacado"):
            groups[key]["doc_sacado"] = only_digits(op["doc_sacado"])
        if not groups[key]["doc_cedente"] and op.get("doc_cedente"):
            groups[key]["doc_cedente"] = only_digits(op["doc_cedente"])
        groups[key]["notas"].append(op)
        groups[key]["total"] += op.get("valor_raw", Decimal("0"))
    result = []
    for g in groups.values():
        g["count"] = len(g["notas"])
        g["valor_total"] = _fmt_brl(g["total"])
        g["notas"].sort(key=lambda n: (n.get("data_vencimento") or "", n.get("nf") or ""))
        result.append(g)
    result.sort(key=lambda g: g["nome_sacado"].upper())
    return result


INVERTIDO_VALOR_MIN = Decimal("10000")

INVERTIDO_PRAZO_MAX = {
    "RPB": 11,
}

INVERTIDO_PRAZO_EXPECTED = {
    "Transdourada":             (90, 70),
    "Posto Sapucaia":           (15, 10),
    "Auto Posto M Timbozao":    (15, 10),
    "Posto Gasol Timbo III":    (15, 10),
    "Posto Timbozao Itaperuna": (15, 10),
    "Posto Pioneiro":           (15, 10),
    "Mirian Cuiaba":            (10, 7),
    "Mirian Varzea":            (10, 7),
    "Petrocal":                 (10, 7),
    "PetroMix":                 (10, 7),
    "PetroVel":                 (10, 7),
}

INVERTIDO_PRAZO_ALIASES = {
    "POSTO GASOL TIMB III": "Posto Gasol Timbo III",
}


def _invertido_sacado_key(name: str) -> str:
    return _normalize_sacado_key(name)


def _invertido_sacado_matches(nome_sacado: str, rule_name: str) -> bool:
    key = _invertido_sacado_key(nome_sacado)
    alias = INVERTIDO_PRAZO_ALIASES.get(key)
    if alias:
        key = _invertido_sacado_key(alias)
    rule_key = _invertido_sacado_key(rule_name)
    if key == rule_key or rule_key in key or key in rule_key:
        return True
    # Tentativa por palavras significativas (ignora artigos/preposições curtas)
    stop = {"DE","DO","DA","DOS","DAS","E","EM","A","O","AS","OS"}
    key_words = {w for w in key.split() if len(w) > 2 and w not in stop}
    rule_words = {w for w in rule_key.split() if len(w) > 2 and w not in stop}
    if key_words and rule_words and len(key_words & rule_words) >= max(1, len(rule_words) - 1):
        return True
    return False


def _invertido_parse_prazo_days(prazo) -> int | None:
    s = str(prazo if prazo is not None else "").strip()
    if not s:
        return None
    m = re.search(r"\d+", s)
    return int(m.group()) if m else None


def _invertido_check_nf(nf: str) -> str | None:
    s = str(nf or "").strip()
    if not s:
        return "NF ausente ou vazia"
    # Suspeito: contém letra(s) misturada(s) ao número (ex: 002141zz04)
    if re.search(r"[A-Za-z]", s):
        return f"NF contém letra(s) ({s})"
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    inner = digits.lstrip("0")
    if not inner:
        return f"NF inválida — apenas zeros ({s})"
    # Suspeito: muitos zeros à esquerda escondendo um número minúsculo
    # (ex: 001, 000000001) — NF normal com zero à esquerda mas número
    # longo (ex: 0010407103) NÃO é suspeito.
    if digits != inner and len(inner) <= 2:
        return f"NF com zeros à esquerda e número muito curto ({s})"
    # Suspeito: número de apenas 1 dígito, sem zeros à esquerda
    if len(digits) <= 1:
        return f"NF com número suspeito — muito curto ({s})"
    return None


def _invertido_check_valor(valor_raw: Decimal) -> str | None:
    if valor_raw < INVERTIDO_VALOR_MIN:
        return f"Valor abaixo de R$ 10.000,00 ({_fmt_brl(valor_raw)})"
    return None


def _invertido_check_prazo(nome_sacado: str, prazo) -> str | None:
    days = _invertido_parse_prazo_days(prazo)
    if days is None:
        for rule_name in list(INVERTIDO_PRAZO_MAX) + list(INVERTIDO_PRAZO_EXPECTED):
            if _invertido_sacado_matches(nome_sacado, rule_name):
                return "Prazo ausente ou inválido"
        return None

    for rule_name, max_days in INVERTIDO_PRAZO_MAX.items():
        if _invertido_sacado_matches(nome_sacado, rule_name):
            if days > max_days:
                return (f"Prazo de {days} dias excede o máximo de {max_days} dias "
                        f"({rule_name})")
            return None

    for rule_name, (expected, min_ok) in INVERTIDO_PRAZO_EXPECTED.items():
        if _invertido_sacado_matches(nome_sacado, rule_name):
            if days < min_ok:
                return (f"Prazo de {days} dias muito abaixo do esperado "
                        f"({expected} dias — {rule_name})")
            if days > expected:
                return (f"Prazo de {days} dias acima do esperado "
                        f"({expected} dias — {rule_name})")
            return None

    return None


def _invertido_collect_alerts(ops: list) -> list:
    alerts = []
    for idx, op in enumerate(ops):
        motivos = []
        nf_msg = _invertido_check_nf(op.get("nf"))
        if nf_msg:
            motivos.append(nf_msg)
        val_msg = _invertido_check_valor(op.get("valor_raw", Decimal("0")))
        if val_msg:
            motivos.append(val_msg)
        prazo_msg = _invertido_check_prazo(op.get("nome_sacado"), op.get("prazo"))
        if prazo_msg:
            motivos.append(prazo_msg)
        if motivos:
            alerts.append({"index": idx, "op": op, "motivos": motivos})
    return alerts


def _invertido_apply_alert_decisions(ops: list, alerts: list, decisions: dict) -> list:
    rejected = {
        item["index"] for item in alerts
        if decisions.get(item["index"]) == "reject"
    }
    return [op for i, op in enumerate(ops) if i not in rejected]


LIMITE_SOBRA_MIN = 100_000


def _parse_data_curta(s):
    """Converte 'dd/mm/aaaa' (ou variações) em date. Retorna None se inválido."""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _fmt_cnpj(digits: str) -> str:
    d = only_digits(digits or "")
    if len(d) != 14:
        return digits or ""
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


RISCO_SACADO_CEDENTE_NOME = "VIBRA ENERGIA S.A"


def _get_risco_sacado_logo_path() -> str:
    path = os.path.join(tempfile.gettempdir(), "risco_sacado_logo.png")
    if not os.path.isfile(path):
        with open(path, "wb") as f:
            f.write(base64.b64decode(RISCO_SACADO_LOGO_B64))
    return path


def build_risco_sacado_email_html(sacado_nome: str, sacado_cnpj: str,
                                   cedente_cnpj: str, notas: list,
                                   taxa_str: str) -> str:
    """Monta o HTML do e-mail 'RISCO SACADO INVERTIDO' no mesmo layout do
    modelo padrão da mesa, com uma linha por nota e subtotal ao final.
    Cada item de `notas` deve conter: nf, data_vencimento, valor_raw,
    valor_liquido (Decimal ou None)."""
    cedente_nome = RISCO_SACADO_CEDENTE_NOME
    cedente_cnpj_fmt = _fmt_cnpj(cedente_cnpj)
    sacado_cnpj_fmt = _fmt_cnpj(sacado_cnpj)

    total_face = Decimal("0")
    total_liq = Decimal("0")
    linhas_html = []
    for op in notas:
        vf = op.get("valor_raw") or Decimal("0")
        vl = op.get("valor_liquido")
        total_face += vf
        if vl is not None:
            total_liq += vl
        linhas_html.append(f"""
<tr>
  <td style="{_TD_STYLE}">{cedente_nome}</td>
  <td style="{_TD_STYLE}">{cedente_cnpj_fmt}</td>
  <td style="{_TD_STYLE}">{sacado_nome}</td>
  <td style="{_TD_STYLE}">{sacado_cnpj_fmt}</td>
  <td style="{_TD_STYLE}">{op.get('nf') or '—'}</td>
  <td style="{_TD_STYLE}">{op.get('data_vencimento') or '—'}</td>
  <td style="{_TD_STYLE}">{_fmt_brl(vf)}</td>
  <td style="{_TD_STYLE}">{_fmt_brl(vl) if vl is not None else '—'}</td>
  <td style="{_TD_STYLE}">{taxa_str or '—'}</td>
</tr>""")

    subtotal_html = f"""
<tr style="background:#F0F0F0">
  <td style="{_TD_STYLE}"><b>Subtotal</b></td>
  <td style="{_TD_STYLE}"></td>
  <td style="{_TD_STYLE}"></td>
  <td style="{_TD_STYLE}"></td>
  <td style="{_TD_STYLE}"></td>
  <td style="{_TD_STYLE}"></td>
  <td style="{_TD_STYLE}"><b>{_fmt_brl(total_face)}</b></td>
  <td style="{_TD_STYLE}"><b>{_fmt_brl(total_liq)}</b></td>
  <td style="{_TD_STYLE}"></td>
</tr>"""

    return f"""<html><head><meta charset="utf-8"></head>
<body style="font-family:'Itau Display',Calibri,sans-serif;color:#333333">
<table cellspacing="3" cellpadding="0" style="max-width:768pt">
<tr>
  <td rowspan="3" style="width:100px;padding:0"><img src="cid:risco_sacado_logo" width="110" height="100" alt="Icone de titulos"></td>
  <td></td>
</tr>
<tr><td><span style="font-size:22.5pt;font-family:'Itau Display',serif;color:#EC7000;font-weight:bold">Notas para antecipação - Risco Sacado Invertido</span></td></tr>
<tr><td><span style="font-size:16.5pt;font-family:'Itau Display',serif;color:#9A9A9A;font-weight:bold">Risco Sacado</span></td></tr>
</table>
<p>&nbsp;</p>
<p>Seguem as notas que foram disponibilizadas nesta data para antecipação.</p>
<table cellspacing="0" cellpadding="0" width="100%" style="border-collapse:collapse">
<thead>
<tr>
  <td style="{_TH_STYLE}">RAZAO SOCIAL CEDENTE</td>
  <td style="{_TH_STYLE}">CNPJ CEDENTE</td>
  <td style="{_TH_STYLE}">RAZAO SOCIAL SACADO</td>
  <td style="{_TH_STYLE}">CNPJ SACADO</td>
  <td style="{_TH_STYLE}">NRO NF/FATURA</td>
  <td style="{_TH_STYLE}">VENCIMENTO</td>
  <td style="{_TH_STYLE}">VALOR NF/FATURA</td>
  <td style="{_TH_STYLE}">VALOR LÍQUIDO</td>
  <td style="{_TH_STYLE}">TX (% AM)</td>
</tr>
</thead>
<tbody>
{''.join(linhas_html)}
{subtotal_html}
</tbody>
</table>
<p>&nbsp;</p>
<p>&quot;Operação é paga ao Fornecedor na conta indicada.</p>
<p>Lembrando que operações antecipadas/pagas não são passíveis de cancelamento. No bankline terá acesso as notas antecipadas através da Rota - Mais Serviços - Consulta de Notas Negociadas</p>
<p>&nbsp;</p>
<p><span style="font-size:16.5pt;font-family:'Itau Display',serif;color:#9A9A9A;font-weight:bold">Bons Negócios,</span></p>
<p><span style="font-size:16.5pt;font-family:'Itau Display',serif;color:#EC7000;font-weight:bold">Isso é Risco Sacado</span></p>
<p style="font-family:Calibri;font-size:9pt;color:#000000;margin-top:20px">Corporativo | Interno</p>
</body></html>"""


_TH_STYLE = ("background:#7EB3F0;padding:4pt;font-family:'Itau Display',serif;"
             "font-size:7.5pt;font-weight:bold;color:#1B120A;text-align:center")
_TD_STYLE = ("border:1pt solid #FAF7F5;padding:4pt;font-family:'Itau Display',serif;"
             "font-size:7.5pt;color:#333333;text-align:center")


def enviar_email_outlook_risco_sacado(subject: str, html_body: str):
    """Abre um popup do Outlook (sem enviar automaticamente) já preenchido
    com título, corpo e imagem embutida, para revisão e envio manual."""
    if not WIN32_OK:
        raise RuntimeError(
            "Integração com Outlook (win32com) não está disponível neste ambiente.")
    did_init = False
    if PYTHONCOM_OK:
        try:
            pythoncom.CoInitialize()
            did_init = True
        except Exception:
            did_init = False
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # olMailItem
        mail.Subject = subject
        img_path = _get_risco_sacado_logo_path()
        attachment = mail.Attachments.Add(img_path)
        try:
            attachment.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001E",
                "risco_sacado_logo")
        except Exception:
            pass
        mail.HTMLBody = html_body
        mail.Display(False)
    finally:
        if did_init:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def calcular_valor_liquido(valor_face: Decimal, taxa_pct_str: str, dias_prazo: int) -> Decimal:
    """Calcula o valor líquido de uma antecipação por desconto comercial
    simples: VL = VF x (1 - taxa x prazo/30), com taxa mensal em % (ex.:
    '1,3950') e prazo em dias corridos a partir de hoje até o vencimento.
    Prazo negativo (nota já vencida) é tratado como zero (sem desconto)."""
    if valor_face is None or not taxa_pct_str:
        return None
    try:
        taxa = Decimal(taxa_pct_str.replace(",", ".")) / Decimal("100")
    except Exception:
        return None
    prazo = max(dias_prazo, 0)
    fator = Decimal("1") - (taxa * Decimal(prazo) / Decimal("30"))
    return (valor_face * fator).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _evaluate_limite_operacao(montante: Decimal, limite_data: dict | None):
    if not limite_data:
        return "nao_validado", "Limites não validados"
    state = limite_data.get("state")
    if state == "processing":
        return "validando", "Validando…"
    if state in ("error", "ltc_expired"):
        return "insuficiente", "Limite insuficiente"
    limite = limite_data.get("limite_disp")
    if limite is None or state not in ("ok", "warn"):
        return "nao_validado", "Limites não validados"
    mont = int(montante.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if mont > limite:
        return "insuficiente", "Limite insuficiente"
    if (limite - mont) < LIMITE_SOBRA_MIN:
        return "quase", "Limite quase lá"
    return "ok", "Limite OK"


def _fmt_limite_int(val) -> str:
    if val is None:
        return "N/D"
    return f"R$ {int(val):,}".replace(",", ".")

BPM_CLIENT_DATA = {
    "Transdourada":             {"CNPJ":"01259730000174","PLATAFORMA":"2939","AG":"1643","CONTA":"99451-8"},
    "RPB":                      {"CNPJ":"07075892000139","PLATAFORMA":"8973","AG":"6627","CONTA":"06471-7"},
    "Posto Arinos":             {"CNPJ":"05798923000154","PLATAFORMA":"8250","AG":"1364","CONTA":"98355-9"},
    "Brasnorte":                {"CNPJ":"00514301000133","PLATAFORMA":"8250","AG":"1364","CONTA":"98654-5"},
    "Mirian Varzea":            {"CNPJ":"16519674000137","PLATAFORMA":"8250","AG":"1689","CONTA":"05136-3"},
    "Mirian Cuiaba":            {"CNPJ":"41240105000103","PLATAFORMA":"8250","AG":"1689","CONTA":"58145-0"},
    "Petrocal":                 {"CNPJ":"12781233000158","PLATAFORMA":"7948","AG":"8251","CONTA":"44190-6"},
    "Posto Sapucaia":           {"CNPJ":"22787055000126","PLATAFORMA":"0352","AG":"1334","CONTA":"57853-9"},
    "PetroMix":                 {"CNPJ":"05684913000198","PLATAFORMA":"7948","AG":"8251","CONTA":"99886-3"},
    "Auto Posto M Timbozao":    {"CNPJ":"04632746000179","PLATAFORMA":"0352","AG":"8296","CONTA":"18100-4"},
    "PetroVel":                 {"CNPJ":"01294927000144","PLATAFORMA":"7948","AG":"8251","CONTA":"99887-1"},
    "Posto Gasol Timbo III":    {"CNPJ":"32179707000101","PLATAFORMA":"0352","AG":"8296","CONTA":"06655-1"},
    "Posto Timbozao Itaperuna": {"CNPJ":"25032853000136","PLATAFORMA":"0352","AG":"1334","CONTA":"49413-3"},
    "Posto Pioneiro":           {"CNPJ":"23184831000166","PLATAFORMA":"0352","AG":"5255","CONTA":"12888-5"},
}

LIMITE_CLIENT_URLS = {
    "Transdourada":  "https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1R10K6&tppes=J&subgrupo=01259730000174&plataforma=2939&ambiente=PRD#/omni/",
    "RPB":           "https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1D1AK3&tppes=J&subgrupo=26727190000137&plataforma=8973&ambiente=PRD#/omni/",
    "Posto Arinos":  "https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1H1UK3&tppes=J&subgrupo=00514301000133&plataforma=8250&ambiente=PRD",
    "Mirian Varzea": "https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1H1UK3&tppes=J&subgrupo=16519674000137&plataforma=8250&ambiente=PRD",
    "Petrocal":      "https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1H1UK3&tppes=J&subgrupo=12781233000158&plataforma=7948&ambiente=PRD",
    "Posto Sapucaia":"https://digital.itau/CF_Digital/IntegracaoQK/IntegracaoQK/redir.aspx?usuario=987400146&senha=1H1UK3&tppes=J&subgrupo=32179707000101&plataforma=0352&ambiente=PRD#/omni/",
}

LIMITE_SHARED_RESULTS = {
    "Posto Arinos":   ["Brasnorte"],
    "Mirian Varzea":  ["Mirian Cuiaba"],
    "Petrocal":       ["PetroMix","PetroVel"],
    "Posto Sapucaia": ["Auto Posto M Timbozao","Posto Gasol Timbo III","Posto Timbozao Itaperuna","Posto Pioneiro"],
}

MAPPED_CLIENTS = {"Transdourada","RPB","Posto Arinos","Mirian Varzea","Petrocal","Posto Sapucaia"}
MIRROR_CLIENTS = {
    "Brasnorte":"Posto Arinos","Mirian Cuiaba":"Mirian Varzea",
    "PetroMix":"Petrocal","PetroVel":"Petrocal",
    "Auto Posto M Timbozao":"Posto Sapucaia","Posto Gasol Timbo III":"Posto Sapucaia",
    "Posto Timbozao Itaperuna":"Posto Sapucaia","Posto Pioneiro":"Posto Sapucaia",
}

REGIAO_TRADER_ESPEC = {
    "21":("Matheus","Lucas"),"22":("Debora","Renata"),
    "23":("Thiago","Paula"),"24":("Thiago","Luiz Gustavo"),
    "25":("Giovana","Lucas"),"26":("Gabriel","Guga"),
    "28":("Rafael","Paula"),"30":("Adriana/Rafael","Luiz Gustavo"),
    "32":("Debora","Renata"),"33":("Giovana","Lucas"),
    "01":("Matheus/Gabriel",""),"02":("Gabriel/Gabriel",""),
}

RE_SPACES        = re.compile(r"\s+")
RE_CNPJ          = re.compile(r"(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})")
RE_CNPJ_LABEL    = re.compile(r"CNPJ[^\d]*([\d./-]{14,20})", re.IGNORECASE)
RE_VALOR_1       = re.compile(r"Valor\s+da\s+Opera[cç][aã]o[^\d]*(R\$\s*[\d\.,]+)", re.IGNORECASE)
RE_VALOR_2       = re.compile(r"R\$\s*[\d\.]{1,},\d{2}")
RE_VALOR_3       = re.compile(r"(R\$\s*[\d\.,]{3,})")
RE_PLATAFORMA    = re.compile(r"Plataforma\D{0,20}(\d{4})", re.IGNORECASE)
RE_REGIAO_PLAT   = re.compile(r"Regi[aã]o(?:\s+da)?\s+Plataforma\D{0,20}(\d{2})", re.IGNORECASE)
RE_REGIAO        = re.compile(r"Regi[aã]o\D{0,20}(\d{2})", re.IGNORECASE)
RE_SPREAD_1      = re.compile(r"(?:Taxa\/Spread|Spread)\D{0,25}([\d\.,]+)\s*%?", re.IGNORECASE)
RE_SPREAD_2      = re.compile(r"Spread\s+m[íi]nimo\D{0,25}([\d\.,]+)\s*%?", re.IGNORECASE)
RE_PRAZO_MIN_1   = re.compile(r"Prazo\s+M[ií]n[ií]mo\s+NF\D{0,25}(\d{1,4})", re.IGNORECASE)
RE_PRAZO_MAX_1   = re.compile(r"Prazo\s+M[áa]ximo\s+NF\D{0,25}(\d{1,4})", re.IGNORECASE)
RE_MODALIDADE_1  = re.compile(r"Modalidade[:\s]+([^\n\.;]+)", re.IGNORECASE)
RE_AUTORIZADAS_CSV   = re.compile(r"\bautorizadas\s+csv\b", re.IGNORECASE)
RE_AUTORIZADAS_SIS   = re.compile(r"\bautorizadas\s+sispag\b", re.IGNORECASE)
RE_RAZAO_1       = re.compile(r"Raz[aã]o\s+Social[:\s]*([^\n]+)", re.IGNORECASE)
RE_RAZAO_ATE     = re.compile(r"Raz[aã]o\s+Social\s*[:\s]*\s*(.+?)(?=\n\s*(?:CNPJ|Plataforma|Modalidade|Regi[aã]o|Conta\s+Corrente|Valor\s+da|Spread|Prazo)\b)", re.IGNORECASE|re.DOTALL)
RE_CONTA_AG_CC   = re.compile(r"\b(\d{4})\s*[/\s]\s*(\d{3,10}(?:[-‑–—−]\d)?)\b")
RE_CONTA_LABEL   = re.compile(r"conta\s+corrente(?:\s+do\s+cliente)?\s*[:\s-]*", re.IGNORECASE)
RE_LIQ_CRED      = re.compile(r"Cr[ée]dito\s+em\s+CC", re.IGNORECASE)
RE_PREMIO        = re.compile(r"com\s+pr[êe]mio", re.IGNORECASE)

def app_base_dir():
    return os.path.dirname(sys.executable) if getattr(sys,"frozen",False) else os.path.dirname(os.path.abspath(__file__))

def resource_path(p):
    return os.path.join(getattr(sys,"_MEIPASS",app_base_dir()), p)

def only_digits(s):
    return re.sub(r"\D","",s or "")

def _bind_digits_only(entry, var, max_len, min_len=None, hint_lbl=None,
                       valid_lengths=None, hint_text=None):
    """Restringe um Entry a aceitar apenas dígitos, até max_len caracteres.
    Se hint_lbl for informado, mostra um aviso minimalista ao perder o foco
    quando o tamanho não corresponde ao exigido (min_len, ou a um dos
    valid_lengths quando especificado — ex.: CPF/CNPJ: 11 ou 14)."""
    def _check_len():
        if hint_lbl is None:
            return
        digits = only_digits(var.get())
        if not digits:
            hint_lbl.configure(text=" ")
            return
        if valid_lengths:
            ok = len(digits) in valid_lengths
        elif min_len is not None and min_len == max_len:
            ok = len(digits) == max_len
        else:
            ok = len(digits) >= (min_len or max_len)
        hint_lbl.configure(text=(hint_text or "número irregular") if not ok else " ")

    def _on_key(e):
        if e.keysym in ("BackSpace","Delete","Tab","Return","Left","Right","Home","End"):
            return
        if e.state & 0x4:
            return
        if e.char and e.char.isdigit():
            if len(only_digits(var.get())) >= max_len and not entry.selection_present():
                return "break"
            return
        if e.char:
            return "break"
    def _on_paste(_e):
        try:
            clip = entry.clipboard_get()
        except Exception:
            return "break"
        var.set(only_digits(clip)[:max_len])
        return "break"
    entry.bind("<KeyPress>", _on_key)
    entry.bind("<<Paste>>", _on_paste)
    entry.bind("<Control-v>", _on_paste)
    if hint_lbl is not None:
        entry.bind("<FocusOut>", lambda _e: _check_len())

def _bind_conta_field(entry, var, hint_lbl):
    """Campo de conta no formato 99999-9 (5 dígitos + dígito verificador).
    O usuário só digita números; o hífen é inserido automaticamente, e um
    aviso minimalista aparece enquanto faltar o dígito verificador."""
    state = {"updating": False}

    def _format(digits):
        digits = digits[:6]
        if len(digits) <= 5:
            return digits
        return f"{digits[:5]}-{digits[5:]}"

    def _refresh_hint(digits):
        if digits and len(digits) < 6:
            hint_lbl.configure(text="falta o dígito (ex.: -8)")
        else:
            hint_lbl.configure(text=" ")

    def _on_keyrelease(_e=None):
        if state["updating"]:
            return
        digits = only_digits(var.get())[:6]
        state["updating"] = True
        var.set(_format(digits))
        entry.icursor("end")
        state["updating"] = False
        hint_lbl.configure(text=" ")

    def _on_focus_out(_e=None):
        digits = only_digits(var.get())[:6]
        _refresh_hint(digits)

    def _on_key(e):
        if e.keysym in ("BackSpace","Delete","Tab","Return","Left","Right","Home","End"):
            entry.after_idle(_on_keyrelease)
            return
        if e.state & 0x4:
            return
        if e.char and e.char.isdigit():
            if len(only_digits(var.get())) >= 6 and not entry.selection_present():
                return "break"
            entry.after_idle(_on_keyrelease)
            return
        if e.char:
            return "break"

    def _on_paste(_e):
        try:
            clip = entry.clipboard_get()
        except Exception:
            return "break"
        digits = only_digits(clip)[:6]
        state["updating"] = True
        var.set(_format(digits))
        state["updating"] = False
        hint_lbl.configure(text=" ")
        return "break"

    entry.bind("<KeyPress>", _on_key)
    entry.bind("<<Paste>>", _on_paste)
    entry.bind("<Control-v>", _on_paste)
    entry.bind("<FocusOut>", _on_focus_out)

LIMITE_INVERTIDO_CNPJS = frozenset(
    only_digits(v["CNPJ"]) for v in BPM_CLIENT_DATA.values() if v.get("CNPJ")
)

def normalize_text_variants(t):
    return RE_SPACES.sub(" ",t).strip(), re.sub(r"\s+","",t or "")

def _plain_label_value_map(pdf_path):
    """Extração geométrica genérica para PDFs em formato de tabela (páginas
    do Salesforce exportadas): os campos aparecem em duas colunas fixas
    (esquerda/direita), com uma linha de rótulos seguida da linha de
    valores logo abaixo, ambos alinhados pela mesma posição X inicial da
    coluna. Detecta as duas posições X mais comuns onde rótulos começam
    (colunas esquerda e direita) e usa essas posições como "trilhos" para
    juntar cada rótulo com seu valor na linha seguinte, evitando pegar
    texto de colunas vizinhas (ex.: barra lateral direita com links).
    Retorna {} se pdfplumber não estiver disponível ou nada for reconhecido."""
    if pdfplumber is None or not pdf_path:
        return {}
    out = {}
    # Rótulos conhecidos do formato "Operação MN / Risco Sacado" do
    # Salesforce. Usados para localizar as linhas de rótulo com confiança,
    # em vez de tentar adivinhar heuristicamente qualquer linha.
    known_labels = [
        "Razão Social", "CNPJ", "Conta Corrente do Cliente",
        "Número da Solicitação", "Evento", "Plataforma",
        "Região da Plataforma", "Região", "Dac Plataforma", "Código do Produto",
        "Número do Contrato", "Data da Devolução", "Data de Início",
        "Data de Vencimento", "Valor da Operação", "Carência",
        "Prazo", "Tipo de Prazo", "Taxa/Spread", "Tipo de Taxa",
        "Data Valor", "Planilha Flex", "Duplo Sim", "SubCarteira",
        "Produto", "Prazo Minimo NF", "Prazo Maximo NF",
        "Tarifa Titulo", "Valor Tarifa", "Spread Minimo", "Spread Rebate",
        "Meio de Transmissão", "Operador Master", "Forma Liquidação",
        "Modalidade", "Valor Tarifa Convênio", "Ganho Financeiro",
        "Percentual Rebate", "Nome Contato Técnico",
        "Telefone Contato Técnico", "Email Contato Técnico",
    ]
    label_re = re.compile(
        "|".join(re.escape(l) for l in sorted(known_labels, key=len, reverse=True)),
        re.IGNORECASE)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for pg in pdf.pages:
                try:
                    words = pg.extract_words()
                except Exception:
                    continue
                if not words:
                    continue
                rows = _word_rows_from_pdf_page(words)
                for i, row in enumerate(rows):
                    line = " ".join(w["text"] for w in row)
                    # Só considera linhas que batem >=1 rótulo conhecido no
                    # começo de algum "segmento" — isto é, colunas cujo
                    # primeiro token corresponde ao início de um rótulo.
                    matches = list(label_re.finditer(line))
                    if not matches:
                        continue
                    # Mapeia cada match de rótulo para a posição X inicial
                    # da palavra correspondente na linha (usa a primeira
                    # palavra do match).
                    segs = []
                    for m in matches:
                        char_pos = 0
                        x0 = None
                        for w in row:
                            wl = len(w["text"])
                            if char_pos <= m.start() < char_pos + wl + 1:
                                x0 = w["x0"]; break
                            char_pos += wl + 1
                        if x0 is None:
                            continue
                        segs.append((m.group(0), x0))
                    segs.sort(key=lambda s: s[1])
                    if not segs:
                        continue
                    x_min = segs[0][1]
                    x_max = segs[-1][1]
                    base_top = row[0]["top"]
                    # Procura a próxima linha (dentro de uma janela vertical
                    # razoável) que tenha ao menos uma palavra dentro do
                    # intervalo X coberto pelos rótulos desta linha — pula
                    # linhas "estranhas" que só têm conteúdo de outra coluna
                    # (ex.: texto da barra lateral direita).
                    vrow = None
                    for cand_row in rows[i + 1:]:
                        if cand_row[0]["top"] - base_top > 30:
                            break
                        if any(x_min - 8 <= w["x0"] <= x_max + 200 for w in cand_row):
                            vrow = cand_row
                            break
                    if vrow is None:
                        continue
                    for k, (label_txt, x0) in enumerate(segs):
                        x1 = segs[k + 1][1] if k + 1 < len(segs) else float("inf")
                        cand = [w for w in vrow if x0 - 8 <= w["x0"] < x1 - 2]
                        if not cand:
                            continue
                        val = " ".join(w["text"] for w in sorted(cand, key=lambda w: w["x0"]))
                        key = label_txt.strip().lower()
                        if key and key not in out:
                            out[key] = val.strip()
    except Exception:
        return out
    return out


def _plain_map_lookup(pmap, *label_variants):
    """Procura o valor de um rótulo no mapa posicional, tolerando pequenas
    variações de grafia (acentuação, maiúsculas)."""
    for variant in label_variants:
        v = pmap.get(variant.lower())
        if v:
            return v
    return None


def extract_text_from_pdf(p):
    # pdfplumber é usado como extrator principal: o PyPDF2 tem um bug
    # conhecido de espaçamento em PDFs com posicionamento de caractere fino
    # (comum em páginas do Salesforce exportadas), quebrando palavras no
    # meio (ex.: "RECON PR OMOC OES EVENT OS EIRELI"). O pdfplumber usa as
    # posições reais dos caracteres e não sofre desse problema.
    if pdfplumber is not None:
        try:
            lo, pl = [], []
            with pdfplumber.open(p) as pdf:
                for pg in pdf.pages:
                    try: lo.append(pg.extract_text(layout=True) or "")
                    except Exception: lo.append("")
                    try: pl.append(pg.extract_text() or "")
                    except Exception: pl.append(lo[-1])
            return "\n".join(lo), "\n".join(pl)
        except Exception:
            pass
    # Fallback: PyPDF2 (apenas se pdfplumber não estiver instalado/falhar).
    if PdfReader is None:
        return "", ""
    r = PdfReader(p, strict=False)
    lo, pl = [], []
    for pg in r.pages:
        try:
            try: lo.append(pg.extract_text(extraction_mode="layout") or "")
            except TypeError: lo.append(pg.extract_text() or "")
        except: lo.append("")
        try: pl.append(pg.extract_text() or "")
        except: pl.append(lo[-1])
    return "\n".join(lo), "\n".join(pl)

def normalize_modalidade(m):
    s = RE_SPACES.sub(" ", (m or "").strip()); sl = s.lower().replace(" ","")
    if "sispag" in sl: return "Autorizadas SISPAG"
    if RE_AUTORIZADAS_SIS.search(s): return "Autorizadas SISPAG"
    if RE_AUTORIZADAS_CSV.search(s): return "Autorizadas CSV"
    if re.search(r"\bsispag\b",s,re.I): return "Autorizadas SISPAG"
    if re.search(r"\bcsv\b",s,re.I): return "Autorizadas CSV"
    return s

def infer_troca(m):
    ml = (m or "").lower().replace(" ","")
    if "sispag" in ml: return "Sispag"
    return "CSV"

def normalize_percent_br(p):
    if not p: return ""
    v = p.strip().replace(" ","").replace(",",".")
    try:
        n = float(re.sub(r"[^\d\.]","",v))
        return f"{int(n)},00% a.a." if n.is_integer() else f"{n:.2f}".replace(".",",")+"%  a.a."
    except:
        v = p.strip().replace(".",",")
        if not v.endswith("%"): v += "%"
        return v+("" if "a.a" in v.lower() else " a.a.")

def trader_espec_from_regiao(reg):
    d = only_digits(reg)
    if len(d) >= 2:
        return REGIAO_TRADER_ESPEC.get(d[-2:], ("",""))
    m = RE_REGIAO_PLAT.search(reg) or RE_REGIAO.search(reg)
    if m: return REGIAO_TRADER_ESPEC.get(m.group(1),("",""))
    return ("","")

def sanitize_razao(s):
    if not s: return s
    s = RE_SPACES.sub(" ", s).strip()
    s = re.sub(r"\s*\(\s*/[^)]+\)","",s)
    s = re.split(r"\bCNPJ\b",s,maxsplit=1,flags=re.I)[0]
    s = re.sub(r"\s*\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\s*$","",s)
    return RE_SPACES.sub(" ",s).strip(" :-–—.,;")

def _razao_stop(ln):
    s = (ln or "").strip().lower()
    if not s: return False
    return bool(re.match(r"^\s*(?:cnpj|plataforma|modalidade|regi[aã]o|conta\s+corrente|valor\s+da\s+opera|spread|prazo|liquida)\b",s)) \
           or (len(only_digits(s))>=14 and len(s)<40)

def _word_rows_from_pdf_page(words, y_tol=3):
    """Agrupa palavras (dicts do pdfplumber com 'top'/'x0') em linhas
    visuais, tolerando pequena variação vertical entre glifos da mesma
    linha, e ordena cada linha por posição horizontal."""
    rows = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        placed = False
        for row in rows:
            if abs(row[0]["top"] - w["top"]) <= y_tol:
                row.append(w); placed = True; break
        if not placed:
            rows.append([w])
    for row in rows:
        row.sort(key=lambda w: w["x0"])
    rows.sort(key=lambda row: row[0]["top"])
    return rows

def extract_razao_social_by_columns(pdf_path):
    """Extração geométrica para PDFs em formato de tabela (ex.: páginas
    exportadas do Salesforce), onde os rótulos ficam todos numa linha e os
    valores na linha logo abaixo, alinhados por coluna (posição X). Localiza
    a linha que contém 'Razão Social' seguido de 'CNPJ' (assinatura única
    desse formato) e lê o valor correspondente na linha de dados abaixo,
    delimitado pela coluna do CNPJ. Retorna None se o padrão não for
    encontrado (PDF em outro formato) — quem chama cai no método por regex."""
    if pdfplumber is None or not pdf_path:
        return None
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for pg in pdf.pages:
                try:
                    words = pg.extract_words()
                except Exception:
                    continue
                if not words:
                    continue
                rows = _word_rows_from_pdf_page(words)
                for i, row in enumerate(rows):
                    texts = [w["text"] for w in row]
                    for j in range(len(texts) - 1):
                        if not (texts[j].lower().startswith("raz")
                                and texts[j+1].lower().startswith("social")):
                            continue
                        label_x0 = row[j]["x0"]
                        stop_x0 = None
                        for k in range(j+2, len(texts)):
                            if texts[k].strip(":").lower() == "cnpj":
                                stop_x0 = row[k]["x0"]; break
                        if stop_x0 is None:
                            continue
                        base_top = row[0]["top"]
                        for vrow in rows[i+1:i+4]:
                            if vrow[0]["top"] - base_top > 25:
                                break
                            cand = [w for w in vrow if label_x0-5 <= w["x0"] < stop_x0-2]
                            if cand:
                                val = " ".join(w["text"] for w in sorted(cand, key=lambda w: w["x0"]))
                                return sanitize_razao(val)
    except Exception:
        return None
    return None

def extract_razao_social(t, lines, tn, tc, t_plain=None, pdf_path=None):
    by_cols = extract_razao_social_by_columns(pdf_path)
    if by_cols and len(by_cols) >= 2:
        return by_cols
    for blk in ([t_plain] if t_plain else []) + [t]:
        if not blk: continue
        m = re.search(r"raz[aã]o\s+social\s*[:\s]*",blk,re.I)
        if m:
            rest = blk[m.end():]; end = len(rest)
            for pat in [re.compile(r"\bCNPJ\b",re.I), RE_CNPJ,
                        re.compile(r"\bPlataforma\b",re.I), re.compile(r"\bModalidade\b",re.I)]:
                rm = pat.search(rest)
                if rm and rm.start() < end: end = rm.start()
            chunk = rest[:end].strip()
            if "\n\n" in chunk: chunk = chunk.split("\n\n")[0].strip()
            if len(chunk) >= 2: return sanitize_razao(chunk)
    return None

def extract_cnpj(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "cnpj")
        if v and only_digits(v):
            return only_digits(v)
    m = RE_CNPJ_LABEL.search(t)
    if m: return m.group(1).strip()
    m = RE_CNPJ.search(t)
    if m: return m.group(1).strip()
    m = re.search(r"cnpj.*?(\d{14})",tc,re.I)
    if m: return m.group(1)
    return None

def extract_conta_corrente(lines, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "conta corrente do cliente", "conta corrente")
        if v:
            # remove ruído de link tipo "(/light…" que pode ter colado
            v = re.sub(r"\(/[^\)]*$", "", v).strip()
            m = RE_CONTA_AG_CC.search(v)
            if m:
                return f"{m.group(1)} / {m.group(2)}"
            if v and not _razao_stop(v):
                return v
    for hay in (tn, tc):
        if not hay: continue
        m = RE_CONTA_AG_CC.search(hay)
        if m: return f"{m.group(1)} / {m.group(2)}"
    for i,line in enumerate(lines):
        if not re.search(r"conta\s+corrente",line,re.I): continue
        mm = RE_CONTA_LABEL.search(line)
        if not mm: continue
        rest = line[mm.end():].strip()
        if rest and not _razao_stop(rest): return rest
        if i+1 < len(lines):
            nxt = lines[i+1].strip()
            if nxt and not _razao_stop(nxt): return nxt
    return None

def extract_plataforma(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "plataforma")
        d = only_digits(v or "")
        if len(d) == 4:
            return d
    m = RE_PLATAFORMA.search(tn) or RE_PLATAFORMA.search(t)
    return m.group(1) if m else None

def extract_regiao(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "região da plataforma", "regiao da plataforma", "região")
        d = only_digits(v or "")
        if 1 <= len(d) <= 2:
            return d.zfill(2)
    m = RE_REGIAO_PLAT.search(tn) or RE_REGIAO_PLAT.search(t)
    if m: return m.group(1)
    m = RE_REGIAO.search(tn) or RE_REGIAO.search(t)
    return m.group(1) if m else None

def _reconstruct_interleaved_currency(raw):
    """Alguns exports do Salesforce sobrepõem dois campos de texto na mesma
    posição da página (ex.: 'Valor da Operação' e 'Omni-Channel (offline)'),
    resultando numa string com os caracteres intercalados, tipo
    'OmnRi$-C 1h.a0n0n0e.0l 0(o0f,f0li0ne)'. Isolando apenas os dígitos e
    separadores decimais dessa string, o valor numérico original é
    recuperado de forma confiável."""
    if not raw:
        return None
    kept = re.findall(r"[\d.,]", raw)
    if not kept:
        return None
    digits_str = "".join(kept)
    if not re.search(r"\d", digits_str):
        return None
    # Exige ao menos um separador de milhar/decimal para reduzir falsos
    # positivos em strings puramente numéricas legítimas.
    if "," not in digits_str and "." not in digits_str:
        return None
    return f"R$ {digits_str}"

def extract_valor(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "valor da operação", "valor da operacao")
        if v:
            # Detecta corrupção por sobreposição de texto: quando letras e
            # dígitos aparecem intercalados na mesma palavra (padrão típico
            # desse bug do Salesforce), um valor monetário legítimo nunca
            # teria letras misturadas character a character nos dígitos.
            has_letters = bool(re.search(r"[A-Za-z]", v))
            clean_match = re.fullmatch(r"\s*R?\$?\s*[\d\.,]+\s*", v)
            if not has_letters or clean_match:
                m = re.search(r"R?\$?\s*[\d\.,]+", v)
                if m and re.search(r"\d", m.group(0)):
                    cand = m.group(0).strip()
                    if not cand.startswith("R$"):
                        cand = "R$ " + cand.lstrip("R$").strip()
                    return cand
            # Campo corrompido por sobreposição de texto (ex.: intercalado
            # com "Omni-Channel (offline)") — tenta reconstruir.
            rec = _reconstruct_interleaved_currency(v)
            if rec:
                return rec
    for pat in [RE_VALOR_1, RE_VALOR_2, RE_VALOR_3]:
        m = pat.search(tn)
        if m: return (m.group(1) if pat is RE_VALOR_1 else m.group(0)).strip()
    return None

def extract_spread(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "taxa/spread", "spread")
        if v:
            m = re.search(r"[\d\.,]+", v)
            if m: return m.group(0).strip()
    m = RE_SPREAD_1.search(tn) or RE_SPREAD_2.search(tn)
    return m.group(1).strip() if m else None

def extract_prazo_min(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "prazo minimo nf", "prazo mínimo nf")
        if v:
            m = re.search(r"\d+", v)
            if m: return m.group(0)
    m = RE_PRAZO_MIN_1.search(tn)
    return m.group(1) if m else None

def extract_prazo_max(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "prazo maximo nf", "prazo máximo nf")
        if v:
            m = re.search(r"\d+", v)
            if m: return m.group(0)
    m = RE_PRAZO_MAX_1.search(tn)
    return m.group(1) if m else None

def extract_modalidade(t, tn, tc, pmap=None):
    if pmap:
        v = _plain_map_lookup(pmap, "modalidade")
        if v:
            return normalize_modalidade(v.strip())
    for hay in (tn, t, tc):
        if not hay: continue
        m = RE_MODALIDADE_1.search(hay)
        if m: return normalize_modalidade(m.group(1).strip())
    return None

class BPMUserCancelled(Exception): pass



class ThreadSafeUIMixin:
    """Permite chamar `self._ui(fn)` de threads de worker sem tocar o Tcl
    diretamente fora da thread principal — evita o crash de GIL/Fatal
    Python error ao mover a janela enquanto uma rotina está em execução."""

    def _init_ui_queue(self):
        self._ui_q = queue.Queue()
        self._poll_ui_queue()

    def _ui(self, fn):
        try:
            self._ui_q.put_nowait(fn)
        except Exception:
            pass

    def _poll_ui_queue(self):
        try:
            while True:
                fn = self._ui_q.get_nowait()
                try:
                    fn()
                except Exception:
                    pass
        except queue.Empty:
            pass
        try:
            if self.winfo_exists():
                self.after(50, self._poll_ui_queue)
        except Exception:
            pass

def make_hairline(parent, orient="h", **kwargs):
    kw = {"bg": C["hair"]}
    kw.update(kwargs)
    if orient == "h":
        return tk.Frame(parent, height=1, **kw)
    else:
        return tk.Frame(parent, width=1, **kw)


def _make_dot(parent, color, size=10, bg=None):
    """Bolinha colorida via Canvas — estilo Notion."""
    bg = bg or parent.cget("bg")
    c = tk.Canvas(parent, width=size + 4, height=size + 4,
                  bg=bg, highlightthickness=0, bd=0)
    c.create_oval(2, 2, size + 2, size + 2, fill=color, outline="")
    return c


def _canvas_round_rect(canvas, x1, y1, x2, y2, radius, **kwargs):
    r = min(radius, (x2 - x1) / 2, (y2 - y1) / 2)
    points = [
        x1 + r, y1, x2 - r, y1,
        x2, y1, x2, y1 + r,
        x2, y2 - r, x2, y2,
        x2 - r, y2, x1 + r, y2,
        x1, y2, x1, y2 - r,
        x1, y1 + r, x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, splinesteps=36, **kwargs)


FRAME_LABELS = {
    "Home":              "Início",
    "Rotinas":           "Rotinas",
    "Share":             "Cadastro Share",
    "BPM_CONFIG":        "Configurar BPM Invertido",
    "BPM_CONFIG_NOVA":   "Configurar BPM — Nova Plataforma",
    "BPM_HUB":           "BPM",
    "BPM":               "BPM — Operações",
    "OperacoesInvertido":"Operações Invertido",
    "LimitesInvertido":  "Limites Invertido",
    "AnalisarOperacoes": "Analisar Operações",
    "TaxasInvertido":    "Taxas (Depara)",
}


def _icon_png_path():
    for path in (resource_path(ICON_FILENAME), os.path.join(app_base_dir(), ICON_FILENAME)):
        if os.path.isfile(path):
            return path
    return None


def _png_to_ico_bytes(png_path):
    with open(png_path, "rb") as f:
        png_data = f.read()
    if len(png_data) < 24 or png_data[:8] != b"\x89PNG\r\n\x1a\n":
        return None
    width = int.from_bytes(png_data[16:20], "big")
    height = int.from_bytes(png_data[20:24], "big")
    w_byte = 0 if width >= 256 else width
    h_byte = 0 if height >= 256 else height
    image_offset = 6 + 16
    header = struct.pack("<HHH", 0, 1, 1)
    entry = struct.pack("<BBBBHHII", w_byte, h_byte, 0, 0, 1, 32, len(png_data), image_offset)
    return header + entry + png_data


def _build_ico_from_png(png_path, ico_path):
    try:
        from PIL import Image
        img = Image.open(png_path).convert("RGBA")
        img.save(ico_path, format="ICO", sizes=[(16, 16), (32, 32), (48, 48), (256, 256)])
        return True
    except Exception:
        return False


def _ensure_ico_path():
    cached = os.path.join(app_base_dir(), "itaulogo.ico")
    png_path = _icon_png_path()
    if png_path:
        needs_build = not os.path.isfile(cached)
        if needs_build and _build_ico_from_png(png_path, cached):
            return cached
    if os.path.isfile(cached):
        return cached
    if not png_path:
        return None
    ico_bytes = _png_to_ico_bytes(png_path)
    if not ico_bytes:
        return None
    try:
        with open(cached, "wb") as f:
            f.write(ico_bytes)
        return cached
    except OSError:
        try:
            fd, tmp = tempfile.mkstemp(suffix=".ico", prefix="mesa_itau_")
            os.close(fd)
            with open(tmp, "wb") as f:
                f.write(ico_bytes)
            return tmp
        except OSError:
            return None


def apply_taskbar_presence(root):
    if sys.platform != "win32":
        return
    try:
        GWL_EXSTYLE = -20
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        SW_HIDE = 0
        SW_SHOW = 5
        root.update_idletasks()
        hwnd = _window_hwnd(root)
        if not hwnd:
            return
        style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        new_style = (style & ~WS_EX_TOOLWINDOW) | WS_EX_APPWINDOW
        if new_style != style:
            ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, new_style)
        ctypes.windll.user32.ShowWindow(hwnd, SW_HIDE)
        ctypes.windll.user32.ShowWindow(hwnd, SW_SHOW)
        root.lift()
    except Exception:
        pass


def apply_window_icon(root):
    ico_path = _ensure_ico_path()
    if not ico_path:
        return
    try:
        root.iconbitmap(default=ico_path)
    except Exception:
        pass
    if sys.platform == "win32":
        try:
            root.update_idletasks()
            hwnd = _window_hwnd(root)
            if not hwnd:
                return
            IMAGE_ICON = 1
            LR_LOADFROMFILE = 0x10
            WM_SETICON = 0x0080
            for size in (16, 32):
                hicon = ctypes.windll.user32.LoadImageW(
                    None, ico_path, IMAGE_ICON, size, size, LR_LOADFROMFILE,
                )
                if hicon:
                    which = 0 if size <= 16 else 1
                    ctypes.windll.user32.SendMessageW(hwnd, WM_SETICON, which, hicon)
        except Exception:
            pass


def apply_windows_shell(root):
    apply_taskbar_presence(root)
    apply_window_icon(root)


def apply_modern_window_chrome(root):
    if sys.platform != "win32":
        return
    try:
        hwnd = _window_hwnd(root)
        dwm = ctypes.windll.dwmapi
        dark = ctypes.c_int(1)
        round_pref = ctypes.c_int(2)
        dwm.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(dark), ctypes.sizeof(dark))
        dwm.DwmSetWindowAttribute(hwnd, 33, ctypes.byref(round_pref), ctypes.sizeof(round_pref))
    except Exception:
        pass


def _window_hwnd(root):
    root.update_idletasks()
    return ctypes.windll.user32.GetParent(root.winfo_id())


def apply_frameless_resize(root):
    if sys.platform != "win32":
        return
    try:
        hwnd = _window_hwnd(root)
        gwl_style = -16
        style = ctypes.windll.user32.GetWindowLongW(hwnd, gwl_style)
        style |= 0x00040000
        style |= 0x00020000
        style |= 0x00010000
        ctypes.windll.user32.SetWindowLongW(hwnd, gwl_style, style)
    except Exception:
        pass


def start_native_window_drag(root):
    """Inicia o arrasto nativo da janela (estilo WM_NCLBUTTONDOWN/HTCAPTION).

    Usa PostMessageW (assíncrono) em vez de SendMessageW: o SendMessageW
    entra no loop modal de mover janela do Windows de forma síncrona,
    dentro da chamada ctypes — que libera o GIL antes de chamar a API do
    Windows. Qualquer callback do Tcl/Tk disparado durante esse loop modal
    tenta reaquisitar o GIL num estado inconsistente, causando
    'Fatal Python error: PyEval_RestoreThread' e o crash da aplicação.
    PostMessageW apenas enfileira a mensagem e retorna imediatamente; o
    loop de mover é processado depois, dentro do laço normal de eventos
    do Tcl, onde o GIL é gerenciado corretamente.
    """
    if sys.platform != "win32":
        return False
    try:
        hwnd = _window_hwnd(root)
        if not hwnd:
            return False
        WM_NCLBUTTONDOWN = 0x00A1
        HTCAPTION = 2
        user32 = ctypes.windll.user32
        user32.ReleaseCapture()
        user32.PostMessageW(ctypes.c_void_p(hwnd), WM_NCLBUTTONDOWN,
                            ctypes.c_void_p(HTCAPTION), ctypes.c_void_p(0))
        return True
    except Exception:
        return False


class AppTitleBar(tk.Frame):
    BG = "#1c1c1c"
    HEIGHT = 36

    def __init__(self, parent, root):
        super().__init__(parent, bg=self.BG, height=self.HEIGHT)
        self.pack_propagate(False)
        self.root = root
        self._drag_offset = None
        self._maximized = False

        row = tk.Frame(self, bg=self.BG)
        row.pack(fill="both", expand=True)

        left = tk.Frame(row, bg=self.BG)
        left.pack(side="left", fill="y", padx=(14, 0))

        tk.Label(left, text="Mesa", bg=self.BG, fg=C["ink"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", pady=8)
        tk.Label(left, text="Itaú", bg=self.BG, fg=C["accent"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=(3, 0), pady=8)
        tk.Label(left, text="·", bg=self.BG, fg="#404040",
                 font=("Segoe UI", 9)).pack(side="left", padx=8, pady=8)
        self._module_lbl = tk.Label(left, text="Início", bg=self.BG, fg="#8a8a8a",
                                    font=("Segoe UI", 9))
        self._module_lbl.pack(side="left", pady=8)

        controls = tk.Frame(row, bg=self.BG)
        controls.pack(side="right", fill="y")

        self._btn_min = self._win_btn(controls, "─", self._minimize)
        self._btn_max = self._win_btn(controls, "□", self._toggle_maximize)
        self._btn_close = self._win_btn(controls, "✕", self.root.destroy, close=True)

        tk.Frame(self, bg="#2e2e2e", height=1).pack(fill="x", side="bottom")

        self._bind_drag(self)
        self._bind_drag(row)
        self._bind_drag(left)
        for w in left.winfo_children():
            if w is not self._module_lbl:
                self._bind_drag(w)
        self._bind_drag(self._module_lbl)
        self.bind("<Double-Button-1>", lambda _e: self._toggle_maximize())

    def _win_btn(self, parent, text, command, close=False):
        hover = "#c95f5f" if close else "#333333"
        lbl = tk.Label(parent, text=text, bg=self.BG, fg="#9a9a9a",
                       font=("Segoe UI", 9), width=4, cursor="hand2")
        lbl.pack(side="left", fill="y")
        lbl.bind("<Button-1>", lambda _e: command())
        lbl.bind("<Enter>", lambda _e, l=lbl, h=hover: l.configure(bg=h, fg="#f2f2f2"))
        lbl.bind("<Leave>", lambda _e, l=lbl: l.configure(bg=self.BG, fg="#9a9a9a"))
        return lbl

    def _bind_drag(self, widget):
        widget.bind("<ButtonPress-1>", self._start_drag, add="+")
        if sys.platform != "win32":
            widget.bind("<B1-Motion>", self._on_drag_fallback, add="+")

    def _start_drag(self, event):
        if start_native_window_drag(self.root):
            return
        if self._maximized:
            return
        self._drag_offset = (event.x_root - self.root.winfo_x(),
                             event.y_root - self.root.winfo_y())

    def _on_drag_fallback(self, event):
        if not self._drag_offset or self._maximized:
            return
        ox, oy = self._drag_offset
        self.root.geometry(f"+{event.x_root - ox}+{event.y_root - oy}")

    def _minimize(self):
        if sys.platform == "win32":
            try:
                ctypes.windll.user32.ShowWindow(_window_hwnd(self.root), 6)
                return
            except Exception:
                pass
        self.root.iconify()

    def _toggle_maximize(self):
        if sys.platform == "win32":
            try:
                hwnd = _window_hwnd(self.root)
                if self._maximized:
                    ctypes.windll.user32.ShowWindow(hwnd, 9)
                    self._maximized = False
                    self._btn_max.configure(text="□")
                else:
                    ctypes.windll.user32.ShowWindow(hwnd, 3)
                    self._maximized = True
                    self._btn_max.configure(text="❐")
                return
            except Exception:
                pass
        if self._maximized:
            self.root.state("normal")
            self._maximized = False
            self._btn_max.configure(text="□")
        else:
            self.root.state("zoomed")
            self._maximized = True
            self._btn_max.configure(text="❐")

    def set_module(self, frame_name):
        self._module_lbl.configure(text=FRAME_LABELS.get(frame_name, frame_name))


class AppStatusBar(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"], height=30)
        self.pack_propagate(False)
        self.controller = controller

        tk.Frame(self, bg="#303030", height=1).pack(fill="x")

        row = tk.Frame(self, bg=C["bg"])
        row.pack(fill="both", expand=True, padx=16)

        left = tk.Frame(row, bg=C["bg"])
        left.pack(side="left", fill="y")

        tk.Label(left, text="Mesa Itaú", bg=C["bg"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", pady=4)
        tk.Label(left, text="·", bg=C["bg"], fg="#3a3a3a",
                 font=("Segoe UI", 8)).pack(side="left", padx=6, pady=4)
        tk.Label(left, text="Risco Sacado", bg=C["bg"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", pady=4)

        right = tk.Frame(row, bg=C["bg"])
        right.pack(side="right", fill="y")

        self._clock_lbl = tk.Label(right, text="", bg=C["bg"], fg=C["ink_faint"],
                                   font=("Segoe UI", 8))
        self._clock_lbl.pack(side="right", padx=(12, 0), pady=4)

        self._module_lbl = tk.Label(right, text="Início", bg=C["bg"], fg="#6b6b6b",
                                    font=("Segoe UI", 8))
        self._module_lbl.pack(side="right", pady=4)

        self._tick_clock()

    def set_module(self, frame_name):
        label = FRAME_LABELS.get(frame_name, frame_name)
        self._module_lbl.configure(text=label)

    def _tick_clock(self):
        now = datetime.now()
        self._clock_lbl.configure(text=now.strftime("%d/%m/%Y  %H:%M"))
        self.after(30_000, self._tick_clock)


def styled_label(parent, text, size=10, weight="normal", color=None, **kwargs):
    return tk.Label(parent, text=text,
                    font=("Segoe UI", size, weight),
                    fg=color or C["ink"],
                    bg=kwargs.pop("bg", C["surface"]),
                    **kwargs)


def styled_button(parent, text, command, accent=False, danger=False, small=False, **kwargs):
    if accent:
        bg  = C["accent_dim"];  fg  = C["accent"]
        abg = C["accent"];      afg = C["bg"]
    elif danger:
        bg  = C["err_dim"];     fg  = C["err"]
        abg = C["err"];         afg = C["bg"]
    else:
        bg  = C["surface2"];    fg  = C["ink_muted"]
        abg = C["surface3"];    afg = C["ink"]
    pad = (7, 3) if small else (13, 6)
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=fg, activebackground=abg, activeforeground=afg,
                    font=("Segoe UI", 8 if small else 9),
                    relief="flat", bd=0, padx=pad[0], pady=pad[1],
                    cursor="hand2", **kwargs)
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg,  fg=fg))
    return btn


def styled_button_limite(parent, text, command, variant="warn", small=False, **kwargs):
    styles = {
        "ok":  (C["ok_dim"],  C["ok"],  C["ok"],  C["bg"]),
        "warn":("#3d3520",    C["warn"], C["warn"], C["bg"]),
        "err": (C["err_dim"],  C["err"], C["err"], C["bg"]),
        "idle":(C["surface2"], C["ink_muted"], C["surface3"], C["ink"]),
    }
    bg, fg, abg, afg = styles.get(variant, styles["warn"])
    pad = (7, 3) if small else (13, 6)
    btn = tk.Button(parent, text=text, command=command,
                    bg=bg, fg=fg, activebackground=abg, activeforeground=afg,
                    font=("Segoe UI", 8 if small else 9),
                    relief="flat", bd=0, padx=pad[0], pady=pad[1],
                    cursor="hand2", **kwargs)
    btn._limite_variant = variant
    btn._limite_bg = bg
    btn._limite_fg = fg
    btn._limite_abg = abg
    btn._limite_afg = afg
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg, fg=fg))
    return btn


def _set_limite_button(btn, text, variant):
    styles = {
        "ok":  (C["ok_dim"],  C["ok"],  C["ok"],  C["bg"]),
        "warn":("#3d3520",    C["warn"], C["warn"], C["bg"]),
        "err": (C["err_dim"],  C["err"], C["err"], C["bg"]),
        "idle":(C["surface2"], C["ink_muted"], C["surface3"], C["ink"]),
    }
    bg, fg, abg, afg = styles.get(variant, styles["warn"])
    btn.configure(text=text, bg=bg, fg=fg, activebackground=abg, activeforeground=afg)
    btn._limite_variant = variant
    btn._limite_bg = bg
    btn._limite_fg = fg
    btn._limite_abg = abg
    btn._limite_afg = afg
    btn.bind("<Enter>", lambda _: btn.configure(bg=abg, fg=afg))
    btn.bind("<Leave>", lambda _: btn.configure(bg=bg, fg=fg))


def styled_entry(parent, textvariable=None, width=20, show=None, **kwargs):
    return tk.Entry(parent, textvariable=textvariable, width=width,
                    show=show or "",
                    bg=C["surface2"], fg=C["ink"],
                    insertbackground=C["accent"],
                    relief="flat", highlightthickness=1,
                    highlightbackground=C["hair"],
                    highlightcolor=C["accent"],
                    font=("Segoe UI", 10), **kwargs)


def card_frame(parent, **kwargs):
    kw = {"bg": C["surface"], "highlightthickness": 1,
          "highlightbackground": C["hair"], "bd": 0}
    kw.update(kwargs)
    return tk.Frame(parent, **kw)


def eyebrow_label(parent, text, bg=None):
    bg = bg or C["bg"]
    return tk.Label(parent, text=text, bg=bg, fg=C["ink_faint"],
                    font=("Segoe UI", 7, "bold"))


def make_hub_option_card(parent, row, col, icon, title, sub, command, color,
                          alert=False, alert_text="● Atenção"):
    """Card de opção usado nas telas-hub (Operações Invertido, BPM, etc.),
    em grid de 3 colunas com design minimalista consistente com o app."""
    pad_x = {0: (0, 6), 1: (6, 6), 2: (6, 0)}.get(col, (6, 6))
    pad_y = (0, 0) if row == 0 else (12, 0)
    outer = tk.Frame(parent, bg=C["surface"],
                     highlightthickness=1, highlightbackground=C["hair"],
                     cursor="hand2")
    outer.grid(row=row, column=col, sticky="nsew", padx=pad_x, pady=pad_y)

    top_line = tk.Frame(outer, bg=C["hair"], height=2)
    top_line.pack(fill="x")

    body_f = tk.Frame(outer, bg=C["surface"], padx=22, pady=22)
    body_f.pack(fill="both", expand=True)

    icon_row = tk.Frame(body_f, bg=C["surface"])
    icon_row.pack(fill="x", anchor="w")
    icon_lbl = tk.Label(icon_row, text=icon, bg=C["surface"], fg=color,
                        font=("Segoe UI", 22))
    icon_lbl.pack(side="left")
    alert_lbl = None
    if alert:
        alert_lbl = tk.Label(icon_row, text=alert_text, bg=C["surface"],
                             fg=C["err"], font=("Segoe UI", 7, "bold"))
        alert_lbl.pack(side="right", anchor="n", pady=(4, 0))
    name_lbl = tk.Label(body_f, text=title, bg=C["surface"], fg=C["ink"],
                        font=("Segoe UI", 12, "bold"), anchor="w")
    name_lbl.pack(anchor="w", pady=(12, 4))
    sub_lbl = tk.Label(body_f, text=sub, bg=C["surface"], fg=C["ink_muted"],
                       font=("Segoe UI", 9), anchor="w", wraplength=200,
                       justify="left")
    sub_lbl.pack(anchor="w")
    arrow_lbl = tk.Label(body_f, text="Abrir →", bg=C["surface"], fg=C["ink_faint"],
                         font=("Segoe UI", 8, "bold"))
    arrow_lbl.pack(anchor="w", pady=(18, 0))

    widgets = [outer, top_line, body_f, icon_row, icon_lbl, name_lbl, sub_lbl, arrow_lbl]

    def _enter(_e=None):
        outer.configure(bg=C["surface2"], highlightbackground=color)
        top_line.configure(bg=color)
        for w in (body_f, icon_row, icon_lbl, name_lbl, sub_lbl):
            w.configure(bg=C["surface2"])
        if alert_lbl is not None:
            alert_lbl.configure(bg=C["surface2"])
        arrow_lbl.configure(bg=C["surface2"], fg=color)

    def _leave(_e=None):
        outer.configure(bg=C["surface"], highlightbackground=C["hair"])
        top_line.configure(bg=C["hair"])
        for w in (body_f, icon_row, icon_lbl, name_lbl, sub_lbl):
            w.configure(bg=C["surface"])
        if alert_lbl is not None:
            alert_lbl.configure(bg=C["surface"])
        arrow_lbl.configure(bg=C["surface"], fg=C["ink_faint"])

    for w in widgets:
        w.bind("<Button-1>", lambda _e: command())
        w.bind("<Enter>", _enter)
        w.bind("<Leave>", _leave)

    return outer


def section_divider(parent, text="", bg=None):
    bg = bg or C["bg"]
    row = tk.Frame(parent, bg=bg)
    if text:
        tk.Label(row, text=text, bg=bg, fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(side="left")
        spacer = tk.Frame(row, bg=C["hair"], height=1)
        spacer.pack(side="left", fill="x", expand=True, padx=(10, 0), pady=(5, 0))
    else:
        tk.Frame(row, bg=C["hair"], height=1).pack(fill="x")
    return row

class MinimalScrollbar(tk.Canvas):
    THUMB_MIN = 28

    def __init__(self, parent, command=None, bg=None, width=6, **kwargs):
        self._track_bg = bg or C["bg"]
        super().__init__(
            parent, width=width, highlightthickness=0, bd=0,
            bg=self._track_bg, cursor="arrow", **kwargs,
        )
        self._command = command
        self._first = 0.0
        self._last = 1.0
        self._thumb_fill = "#5c5c5c"
        self._thumb_hover = "#787878"
        self._drag_y = 0
        self._thumb_rect = None

        self.bind("<Configure>", self._redraw, add="+")
        self.bind("<Button-1>", self._on_press)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Enter>", lambda _e: self._paint_thumb(self._thumb_hover))
        self.bind("<Leave>", lambda _e: self._paint_thumb(self._thumb_fill))

    def set(self, first, last):
        f, l = float(first), float(last)
        if f == self._first and l == self._last:
            return
        self._first, self._last = f, l
        self._redraw()

    def _visible(self):
        return self._first > 0.001 or self._last < 0.999

    def _thumb_geometry(self):
        h = max(self.winfo_height(), 1)
        w = max(self.winfo_width(), 1)
        span = max(self._last - self._first, 0.001)
        thumb_h = max(self.THUMB_MIN, int(h * span))
        thumb_y = int(h * self._first)
        if thumb_y + thumb_h > h:
            thumb_y = max(0, h - thumb_h)
        return w, h, thumb_y, thumb_h

    def _redraw(self, _event=None):
        self.delete("all")
        self._thumb_rect = None
        if not self._visible():
            return
        w, _h, thumb_y, thumb_h = self._thumb_geometry()
        margin = 1
        x0, x1 = margin, max(margin + 2, w - margin)
        radius = (x1 - x0) / 2
        if thumb_h <= (x1 - x0):
            self.create_oval(x0, thumb_y, x1, thumb_y + (x1 - x0),
                             fill=self._thumb_fill, outline="", tags="thumb")
            self.create_oval(x0, thumb_y + thumb_h - (x1 - x0), x1, thumb_y + thumb_h,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self._thumb_rect = (x0, thumb_y, x1, thumb_y + thumb_h)
        else:
            mid_top = thumb_y + radius
            mid_bot = thumb_y + thumb_h - radius
            self.create_oval(x0, thumb_y, x1, mid_top + radius,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self.create_rectangle(x0, mid_top, x1, mid_bot,
                                  fill=self._thumb_fill, outline="", tags="thumb")
            self.create_oval(x0, mid_bot - radius, x1, thumb_y + thumb_h,
                             fill=self._thumb_fill, outline="", tags="thumb")
            self._thumb_rect = (x0, thumb_y, x1, thumb_y + thumb_h)

    def _paint_thumb(self, color):
        for item in self.find_withtag("thumb"):
            self.itemconfig(item, fill=color)

    def _on_press(self, event):
        if not self._command or not self._visible():
            return
        w, h, thumb_y, thumb_h = self._thumb_geometry()
        if self._thumb_rect and self._thumb_rect[1] <= event.y <= self._thumb_rect[3]:
            self._drag_y = event.y - thumb_y
            return
        if event.y > thumb_y + thumb_h:
            self._command("scroll", 1, "pages")
        elif event.y < thumb_y:
            self._command("scroll", -1, "pages")

    def _on_drag(self, event):
        if not self._command or not self._visible():
            return
        w, h, _thumb_y, thumb_h = self._thumb_geometry()
        span = max(h - thumb_h, 1)
        frac = (event.y - self._drag_y) / span
        self._command("moveto", max(0.0, min(1.0, frac)))

    def _on_release(self, _event):
        self._drag_y = 0


def bind_text_mousewheel(text_widget):
    def _mw(event):
        try:
            if not text_widget.winfo_exists():
                return
        except tk.TclError:
            return
        if getattr(event, "delta", 0):
            text_widget.yview_scroll(int(-event.delta / 120), "units")
        elif event.num == 4:
            text_widget.yview_scroll(-3, "units")
        elif event.num == 5:
            text_widget.yview_scroll(3, "units")
    text_widget.bind("<MouseWheel>", _mw)
    text_widget.bind("<Button-4>", _mw)
    text_widget.bind("<Button-5>", _mw)


class ScrollableFrame(tk.Frame):
    def __init__(self, parent, bg=None, **kwargs):
        bg = bg or C["bg"]
        super().__init__(parent, bg=bg, **kwargs)
        self._bg = bg
        self._wheel_roots = []
        self._canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self._vbar = MinimalScrollbar(self, command=self._canvas.yview, bg=bg, width=6)
        self._canvas.configure(yscrollcommand=self._vbar.set)
        self._vbar.pack(side="right", fill="y", padx=(0, 2), pady=4)
        self._canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self._canvas, bg=bg)
        self._win = self._canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", self._on_inner)
        self._canvas.bind("<Configure>", self._on_canvas)
        self.bind("<Destroy>", self._on_destroy)
        self._wheel_roots = [self]
        self._scroll_job = None
        self.refresh_bindings()

    def _canvas_alive(self):
        try:
            return bool(self._canvas.winfo_exists())
        except tk.TclError:
            return False

    def _scroll_mousewheel(self, event):
        if not self._canvas_alive():
            return
        try:
            if getattr(event, "delta", 0):
                self._canvas.yview_scroll(int(-event.delta / 120), "units")
            elif event.num == 4:
                self._canvas.yview_scroll(-3, "units")
            elif event.num == 5:
                self._canvas.yview_scroll(3, "units")
        except tk.TclError:
            pass

    def _bind_mousewheel_tree(self, widget):
        if isinstance(widget, (tk.Text, tk.Listbox)):
            return
        widget.bind("<MouseWheel>", self._scroll_mousewheel)
        widget.bind("<Button-4>", self._scroll_mousewheel)
        widget.bind("<Button-5>", self._scroll_mousewheel)
        for child in widget.winfo_children():
            self._bind_mousewheel_tree(child)

    def refresh_bindings(self):
        if not self._canvas_alive():
            return
        for root in self._wheel_roots:
            try:
                if root.winfo_exists():
                    self._bind_mousewheel_tree(root)
            except tk.TclError:
                pass

    def link_wheel(self, container):
        if container not in self._wheel_roots:
            self._wheel_roots.append(container)
        self.refresh_bindings()

    def _on_inner(self, _event):
        if not self._canvas_alive():
            return
        if getattr(self, "_scroll_job", None):
            try:
                self.after_cancel(self._scroll_job)
            except Exception:
                pass
        self._scroll_job = self.after(16, self._sync_scrollregion)

    def _sync_scrollregion(self):
        self._scroll_job = None
        if not self._canvas_alive():
            return
        try:
            self._canvas.update_idletasks()
            self._canvas.configure(scrollregion=self._canvas.bbox("all"))
        except tk.TclError:
            pass

    def _on_canvas(self, event):
        if self._canvas_alive():
            self._canvas.itemconfigure(self._win, width=event.width)

    def _on_destroy(self, event):
        pass


class Sidebar(tk.Frame):
    NAV = [
        ("Home",              "⌂",  "Início"),
        ("Rotinas",           "◈",  "Rotinas"),
        ("Share",             "⊕",  "Cadastro Share"),
        ("BPM",               "⚡",  "BPM"),
        ("OperacoesInvertido","⬡",  "Operações Invertido"),
    ]

    def __init__(self, parent, controller, **kwargs):
        super().__init__(parent, bg=C["surface"], width=210, **kwargs)
        self.pack_propagate(False)
        self.controller = controller
        self._btns = {}
        self._build()

    def _build(self):
        top = tk.Frame(self, bg=C["surface"])
        top.pack(fill="x", padx=18, pady=(14, 0))

        logo_row = tk.Frame(top, bg=C["surface"])
        logo_row.pack(fill="x")
        tk.Label(logo_row, text="Mesa", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Label(logo_row, text=" Itaú", bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Label(top, text="Risco Sacado", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(3, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(18, 12))

        nav_outer = tk.Frame(self, bg=C["surface"])
        nav_outer.pack(fill="both", expand=True, anchor="n")

        for name, icon, label in self.NAV:
            row = tk.Frame(nav_outer, bg=C["surface"], cursor="hand2")
            row.pack(fill="x", pady=1, padx=6)

            bar = tk.Frame(row, bg=C["surface"], width=3)
            bar.pack(side="left", fill="y")

            inner = tk.Frame(row, bg=C["surface"], padx=8, pady=7)
            inner.pack(side="left", fill="x", expand=True)

            icon_lbl = tk.Label(inner, text=icon, bg=C["surface"],
                                fg=C["ink_faint"], font=("Segoe UI", 12), width=2, anchor="w")
            icon_lbl.pack(side="left")

            text_lbl = tk.Label(inner, text=label, bg=C["surface"],
                                fg=C["ink_muted"], font=("Segoe UI", 9), anchor="w")
            text_lbl.pack(side="left", padx=(6, 0))

            def _click(n=name): self.controller.show_frame(n)

            def _enter(e, r=row, inn=inner, il=icon_lbl, tl=text_lbl):
                active = getattr(self.controller, "_active_frame", None)
                n_     = ""
                for _n, _b in self._btns.items():
                    if _b["row"] is r: n_ = _n; break
                if n_ != active:
                    for w in (r, inn, il, tl):
                        try: w.configure(bg=C["surface2"])
                        except: pass

            def _leave(e, r=row, inn=inner, il=icon_lbl, tl=text_lbl):
                active = getattr(self.controller, "_active_frame", None)
                n_     = ""
                for _n, _b in self._btns.items():
                    if _b["row"] is r: n_ = _n; break
                if n_ != active:
                    for w in (r, inn, il, tl):
                        try: w.configure(bg=C["surface"])
                        except: pass

            for w in (row, inner, icon_lbl, text_lbl):
                w.bind("<Button-1>", lambda _, n=name: _click(n))
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)

            self._btns[name] = {
                "row": row, "inner": inner,
                "icon": icon_lbl, "text": text_lbl, "bar": bar
            }

    def set_active(self, name):
        for n, w in self._btns.items():
            is_active = (n == name)
            row_bg  = C["surface2"] if is_active else C["surface"]
            icon_fg = C["accent"]   if is_active else C["ink_faint"]
            text_fg = C["ink"]      if is_active else C["ink_muted"]
            bar_bg  = C["accent"]   if is_active else C["surface"]

            w["row"].configure(bg=row_bg)
            w["inner"].configure(bg=row_bg)
            w["icon"].configure(bg=row_bg, fg=icon_fg)
            w["text"].configure(bg=row_bg, fg=text_fg)
            w["bar"].configure(bg=bar_bg)

def get_market_status(now=None):
    """Retorna (aberto: bool, texto: str) com base na tabela de horários de
    antecipação. Fora do horário comercial ou em finais de semana, o
    mercado é considerado fechado."""
    now = now or datetime.now()
    if now.weekday() >= 5:
        return False, "Mercado fechado"
    t = now.hour * 60 + now.minute
    ranges = [
        (7*60,        8*60+50,  "Pagamento a partir das 09h00"),
        (8*60+51,     9*60+35,  "Pagamento a partir das 09h40"),
        (9*60+36,     10*60+5,  "Pagamento a partir das 10h10"),
        (10*60+6,     10*60+20, "Pagamento a partir das 10h30"),
        (10*60+21,    16*60+35, "Pagamento entre 30 e 40' de intervalo"),
        (16*60+36,    17*60,    "Pagamento a partir das 17h10"),
    ]
    for start, end, label in ranges:
        if start <= t <= end:
            return True, label
    return False, "Mercado fechado"


class HomeFrame(tk.Frame):
    MODULES = [
        {"name": "Cadastro Share",     "sub": "Extração e análise de PDF",          "icon": "⊕", "frame": "Share",            "color": "#5a9e72"},
        {"name": "BPM",                "sub": "Abertura de solicitações",            "icon": "⚡", "frame": "BPM_HUB",          "color": "#EC7000"},
        {"name": "Operações Invertido","sub": "Limites, LTC e análise de planilhas", "icon": "⬡", "frame": "OperacoesInvertido","color": "#c87941"},
        {"name": "Rotinas",            "sub": "Sequências configuráveis",            "icon": "◈", "frame": "Rotinas",          "color": "#8b72c9"},
    ]

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._build()

    def _build(self):
        self._sf = ScrollableFrame(self)
        self._sf.pack(fill="both", expand=True)
        inner = self._sf.inner
        inner.configure(bg=C["bg"])
        inner.columnconfigure(0, weight=1)

        greet = tk.Frame(inner, bg=C["bg"])
        greet.pack(fill="x", padx=44, pady=(40, 0))

        now  = datetime.now()
        hour = now.hour
        saudacao = "Bom dia" if hour < 12 else ("Boa tarde" if hour < 18 else "Boa noite")

        eyebrow_label(greet, "MESA DE OPERAÇÕES").pack(anchor="w")
        tk.Label(greet, text=f"{saudacao}.", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 26, "bold"), anchor="w").pack(anchor="w", pady=(6, 0))
        tk.Label(greet,
                 text=format_data_pt_br(now),
                 bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 0))

        self._market_row = tk.Frame(greet, bg=C["bg"])
        self._market_row.pack(anchor="w", pady=(10, 0))
        self._market_dot = tk.Canvas(self._market_row, width=9, height=9,
                                     bg=C["bg"], highlightthickness=0, bd=0)
        self._market_dot.pack(side="left", padx=(0, 7))
        self._market_lbl = tk.Label(self._market_row, bg=C["bg"],
                                    font=("Segoe UI", 9, "bold"))
        self._market_lbl.pack(side="left")
        self._paint_market_status()

        make_hairline(inner, bg=C["hair"]).pack(fill="x", padx=44, pady=(28, 26))

        eyebrow_label(inner, "MÓDULOS").pack(anchor="w", padx=44, pady=(0, 14))

        grid = tk.Frame(inner, bg=C["bg"])
        grid.pack(fill="x", padx=44)
        grid.columnconfigure(0, weight=1, uniform="m")
        grid.columnconfigure(1, weight=1, uniform="m")

        for i, mod in enumerate(self.MODULES):
            r, c = divmod(i, 2)
            self._make_module_card(grid, mod, r, c)

        make_hairline(inner, bg=C["hair"]).pack(fill="x", padx=44, pady=(30, 24))
        eyebrow_label(inner, "ROTINAS DE HOJE").pack(anchor="w", padx=44, pady=(0, 8))

        self._rot_container = tk.Frame(inner, bg=C["bg"])
        self._rot_container.pack(fill="x", padx=44, pady=(0, 40))
        self._build_rotinas_hoje()

    def on_show(self):
        self.refresh_rotinas()
        if not getattr(self, "_market_loop_started", False):
            self._market_loop_started = True
            self._refresh_market_status()
        else:
            self._paint_market_status()

    def _refresh_market_status(self):
        self._paint_market_status()
        self.after(30_000, self._refresh_market_status)

    def _paint_market_status(self):
        aberto, texto = get_market_status()
        color = C["ok"] if aberto else C["err"]
        glow  = C["ok_dim"] if aberto else C["err_dim"]

        dot = self._market_dot
        dot.delete("all")
        dot.create_oval(0, 0, 9, 9, fill=glow, outline="")
        dot.create_oval(2, 2, 7, 7, fill=color, outline="")

        self._market_lbl.configure(
            text=(texto if aberto else "Mercado fechado"),
            fg=(C["ok"] if aberto else C["ink_muted"]))

    def refresh_rotinas(self):
        if not hasattr(self, "_rot_container"):
            return
        for w in self._rot_container.winfo_children():
            w.destroy()
        self._build_rotinas_hoje()

    def _build_rotinas_hoje(self):
        parent = self._rot_container
        data   = RotinasData.get()
        rots   = data.today_rotinas()
        done_n, total = data.today_stats()

        if not rots:
            tk.Label(parent, text="Nenhuma rotina para hoje — configure em Rotinas.",
                     bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 9)).pack(anchor="w")
            return

        prog_row = tk.Frame(parent, bg=C["bg"])
        prog_row.pack(fill="x", pady=(0, 10))
        prog_color = C["ok"] if done_n == total else C["accent"]
        tk.Label(prog_row, text=f"{done_n}/{total} concluída{'s' if total > 1 else ''}",
                 bg=C["bg"], fg=prog_color,
                 font=("Segoe UI", 8, "bold")).pack(side="left")
        lnk = tk.Button(prog_row, text="ver todas →",
                        command=lambda: self.controller.show_frame("Rotinas"),
                        bg=C["bg"], fg=C["ink_faint"],
                        activebackground=C["bg"], activeforeground=C["accent"],
                        font=("Segoe UI", 8), relief="flat", bd=0, padx=0,
                        cursor="hand2")
        lnk.pack(side="right")
        lnk.bind("<Enter>", lambda e: lnk.configure(fg=C["accent"]))
        lnk.bind("<Leave>", lambda e: lnk.configure(fg=C["ink_faint"]))

        sorted_rots = sorted(rots, key=RotinasData._rot_sort_key)
        for rot in sorted_rots[:6]:
            self._make_home_rot_row(parent, rot, data)

        if len(rots) > 6:
            tk.Label(parent,
                     text=f"+ {len(rots)-6} rotinas — veja todas em Rotinas",
                     bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 8)).pack(anchor="w", pady=(4, 0))

    def _make_home_rot_row(self, parent, rot, data):
        from tkinter import font as tkfont
        rid   = rot["id"]
        color = rot.get("cor", C["accent"])
        nome  = rot.get("nome", "Rotina")
        _als  = rot.get("alertas") or []
        hora  = _als[0] if _als else (rot.get("hora_alerta") or "")

        row = tk.Frame(parent, bg=C["bg"], pady=4)
        row.pack(fill="x")

        chk = tk.Canvas(row, width=16, height=16,
                        bg=C["bg"], highlightthickness=0, bd=0, cursor="hand2")
        chk.pack(side="left", padx=(0, 8))

        if hora:
            tk.Label(row, text=hora, bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 7), width=5, anchor="w").pack(side="left")

        _make_dot(row, color, 6, bg=C["bg"]).pack(side="left", padx=(0, 6))

        name_lbl = tk.Label(row, text=nome, bg=C["bg"],
                            fg=C["ink"], font=("Segoe UI", 9), anchor="w")
        name_lbl.pack(side="left")

        def _refresh_chk(done):
            chk.delete("all")
            if done:
                chk.create_oval(0, 0, 15, 15, fill=color, outline="")
                chk.create_text(8, 8, text="✓", fill=C["bg"],
                                font=("Segoe UI", 7, "bold"))
                name_lbl.configure(fg=C["ink_faint"],
                                   font=("Segoe UI", 9, "overstrike"))
            else:
                chk.create_oval(0, 0, 15, 15, fill="",
                                outline=C["ink_faint"], width=1.5)
                name_lbl.configure(fg=C["ink"], font=("Segoe UI", 9))

        _refresh_chk(data.is_done(rid))

        def toggle(e=None):
            new = not data.is_done(rid)
            data.set_done(rid, new)
            _refresh_chk(new)
            self.refresh_rotinas()

        for w in [row, chk, name_lbl]:
            try: w.bind("<Button-1>", toggle)
            except Exception: pass

    def _make_module_card(self, parent, mod, row, col):
        dot_color = mod.get("color", C["accent"])
        pad = (0, 6) if col == 0 else (6, 0)

        outer = tk.Frame(parent, bg=C["surface"],
                         highlightthickness=1, highlightbackground=C["hair"],
                         cursor="hand2")
        outer.grid(row=row, column=col, sticky="nsew", padx=pad, pady=6)

        top_line = tk.Frame(outer, bg=C["hair"], height=2)
        top_line.pack(fill="x")

        body = tk.Frame(outer, bg=C["surface"], padx=18, pady=16)
        body.pack(fill="both", expand=True)

        icon_row = tk.Frame(body, bg=C["surface"])
        icon_row.pack(fill="x")
        dot = _make_dot(icon_row, dot_color, size=8, bg=C["surface"])
        dot.pack(side="left", pady=(3, 0))
        tk.Label(icon_row, text=mod["icon"], bg=C["surface"], fg=dot_color,
                 font=("Segoe UI", 16)).pack(side="left", padx=(6, 0))

        name_lbl = tk.Label(body, text=mod["name"], bg=C["surface"], fg=C["ink"],
                             font=("Segoe UI", 11, "bold"), anchor="w")
        name_lbl.pack(anchor="w", pady=(8, 2))

        sub_lbl = tk.Label(body, text=mod["sub"], bg=C["surface"], fg=C["ink_muted"],
                           font=("Segoe UI", 8), anchor="w",
                           wraplength=160, justify="left")
        sub_lbl.pack(anchor="w")

        def cmd(f=mod["frame"]): self.controller.show_frame(f)

        def _enter(e):
            outer.configure(bg=C["surface2"], highlightbackground=dot_color)
            top_line.configure(bg=dot_color)
            body.configure(bg=C["surface2"])
            icon_row.configure(bg=C["surface2"])
            name_lbl.configure(bg=C["surface2"])
            sub_lbl.configure(bg=C["surface2"])
            dot.configure(bg=C["surface2"])
            for w in icon_row.winfo_children():
                try: w.configure(bg=C["surface2"])
                except: pass

        def _leave(e):
            outer.configure(bg=C["surface"], highlightbackground=C["hair"])
            top_line.configure(bg=C["hair"])
            body.configure(bg=C["surface"])
            icon_row.configure(bg=C["surface"])
            name_lbl.configure(bg=C["surface"])
            sub_lbl.configure(bg=C["surface"])
            dot.configure(bg=C["surface"])
            for w in icon_row.winfo_children():
                try: w.configure(bg=C["surface"])
                except: pass

        for w in [outer, body, icon_row, name_lbl, sub_lbl, dot, top_line]:
            try:
                w.bind("<Button-1>", lambda _: cmd())
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)
            except: pass


# ─── Rotinas Data Layer ─────────────────────────────────────────────────────

def _rotinas_data_path():
    return os.path.join(app_base_dir(), "rotinas_data.json")


SHARED_TAXAS_PATH = (
    r"\\BBAPROD3\fo\Diretoria de Produtos Ativos\Ativos em Reais\Risco Sacado"
    r"\Comum\AppMiddle\bancodedados\dados.json"
)


class RotinasData:
    _instance = None

    @classmethod
    def get(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def __init__(self):
        self._rotinas    = []
        self._conclusoes = {}
        self._load()

    def _load(self):
        path = _rotinas_data_path()
        if not os.path.isfile(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = _json_mod.load(f)
            self._rotinas    = data.get("rotinas",    [])
            self._conclusoes = data.get("conclusoes", {})
            for r in self._rotinas:
                if "dias" not in r:
                    freq = r.get("frequencia", "diaria")
                    if freq == "diaria":
                        r["dias"] = [0, 1, 2, 3, 4, 5, 6]
                    elif freq == "dias_uteis":
                        r["dias"] = [0, 1, 2, 3, 4]
                    elif freq.startswith("semanal_"):
                        try:
                            r["dias"] = [int(freq.split("_")[1])]
                        except Exception:
                            r["dias"] = [0, 1, 2, 3, 4, 5, 6]
                    else:
                        r["dias"] = [0, 1, 2, 3, 4, 5, 6]
                if "alertas" not in r:
                    h = (r.get("hora_alerta") or "").strip()
                    r["alertas"] = [h] if h else []
        except Exception:
            pass

    def save(self):
        path = _rotinas_data_path()
        try:
            with open(path, "w", encoding="utf-8") as f:
                _json_mod.dump(
                    {"rotinas": self._rotinas, "conclusoes": self._conclusoes},
                    f, ensure_ascii=False, indent=2
                )
        except Exception:
            pass

    def rotinas(self):
        return list(self._rotinas)

    def add_rotina(self, nome, dias, alertas, cor, notas=""):
        r = {
            "id":      str(_uuid_mod.uuid4())[:8],
            "nome":    nome,
            "dias":    dias,
            "alertas": alertas,
            "cor":     cor,
            "notas":   notas,
            "ativa":   True,
        }
        self._rotinas.append(r)
        self.save()
        return r

    def update_rotina(self, rid, **kwargs):
        for r in self._rotinas:
            if r["id"] == rid:
                r.update(kwargs)
                self.save()
                return True
        return False

    def delete_rotina(self, rid):
        self._rotinas = [r for r in self._rotinas if r["id"] != rid]
        self.save()

    @staticmethod
    def _today_key():
        return date.today().isoformat()

    def is_done(self, rid, day=None):
        key = day or self._today_key()
        return self._conclusoes.get(key, {}).get(rid, False)

    def set_done(self, rid, done: bool, day=None):
        key = day or self._today_key()
        if key not in self._conclusoes:
            self._conclusoes[key] = {}
        self._conclusoes[key][rid] = done
        today = date.today()
        old_keys = [k for k in list(self._conclusoes)
                    if (today - date.fromisoformat(k)).days > 90]
        for k in old_keys:
            del self._conclusoes[k]
        self.save()

    def today_rotinas(self):
        today = date.today()
        wd    = today.weekday()
        result = []
        for r in self._rotinas:
            if not r.get("ativa", True):
                continue
            dias = r.get("dias")
            if dias is not None:
                if wd in dias:
                    result.append(r)
            else:
                freq = r.get("frequencia", "diaria")
                if freq == "diaria":
                    result.append(r)
                elif freq == "dias_uteis" and wd < 5:
                    result.append(r)
                elif freq.startswith("semanal_"):
                    try:
                        if wd == int(freq.split("_")[1]):
                            result.append(r)
                    except Exception:
                        pass
        return result

    @staticmethod
    def _rot_sort_key(rot):
        alertas = rot.get("alertas") or []
        if alertas:
            return alertas[0]
        return (rot.get("hora_alerta") or "99:99")

    def today_stats(self):
        rots  = self.today_rotinas()
        total = len(rots)
        done  = sum(1 for r in rots if self.is_done(r["id"]))
        return done, total


# ─── Taxas (Depara) Data Layer ──────────────────────────────────────────────
# Persistido no MESMO arquivo de rotinas (rotinas_data.json), chave "taxas".
# Estrutura: { cnpj: {"taxa": "1,3950", "validade_mes": "2026-06"} }

def _mes_atual_key():
    today = date.today()
    return f"{today.year:04d}-{today.month:02d}"


class TaxasData:
    """Camada de dados de Taxas (Depara) — usa EXCLUSIVAMENTE o arquivo de
    rede compartilhado. Se a rede estiver indisponível, não há fallback
    local: os dados ficam temporariamente inacessíveis e uma rotina em
    background tenta reconectar a cada 30s, silenciosamente."""

    _instance = None
    RETRY_SECONDS = 30

    @classmethod
    def get(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def __init__(self):
        self._taxas = {}
        self._available = False
        self._on_reconnect_callbacks = []
        self._retry_timer = None
        self._try_load()
        if not self._available:
            self._schedule_retry()

    # ── Conectividade ────────────────────────────────────────────────────
    def is_available(self):
        return self._available

    def on_reconnect(self, callback):
        """Registra um callback (sem argumentos) chamado quando a conexão
        com o arquivo de rede for restabelecida."""
        self._on_reconnect_callbacks.append(callback)

    def _schedule_retry(self):
        if self._retry_timer is not None:
            return
        self._retry_timer = threading.Timer(self.RETRY_SECONDS, self._retry_tick)
        self._retry_timer.daemon = True
        self._retry_timer.start()

    def _retry_tick(self):
        self._retry_timer = None
        was_available = self._available
        self._try_load()
        if self._available and not was_available:
            for cb in list(self._on_reconnect_callbacks):
                try:
                    cb()
                except Exception:
                    pass
        if not self._available:
            self._schedule_retry()

    def _network_reachable(self):
        try:
            return os.path.isdir(os.path.dirname(SHARED_TAXAS_PATH))
        except Exception:
            return False

    def _try_load(self):
        if not self._network_reachable():
            self._available = False
            return
        try:
            if os.path.isfile(SHARED_TAXAS_PATH):
                conteudo = ""
                with open(SHARED_TAXAS_PATH, "r", encoding="utf-8") as f:
                    conteudo = f.read().strip()
                if conteudo:
                    try:
                        data = _json_mod.loads(conteudo)
                    except Exception:
                        # JSON corrompido/inválido: trata como vazio, mas a
                        # rede está acessível — não é motivo pra indisponível.
                        data = {}
                    self._taxas = data.get("taxas", {})
                else:
                    # Arquivo existe mas está vazio: inicializa estrutura.
                    self._taxas = {}
            # Rede acessível (com ou sem arquivo ainda criado) = disponível.
            self._available = True
        except Exception:
            self._available = False

    def _load(self):
        self._try_load()

    # ── Persistência ─────────────────────────────────────────────────────
    def save(self):
        if not self._network_reachable():
            self._available = False
            self._schedule_retry()
            return False
        data = {}
        try:
            if os.path.isfile(SHARED_TAXAS_PATH):
                with open(SHARED_TAXAS_PATH, "r", encoding="utf-8") as f:
                    conteudo = f.read().strip()
                if conteudo:
                    data = _json_mod.loads(conteudo)
        except Exception:
            data = {}
        data["taxas"] = self._taxas
        try:
            with open(SHARED_TAXAS_PATH, "w", encoding="utf-8") as f:
                _json_mod.dump(data, f, ensure_ascii=False, indent=2)
            self._available = True
            return True
        except Exception:
            self._available = False
            self._schedule_retry()
            return False

    def get_taxa(self, cnpj):
        return self._taxas.get(cnpj)

    def set_taxa(self, cnpj, valor_str):
        self._taxas[cnpj] = {
            "taxa": valor_str,
            "validade_mes": _mes_atual_key(),
        }
        return self.save()

    def is_vigente(self, cnpj):
        info = self._taxas.get(cnpj)
        if not info:
            return False
        return info.get("validade_mes") == _mes_atual_key()

    def todas_vigentes(self, cnpjs):
        return all(self.is_vigente(c) for c in cnpjs if c)

    def vencidas(self, cnpjs):
        return [c for c in cnpjs if c and not self.is_vigente(c)]


# ─── RotinasFrame ────────────────────────────────────────────────────────────

class RotinasFrame(tk.Frame):
    FREQ_LABELS = {
        "diaria":     "Diária",
        "dias_uteis": "Dias úteis (Seg–Sex)",
        "semanal_0":  "Semanal — Segunda",
        "semanal_1":  "Semanal — Terça",
        "semanal_2":  "Semanal — Quarta",
        "semanal_3":  "Semanal — Quinta",
        "semanal_4":  "Semanal — Sexta",
        "semanal_5":  "Semanal — Sábado",
        "semanal_6":  "Semanal — Domingo",
    }
    FREQ_OPTIONS = list(FREQ_LABELS.values())
    FREQ_KEYS    = list(FREQ_LABELS.keys())

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._data      = RotinasData.get()
        self._tab       = "hoje"
        self._build()
        self.after(500, self._start_alert_checker)

    def _build(self):
        hdr_wrap = tk.Frame(self, bg=C["bg"])
        hdr_wrap.pack(fill="x", padx=44, pady=(36, 0))
        eyebrow_label(hdr_wrap, "PLANEJAMENTO DO DIA").pack(anchor="w")
        tk.Label(hdr_wrap, text="Rotinas", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 22, "bold")).pack(anchor="w", pady=(6, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", pady=(20, 0))

        tab_row = tk.Frame(self, bg=C["bg"])
        tab_row.pack(fill="x", padx=44)
        self._tab_btns = {}
        for key, label in [("hoje", "Hoje"), ("gerenciar", "Gerenciar")]:
            btn = tk.Button(tab_row, text=label,
                            command=lambda k=key: self._switch_tab(k),
                            bg=C["bg"], fg=C["ink_muted"],
                            activebackground=C["bg"], activeforeground=C["ink"],
                            font=("Segoe UI", 9), relief="flat", bd=0,
                            padx=0, pady=10, cursor="hand2")
            btn.pack(side="left", padx=(0, 22))
            self._tab_btns[key] = btn

        make_hairline(self, bg=C["hair"]).pack(fill="x")

        self._content_area = tk.Frame(self, bg=C["bg"])
        self._content_area.pack(fill="both", expand=True)
        self._switch_tab("hoje")

    def _switch_tab(self, tab):
        self._tab = tab
        for k, btn in self._tab_btns.items():
            if k == tab:
                btn.configure(fg=C["accent"], font=("Segoe UI", 9, "bold"))
            else:
                btn.configure(fg=C["ink_muted"], font=("Segoe UI", 9))
        for w in self._content_area.winfo_children():
            w.destroy()
        if tab == "hoje":
            self._build_hoje(self._content_area)
        else:
            self._build_gerenciar(self._content_area)

    def on_show(self):
        self._switch_tab(self._tab)

    def _build_hoje(self, parent):
        sf    = ScrollableFrame(parent, bg=C["bg"])
        sf.pack(fill="both", expand=True)
        inner = sf.inner
        inner.configure(bg=C["bg"])

        now        = datetime.now()
        today_rots = self._data.today_rotinas()
        done_n, total = self._data.today_stats()

        hdr = tk.Frame(inner, bg=C["bg"])
        hdr.pack(fill="x", padx=44, pady=(28, 0))
        tk.Label(hdr, text=format_data_pt_br(now), bg=C["bg"],
                 fg=C["ink_muted"], font=("Segoe UI", 10)).pack(anchor="w")

        if total > 0:
            prog_text = f"{done_n} de {total} concluída{'s' if total > 1 else ''}"
            prog_color = C["ok"] if done_n == total else C["accent"]
        else:
            prog_text  = "Nenhuma rotina para hoje"
            prog_color = C["ink_faint"]

        tk.Label(hdr, text=prog_text, bg=C["bg"], fg=prog_color,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(6, 0))

        if total > 0:
            bar_bg = tk.Frame(hdr, bg=C["surface2"], height=4)
            bar_bg.pack(fill="x", pady=(8, 0))
            bar_bg.pack_propagate(False)
            bar_fg = tk.Frame(bar_bg, bg=prog_color, height=4)
            bar_fg.place(x=0, y=0, relheight=1.0, relwidth=min(done_n / total, 1.0))

        make_hairline(inner, bg=C["hair"]).pack(fill="x", padx=44, pady=(20, 4))

        if not today_rots:
            empty = tk.Frame(inner, bg=C["bg"])
            empty.pack(fill="x", padx=44, pady=40)
            tk.Label(empty, text="◈", bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 24)).pack()
            tk.Label(empty, text="Nenhuma rotina configurada para hoje",
                     bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 11)).pack(pady=(8, 0))
            tk.Label(empty, text="Vá em Gerenciar para criar suas rotinas.",
                     bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 9)).pack(pady=(4, 0))
            styled_button(empty, "→ Gerenciar rotinas",
                          lambda: self._switch_tab("gerenciar"),
                          accent=True).pack(pady=(16, 0))
            return

        sorted_rots = sorted(today_rots, key=RotinasData._rot_sort_key)
        for rot in sorted_rots:
            self._make_hoje_row(inner, rot)
        tk.Frame(inner, bg=C["bg"], height=40).pack()
        sf.refresh_bindings()

    def _make_hoje_row(self, parent, rot):
        rid   = rot["id"]
        color = rot.get("cor", C["accent"])
        _als  = rot.get("alertas") or []
        hora  = _als[0] if _als else (rot.get("hora_alerta") or "")
        nome  = rot.get("nome", "Rotina")
        notas = (rot.get("notas") or "").strip()

        row = tk.Frame(parent, bg=C["bg"])
        row.pack(fill="x", padx=44, pady=2)
        inner_row = tk.Frame(row, bg=C["bg"], pady=9)
        inner_row.pack(fill="x")

        chk = tk.Canvas(inner_row, width=20, height=20,
                        bg=C["bg"], highlightthickness=0, bd=0, cursor="hand2")
        chk.pack(side="left", padx=(0, 12))

        if hora:
            tk.Label(inner_row, text=hora, bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 8), width=5, anchor="w").pack(side="left")

        dot = _make_dot(inner_row, color, 8, bg=C["bg"])
        dot.pack(side="left", padx=(0, 8))

        name_lbl = tk.Label(inner_row, text=nome, bg=C["bg"],
                            fg=C["ink"], font=("Segoe UI", 10), anchor="w")
        name_lbl.pack(side="left", fill="x", expand=True)

        note_lbl = None
        if notas:
            note_lbl = tk.Label(parent, text=f"      {notas}", bg=C["bg"],
                                fg=C["ink_faint"], font=("Segoe UI", 8),
                                anchor="w", wraplength=520, justify="left")
            note_lbl.pack(fill="x", padx=44, pady=(0, 2))

        make_hairline(parent, bg=C["surface2"]).pack(fill="x", padx=44)

        def _refresh_visual(done):
            self._draw_checkbox(chk, done, color)
            if done:
                name_lbl.configure(fg=C["ink_faint"],
                                   font=("Segoe UI", 10, "overstrike"))
            else:
                name_lbl.configure(fg=C["ink"],
                                   font=("Segoe UI", 10))

        _refresh_visual(self._data.is_done(rid))

        def toggle(e=None):
            new = not self._data.is_done(rid)
            self._data.set_done(rid, new)
            _refresh_visual(new)
            home = self.controller.frames.get("Home")
            if home and hasattr(home, "refresh_rotinas"):
                home.refresh_rotinas()

        for w in [row, inner_row, chk, name_lbl, dot]:
            try: w.bind("<Button-1>", toggle)
            except Exception: pass

    @staticmethod
    def _draw_checkbox(canvas, checked, color):
        canvas.delete("all")
        if checked:
            canvas.create_oval(1, 1, 19, 19, fill=color, outline="")
            canvas.create_text(10, 10, text="✓", fill=C["bg"],
                               font=("Segoe UI", 9, "bold"))
        else:
            canvas.create_oval(1, 1, 19, 19, fill="", outline=C["ink_faint"],
                               width=1.5)

    def _build_gerenciar(self, parent):
        act = tk.Frame(parent, bg=C["bg"])
        act.pack(fill="x", padx=44, pady=(18, 2))
        styled_button(act, "+ Nova rotina", self._open_nova_dialog,
                      accent=True, small=True).pack(side="right")

        sf = ScrollableFrame(parent, bg=C["bg"])
        sf.pack(fill="both", expand=True)
        self._ger_inner = sf.inner
        self._ger_sf    = sf
        self._ger_inner.configure(bg=C["bg"])
        self._refresh_gerenciar()

    def _refresh_gerenciar(self):
        inner = getattr(self, "_ger_inner", None)
        if inner is None:
            return
        for w in inner.winfo_children():
            w.destroy()
        rots = self._data.rotinas()
        if not rots:
            e = tk.Frame(inner, bg=C["bg"])
            e.pack(fill="x", padx=44, pady=60)
            tk.Label(e, text="◈", bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 24)).pack()
            tk.Label(e, text="Nenhuma rotina cadastrada", bg=C["bg"],
                     fg=C["ink_faint"], font=("Segoe UI", 11)).pack(pady=(8, 0))
            styled_button(e, "+ Nova rotina", self._open_nova_dialog,
                          accent=True).pack(pady=(16, 0))
        else:
            for rot in rots:
                self._make_ger_row(inner, rot)
            tk.Frame(inner, bg=C["bg"], height=40).pack()
        sf = getattr(self, "_ger_sf", None)
        if sf:
            sf.refresh_bindings()

    @staticmethod
    def _dias_label(dias):
        if dias is None:
            return "Diária"
        s = sorted(dias)
        if s == list(range(7)):
            return "Diária"
        if s == list(range(5)):
            return "Dias úteis"
        abbr = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        return "  ".join(abbr[d] for d in s)

    def _make_ger_row(self, parent, rot):
        color    = rot.get("cor", C["accent"])
        nome     = rot.get("nome", "Rotina")
        dias     = rot.get("dias")
        alertas  = rot.get("alertas") or []
        if not alertas:
            h = (rot.get("hora_alerta") or "").strip()
            if h: alertas = [h]
        notas    = (rot.get("notas") or "").strip()
        dias_lbl = self._dias_label(dias)
        horas_lbl = "  ·  " + ", ".join(alertas) if alertas else ""

        row_o = tk.Frame(parent, bg=C["bg"])
        row_o.pack(fill="x")
        row_i = tk.Frame(row_o, bg=C["bg"], padx=44, pady=10)
        row_i.pack(fill="x")

        dot_col = tk.Frame(row_i, bg=C["bg"], width=20)
        dot_col.pack(side="left", fill="y")
        dot_col.pack_propagate(False)
        _make_dot(dot_col, color, 10, bg=C["bg"]).pack(anchor="n", pady=(5, 0))

        center = tk.Frame(row_i, bg=C["bg"])
        center.pack(side="left", fill="both", expand=True, padx=(10, 16))
        tk.Label(center, text=nome, bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 10, "bold"), anchor="w").pack(anchor="w")
        meta = dias_lbl + horas_lbl
        tk.Label(center, text=meta, bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI", 8), anchor="w").pack(anchor="w", pady=(2, 0))
        if notas:
            tk.Label(center, text=notas, bg=C["bg"], fg=C["ink_faint"],
                     font=("Segoe UI", 8), anchor="w",
                     wraplength=380, justify="left").pack(anchor="w", pady=(1, 0))

        right = tk.Frame(row_i, bg=C["bg"])
        right.pack(side="right", fill="y", pady=(2, 0))
        styled_button(right, "Editar",
                      lambda r=rot: self._open_edit_dialog(r),
                      small=True).pack(side="left")
        styled_button(right, "✕",
                      lambda r=rot: self._del_rotina(r),
                      danger=True, small=True).pack(side="left", padx=(4, 0))

        make_hairline(parent, bg=C["hair"]).pack(fill="x", padx=44)

        def _bg_all(bg):
            for w in [row_o, row_i, dot_col, center, right]:
                try: w.configure(bg=bg)
                except Exception: pass
            for w in center.winfo_children() + right.winfo_children() + \
                     dot_col.winfo_children():
                try: w.configure(bg=bg)
                except Exception: pass
        row_o.bind("<Enter>", lambda _: _bg_all(C["surface"]))
        row_o.bind("<Leave>", lambda _: _bg_all(C["bg"]))
        row_i.bind("<Enter>", lambda _: _bg_all(C["surface"]))
        row_i.bind("<Leave>", lambda _: _bg_all(C["bg"]))

    def _del_rotina(self, rot):
        if messagebox.askyesno("Remover", f"Remover '{rot['nome']}'?", parent=self):
            self._data.delete_rotina(rot["id"])
            self._refresh_gerenciar()
            home = self.controller.frames.get("Home")
            if home and hasattr(home, "refresh_rotinas"):
                home.refresh_rotinas()

    def _open_nova_dialog(self):
        self._open_rotina_dialog(None)

    def _open_edit_dialog(self, rot):
        self._open_rotina_dialog(rot)

    def _open_rotina_dialog(self, rot):
        editing = rot is not None
        dlg = tk.Toplevel(self)
        dlg.title("Editar Rotina" if editing else "Nova Rotina")
        dlg.configure(bg=C["surface"])
        dlg.geometry("480x580")
        dlg.resizable(False, True)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=C["surface"], padx=24)
        hdr.pack(fill="x", pady=(22, 0))
        color_var = tk.StringVar(value=(rot or {}).get("cor", C["accent"]))

        title_row = tk.Frame(hdr, bg=C["surface"])
        title_row.pack(fill="x")
        title_dot = _make_dot(title_row, color_var.get(), 12, bg=C["surface"])
        title_dot.pack(side="left", pady=(3, 0))
        tk.Label(title_row,
                 text="Editar Rotina" if editing else "Nova Rotina",
                 bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 14, "bold")).pack(side="left", padx=(8, 0))
        make_hairline(dlg, bg=C["hair"]).pack(fill="x", pady=(18, 0))

        sf   = ScrollableFrame(dlg, bg=C["surface"])
        sf.pack(fill="both", expand=True)
        sf.link_wheel(dlg)
        form = sf.inner
        form.configure(bg=C["surface"])
        pad  = tk.Frame(form, bg=C["surface"])
        pad.pack(fill="both", expand=True, padx=24)

        tk.Label(pad, text="NOME", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(16, 0))
        nome_var = tk.StringVar(value=(rot or {}).get("nome", ""))
        styled_entry(pad, textvariable=nome_var).pack(fill="x", pady=(4, 0))

        tk.Label(pad, text="DIAS DA SEMANA", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(16, 0))

        rot_dias = (rot or {}).get("dias")
        if rot_dias is None:
            freq0 = (rot or {}).get("frequencia", "diaria")
            if freq0 == "dias_uteis":
                rot_dias = [0, 1, 2, 3, 4]
            elif freq0.startswith("semanal_"):
                try:
                    rot_dias = [int(freq0.split("_")[1])]
                except Exception:
                    rot_dias = [0, 1, 2, 3, 4, 5, 6]
            else:
                rot_dias = [0, 1, 2, 3, 4, 5, 6]

        dias_sel  = set(rot_dias)
        DIA_ABBR  = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        dias_btns = {}

        dias_row = tk.Frame(pad, bg=C["surface"])
        dias_row.pack(anchor="w", pady=(6, 0))

        def _render_dias_btns():
            for d, btn in dias_btns.items():
                sel = d in dias_sel
                btn.configure(
                    bg=C["accent"]   if sel else C["surface3"],
                    fg=C["bg"]       if sel else C["ink_muted"],
                    font=("Segoe UI", 8, "bold") if sel else ("Segoe UI", 8),
                )

        def _toggle_dia(d):
            if d in dias_sel:
                if len(dias_sel) > 1:
                    dias_sel.discard(d)
            else:
                dias_sel.add(d)
            _render_dias_btns()

        for d, lbl in enumerate(DIA_ABBR):
            sel = d in dias_sel
            btn = tk.Button(dias_row, text=lbl,
                            command=lambda dd=d: _toggle_dia(dd),
                            bg=C["accent"] if sel else C["surface3"],
                            fg=C["bg"] if sel else C["ink_muted"],
                            font=("Segoe UI", 8, "bold") if sel else ("Segoe UI", 8),
                            relief="flat", bd=0, padx=9, pady=5,
                            cursor="hand2")
            btn.pack(side="left", padx=(0, 3))
            dias_btns[d] = btn

        preset_row = tk.Frame(pad, bg=C["surface"])
        preset_row.pack(anchor="w", pady=(5, 0))

        def _preset(days):
            dias_sel.clear()
            dias_sel.update(days)
            _render_dias_btns()

        for ptxt, pdays in [("Todos os dias", range(7)), ("Dias úteis", range(5))]:
            tk.Button(preset_row, text=ptxt,
                      command=lambda d=pdays: _preset(d),
                      bg=C["surface3"], fg=C["ink_faint"],
                      font=("Segoe UI", 7), relief="flat", bd=0,
                      padx=6, pady=3, cursor="hand2").pack(side="left", padx=(0, 5))

        tk.Label(pad, text="ALERTAS", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(16, 0))

        rot_als = list((rot or {}).get("alertas") or [])
        if not rot_als:
            h0 = (rot or {}).get("hora_alerta") or ""
            if h0:
                rot_als = [h0]
        alertas_list = list(rot_als)

        alertas_frame = tk.Frame(pad, bg=C["surface"])
        alertas_frame.pack(fill="x", pady=(4, 0))

        def _render_alertas():
            for w in alertas_frame.winfo_children():
                w.destroy()
            if not alertas_list:
                tk.Label(alertas_frame, text="Nenhum alerta configurado",
                         bg=C["surface"], fg=C["ink_faint"],
                         font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))
                return
            for i, hora in enumerate(alertas_list):
                r = tk.Frame(alertas_frame, bg=C["surface"])
                r.pack(anchor="w", pady=2)
                tk.Label(r, text=hora,
                         bg=C["surface2"], fg=C["ink"],
                         font=("Segoe UI", 9, "bold"),
                         padx=10, pady=3).pack(side="left")
                tk.Button(r, text="✕",
                          command=lambda idx=i: (_del_alerta(idx)),
                          bg=C["surface"], fg=C["ink_faint"],
                          activebackground=C["surface"], activeforeground=C["err"],
                          font=("Segoe UI", 8), relief="flat", bd=0,
                          padx=4, cursor="hand2").pack(side="left", padx=(3, 0))

        def _del_alerta(idx):
            if 0 <= idx < len(alertas_list):
                alertas_list.pop(idx)
                _render_alertas()

        _render_alertas()

        make_hairline(pad, bg=C["surface3"]).pack(fill="x", pady=(10, 0))
        add_lbl = tk.Frame(pad, bg=C["surface"])
        add_lbl.pack(fill="x", pady=(8, 0))
        tk.Label(add_lbl, text="Nova hora", bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 8)).pack(side="left")

        add_row = tk.Frame(pad, bg=C["surface"])
        add_row.pack(fill="x", pady=(4, 0))
        add_hora_var = tk.StringVar()
        styled_entry(add_row, textvariable=add_hora_var, width=8).pack(side="left")
        tk.Label(add_row, text="HH:MM", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7)).pack(side="left", padx=(5, 12))

        def _add_hora():
            h = add_hora_var.get().strip()
            if not re.match(r"^\d{2}:\d{2}$", h):
                return
            if h not in alertas_list:
                alertas_list.append(h)
                alertas_list.sort()
            add_hora_var.set("")
            _render_alertas()

        tk.Button(add_row, text="+ Adicionar", command=_add_hora,
                  bg=C["surface3"], fg=C["ink_muted"],
                  activebackground=C["accent_dim"], activeforeground=C["ink"],
                  font=("Segoe UI", 8), relief="flat", bd=0,
                  padx=8, pady=3, cursor="hand2").pack(side="left")

        antes_row = tk.Frame(pad, bg=C["surface"])
        antes_row.pack(fill="x", pady=(8, 0))
        antes_var = tk.StringVar(value="15")
        tk.Label(antes_row, text="ou", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        styled_entry(antes_row, textvariable=antes_var, width=4).pack(side="left")
        tk.Label(antes_row, text="min antes de", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(side="left", padx=(5, 5))
        ref_var = tk.StringVar()
        styled_entry(antes_row, textvariable=ref_var, width=8).pack(side="left")
        tk.Label(antes_row, text="HH:MM", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7)).pack(side="left", padx=(4, 8))

        def _add_antes():
            try:
                mins = int(antes_var.get().strip())
                ref  = ref_var.get().strip()
                if not re.match(r"^\d{2}:\d{2}$", ref):
                    return
                h, m  = int(ref[:2]), int(ref[3:])
                total = (h * 60 + m - mins) % (24 * 60)
                nova  = f"{total // 60:02d}:{total % 60:02d}"
                if nova not in alertas_list:
                    alertas_list.append(nova)
                    alertas_list.sort()
                ref_var.set("")
                _render_alertas()
            except Exception:
                pass

        tk.Button(antes_row, text="OK", command=_add_antes,
                  bg=C["surface3"], fg=C["ink_muted"],
                  activebackground=C["accent_dim"], activeforeground=C["ink"],
                  font=("Segoe UI", 8), relief="flat", bd=0,
                  padx=8, pady=3, cursor="hand2").pack(side="left")

        make_hairline(pad, bg=C["surface3"]).pack(fill="x", pady=(10, 0))

        tk.Label(pad, text="COR", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(14, 0))
        color_row  = tk.Frame(pad, bg=C["surface"])
        color_row.pack(anchor="w", pady=(6, 0))
        color_btns = {}

        def _pick(c_val):
            color_var.set(c_val)
            title_dot.itemconfig(1, fill=c_val)
            for cv2, cb2 in color_btns.items():
                r2 = 2 if cv2 == c_val else 0
                cb2.configure(highlightthickness=r2,
                              highlightbackground=C["ink"] if cv2 == c_val else C["hair"])

        for dc in DOT_COLORS:
            cv = tk.Canvas(color_row, width=20, height=20, bg=C["surface"],
                           highlightthickness=2 if dc == color_var.get() else 0,
                           highlightbackground=C["ink"] if dc == color_var.get() else C["hair"],
                           cursor="hand2")
            cv.pack(side="left", padx=3)
            cv.create_oval(3, 3, 17, 17, fill=dc, outline="")
            cv.bind("<Button-1>", lambda _, d=dc: _pick(d))
            color_btns[dc] = cv

        tk.Label(pad, text="NOTAS  (opcional)", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(14, 0))
        notas_var = tk.StringVar(value=(rot or {}).get("notas", "") or "")
        styled_entry(pad, textvariable=notas_var).pack(fill="x", pady=(4, 12))

        make_hairline(dlg, bg=C["hair"]).pack(fill="x")
        foot = tk.Frame(dlg, bg=C["surface"], padx=24, pady=12)
        foot.pack(fill="x")

        def salvar():
            nome = nome_var.get().strip()
            if not nome:
                messagebox.showwarning("Campo obrigatório", "Informe um nome.",
                                       parent=dlg)
                return
            dias_final = sorted(dias_sel)
            als_final  = alertas_list[:]

            if editing:
                self._data.update_rotina(
                    rot["id"], nome=nome, dias=dias_final,
                    alertas=als_final, cor=color_var.get(),
                    notas=notas_var.get().strip()
                )
            else:
                self._data.add_rotina(nome, dias_final, als_final,
                                      color_var.get(), notas_var.get().strip())
            dlg.destroy()
            self._refresh_gerenciar()
            home = self.controller.frames.get("Home")
            if home and hasattr(home, "refresh_rotinas"):
                home.refresh_rotinas()

        if editing:
            def _excluir():
                if messagebox.askyesno("Excluir", f"Remover '{rot['nome']}'?",
                                       parent=dlg):
                    self._data.delete_rotina(rot["id"])
                    dlg.destroy()
                    self._refresh_gerenciar()
                    home = self.controller.frames.get("Home")
                    if home and hasattr(home, "refresh_rotinas"):
                        home.refresh_rotinas()
            styled_button(foot, "Excluir", _excluir,
                          danger=True, small=True).pack(side="left")

        styled_button(foot, "Cancelar", dlg.destroy, small=True).pack(
            side="right", padx=(6, 0))
        styled_button(foot, "Salvar", salvar, accent=True, small=True).pack(side="right")

    # ── Alertas ──────────────────────────────────────────────────────────────
    def _start_alert_checker(self):
        self._alerted_flags: set = set()
        self._alerted_date:  str = ""
        self._check_alerts()

    def _check_alerts(self):
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return

        now       = datetime.now()
        now_str   = now.strftime("%H:%M")
        today_str = now.strftime("%Y-%m-%d")

        if self._alerted_date != today_str:
            self._alerted_flags = set()
            self._alerted_date  = today_str

        for rot in self._data.today_rotinas():
            if self._data.is_done(rot["id"]):
                continue
            alertas = rot.get("alertas") or []
            if not alertas:
                h = (rot.get("hora_alerta") or "").strip()
                if h:
                    alertas = [h]
            for hora in alertas:
                if hora.strip() != now_str:
                    continue
                flag = (rot["id"], hora)
                if flag in self._alerted_flags:
                    continue
                self._alerted_flags.add(flag)
                nome = rot["nome"]
                self.after(0, lambda n=nome, h=hora: self._show_alert(n, h))

        self.after(30_000, self._check_alerts)

    # ── PATCH 1: popup de alerta visual customizado ──────────────────────────
    def _show_alert(self, nome, hora):
        try:
            import winsound
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception:
            pass

        try:
            if not self.controller.winfo_exists():
                return
        except Exception:
            return

        dlg = tk.Toplevel(self.controller)
        dlg.title("Lembrete")
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.overrideredirect(True)
        dlg.attributes("-topmost", True)
        dlg.lift()
        dlg.focus_force()

        # Força ao topo no Windows — aparece sobre qualquer app aberto
        if sys.platform == "win32":
            try:
                import ctypes as _ctypes
                _hwnd = _ctypes.windll.user32.GetParent(dlg.winfo_id())
                if not _hwnd:
                    _hwnd = dlg.winfo_id()
                _ctypes.windll.user32.SetForegroundWindow(_hwnd)
                _ctypes.windll.user32.BringWindowToTop(_hwnd)
                _ctypes.windll.user32.FlashWindow(_hwnd, True)
            except Exception:
                pass

        # Borda laranja no topo (identidade visual do app)
        tk.Frame(dlg, bg=C["accent"], height=3).pack(fill="x")

        body = tk.Frame(dlg, bg=C["surface"], padx=28, pady=20)
        body.pack(fill="both", expand=True)

        # Cabeçalho: sino + hora
        hdr_row = tk.Frame(body, bg=C["surface"])
        hdr_row.pack(fill="x")
        tk.Label(hdr_row, text="\U0001f514", bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 18)).pack(side="left")
        tk.Label(hdr_row, text=hora, bg=C["surface"], fg=C["accent"],
                 font=("Segoe UI", 22, "bold")).pack(side="left", padx=(10, 0))

        make_hairline(body, bg=C["hair"]).pack(fill="x", pady=(14, 12))

        # Nome da rotina
        tk.Label(body, text=nome, bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 13, "bold"),
                 wraplength=300, justify="left").pack(anchor="w")
        tk.Label(body, text="Rotina programada para agora", bg=C["surface"],
                 fg=C["ink_muted"], font=("Segoe UI", 8)).pack(anchor="w", pady=(4, 0))

        make_hairline(body, bg=C["hair"]).pack(fill="x", pady=(16, 0))

        foot = tk.Frame(body, bg=C["surface"])
        foot.pack(fill="x", pady=(12, 0))

        def _dismiss():
            dlg.destroy()

        styled_button(foot, "Dispensar", _dismiss, small=True).pack(side="right")
        styled_button(foot, "Ok, entendido", _dismiss, accent=True, small=True).pack(
            side="right", padx=(0, 6))

        # Centraliza na tela
        dlg.update_idletasks()
        sw = dlg.winfo_screenwidth()
        sh = dlg.winfo_screenheight()
        w  = dlg.winfo_reqwidth()
        h  = dlg.winfo_reqheight()
        dlg.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

        dlg.bind("<Escape>", lambda _e: _dismiss())
        dlg.grab_set()


class BPMHubFrame(tk.Frame):
    """Hub do BPM — escolha entre BPM Invertido (clientes/valores fixos da
    mesa) e BPM Nova Plataforma (CNPJ/agência/conta/plataforma/valor
    variáveis, informados pelo usuário)."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=44, pady=(36, 0))
        eyebrow_label(hdr, "MESA DE OPERAÇÕES").pack(anchor="w")
        tk.Label(hdr, text="BPM", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 22, "bold")).pack(anchor="w", pady=(6, 0))
        tk.Label(hdr, text="Escolha a plataforma para abrir a solicitação.",
                 bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", pady=(20, 0))

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=44, pady=(32, 0))
        for c in range(3):
            body.columnconfigure(c, weight=1, uniform="bpmhub")

        make_hub_option_card(
            body, 0, 0, "⬡", "BPM Invertido",
            "Clientes e valores de Operações Invertido já mapeados.",
            lambda: self.controller.show_frame("BPM_CONFIG"), C["accent"])

        make_hub_option_card(
            body, 0, 1, "⚡", "BPM Nova Plataforma",
            "Informe CNPJ, agência, conta, plataforma e valor da operação.",
            lambda: self.controller.show_frame("BPM_CONFIG_NOVA"), "#EC7000")

    def on_show(self):
        pass


class BPMConfigFrame(tk.Frame):
    CLIENTS = list(BPM_CLIENT_DATA.keys())

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._selected  = {}
        self._func_var  = tk.StringVar()
        self._senha_var = tk.StringVar()
        self._func_var.trace_add(
            "write", lambda *_: setattr(self.controller, "bpm_funcional", self._func_var.get()))
        self._senha_var.trace_add(
            "write", lambda *_: setattr(self.controller, "bpm_password", self._senha_var.get()))
        self._build()

    def _only_digits(self, s):
        return "".join(c for c in (s or "") if c.isdigit())

    def on_show(self):
        self._func_var.set(getattr(self.controller, "bpm_funcional", "") or "")
        self._senha_var.set(getattr(self.controller, "bpm_password", "") or "")
        if hasattr(self, "_sf"):
            self._sf.refresh_bindings()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24,0))
        tk.Label(hdr, text="Configurar BPM Invertido", bg=C["bg"], fg=C["ink"],
                 font=("Georgia",18,"bold")).pack(side="left")
        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(14,0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        body = self._sf.inner
        body.configure(bg=C["bg"])

        sec = card_frame(body)
        sec.pack(fill="x", padx=32, pady=(20,0))
        tk.Label(sec, text="Credenciais do Painel de Serviços", bg=C["surface"],
                 fg=C["ink"], font=("Segoe UI",10,"bold")).pack(anchor="w", padx=18, pady=(14,0))
        tk.Label(sec, text="Funcional e senha para login no Painel BPM",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI",8)).pack(anchor="w", padx=18, pady=(2,10))
        make_hairline(sec, bg=C["hair"]).pack(fill="x")

        cred_form = tk.Frame(sec, bg=C["surface"], padx=18, pady=14)
        cred_form.pack(fill="x")
        cred_form.columnconfigure(0, weight=1)
        cred_form.columnconfigure(1, weight=1)

        fc = tk.Frame(cred_form, bg=C["surface"])
        fc.grid(row=0, column=0, sticky="ew", padx=(0,8))
        tk.Label(fc, text="Funcional (somente números)", bg=C["surface"],
                 fg=C["ink_muted"], font=("Segoe UI",8)).pack(anchor="w")
        self._ent_func = styled_entry(fc, textvariable=self._func_var)
        self._ent_func.pack(fill="x", pady=(4,0))

        sc2 = tk.Frame(cred_form, bg=C["surface"])
        sc2.grid(row=0, column=1, sticky="ew", padx=(8,0))
        tk.Label(sc2, text="Senha (até 6 dígitos)", bg=C["surface"],
                 fg=C["ink_muted"], font=("Segoe UI",8)).pack(anchor="w")
        self._ent_senha = styled_entry(sc2, textvariable=self._senha_var, show="•")
        self._ent_senha.pack(fill="x", pady=(4,0))

        def _key_func(e):
            if e.keysym in ("BackSpace","Delete","Tab","Return","Left","Right","Home","End"): return
            if e.state & 0x4: return
            if e.char and e.char.isdigit(): return
            if e.char: return "break"

        def _key_senha(e):
            if e.keysym in ("BackSpace","Delete","Tab","Return","Left","Right","Home","End"): return
            if e.state & 0x4: return
            if not e.char: return
            if not e.char.isdigit(): return "break"
            if len(self._ent_senha.get()) >= 6 and not self._ent_senha.selection_present():
                return "break"

        self._ent_func.bind("<KeyPress>", _key_func)
        self._ent_senha.bind("<KeyPress>", _key_senha)

        sec2 = card_frame(body)
        sec2.pack(fill="x", padx=32, pady=(16,0))
        tk.Label(sec2, text="Clientes e Valores", bg=C["surface"],
                 fg=C["ink"], font=("Segoe UI",10,"bold")).pack(anchor="w", padx=18, pady=(14,0))
        tk.Label(sec2, text="Selecione os clientes e informe o valor da operação",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI",8)).pack(anchor="w", padx=18, pady=(2,10))
        make_hairline(sec2, bg=C["hair"]).pack(fill="x")

        clients_frame = tk.Frame(sec2, bg=C["surface"], padx=18, pady=12)
        clients_frame.pack(fill="x")

        self._client_rows = {}
        for cli in self.CLIENTS:
            self._make_client_row(clients_frame, cli)

        make_hairline(body, bg=C["hair"]).pack(fill="x", padx=32, pady=(20,0))

        foot = tk.Frame(body, bg=C["bg"])
        foot.pack(fill="x", padx=32, pady=16)
        styled_button(foot, "← Voltar",
                      lambda: self.controller.show_frame("BPM_HUB")).pack(side="left")
        styled_button(foot, "Trocar p/ Nova Plataforma →",
                      lambda: self.controller.show_frame("BPM_CONFIG_NOVA")).pack(side="left", padx=(8, 0))
        self._run_btn = styled_button(foot, "▶  Iniciar BPM",
                                      self._start_bpm, accent=True)
        self._run_btn.pack(side="right")

    def _make_client_row(self, parent, cli):
        row = tk.Frame(parent, bg=C["surface"], pady=3)
        row.pack(fill="x")
        row.columnconfigure(1, weight=1)

        selected = tk.BooleanVar(value=False)
        val_var  = tk.StringVar(value="R$ 0,00")
        val_digits = [""]

        cb = tk.Checkbutton(row, variable=selected, bg=C["surface"],
                            selectcolor=C["bg"], activebackground=C["surface"],
                            fg=C["ink_muted"], font=("Segoe UI",9))
        cb.grid(row=0, column=0, sticky="w")

        name_lbl = tk.Label(row, text=cli, bg=C["surface"], fg=C["ink_muted"],
                            font=("Segoe UI",9))
        name_lbl.grid(row=0, column=1, sticky="w", padx=(4,8))

        ent = styled_entry(row, textvariable=val_var, width=14)
        ent.grid(row=0, column=2, sticky="e")
        ent.configure(state="disabled", disabledbackground=C["bg"], disabledforeground=C["ink_faint"])

        def fmt_val():
            if not val_digits[0]: val_var.set("R$ 0,00"); return
            d = Decimal(int(val_digits[0])) / Decimal("100")
            val_var.set(_fmt_brl(d))

        def on_key(e):
            if e.keysym == "BackSpace": val_digits[0] = val_digits[0][:-1]; fmt_val(); return "break"
            if e.char and e.char.isdigit(): val_digits[0] += e.char; fmt_val(); return "break"
            if e.keysym in {"Tab","Left","Right","Home","End"}: return
            return "break"

        def on_paste(_):
            try: clip = row.clipboard_get()
            except: return "break"
            d = _parse_brl(clip)
            if d is None: return "break"
            val_digits[0] = str(max(int((d*100).quantize(Decimal("1"))),0))
            fmt_val(); return "break"

        def on_toggle(*_):
            if selected.get():
                ent.configure(state="normal", bg=C["bg"], fg=C["ink"])
                self._selected[cli] = val_var
            else:
                ent.configure(state="disabled", disabledbackground=C["bg"], disabledforeground=C["ink_faint"])
                self._selected.pop(cli, None)

        cb.configure(command=on_toggle)
        ent.bind("<KeyPress>", on_key)
        ent.bind("<<Paste>>", on_paste)
        ent.bind("<Control-v>", on_paste)
        self._client_rows[cli] = {"selected": selected, "val_var": val_var, "val_digits": val_digits, "entry": ent}

    def _start_bpm(self):
        func = self._only_digits(self._func_var.get())
        senha = self._only_digits(self._senha_var.get())
        if not func:
            messagebox.showwarning("Funcional obrigatório", "Informe o funcional (somente números)."); return
        if not (1 <= len(senha) <= 6):
            messagebox.showwarning("Senha inválida", "Senha deve ter 1 a 6 dígitos."); return

        selection = []
        for cli in self.CLIENTS:
            row = self._client_rows[cli]
            if not row["selected"].get(): continue
            raw = row["val_var"].get()
            d = _parse_brl(raw)
            if d is None or d == Decimal("0.00"):
                messagebox.showwarning("Valor inválido", f"Informe um valor válido para {cli}."); return
            selection.append({"cliente": cli, "valor": _fmt_brl_from_raw(raw)})

        if not selection:
            messagebox.showwarning("Seleção vazia", "Selecione ao menos um cliente."); return

        self.controller.bpm_funcional    = func
        self.controller.bpm_password     = senha
        self.controller.bpm_run_selection = selection
        self.controller.bpm_run_mode      = "invertido"
        self.controller.show_frame("BPM")

class BPMNovaConfigFrame(tk.Frame):
    """Configuração de BPM para a Nova Plataforma — diferente do Invertido,
    aqui CNPJ, agência, conta, plataforma e valor são todos informados
    livremente pelo usuário, linha por linha."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._func_var  = tk.StringVar()
        self._senha_var = tk.StringVar()
        self._func_var.trace_add(
            "write", lambda *_: setattr(self.controller, "bpm_funcional", self._func_var.get()))
        self._senha_var.trace_add(
            "write", lambda *_: setattr(self.controller, "bpm_password", self._senha_var.get()))
        self._rows = []
        self._build()

    def on_show(self):
        self._func_var.set(getattr(self.controller, "bpm_funcional", "") or "")
        self._senha_var.set(getattr(self.controller, "bpm_password", "") or "")
        if hasattr(self, "_sf"):
            self._sf.refresh_bindings()
        if not self._rows:
            self._add_row()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24, 0))
        tk.Label(hdr, text="Configurar BPM — Nova Plataforma", bg=C["bg"], fg=C["ink"],
                 font=("Georgia", 18, "bold")).pack(side="left")
        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(14, 0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        body = self._sf.inner
        body.configure(bg=C["bg"])

        sec = card_frame(body)
        sec.pack(fill="x", padx=32, pady=(20, 0))
        tk.Label(sec, text="Credenciais do Painel de Serviços", bg=C["surface"],
                 fg=C["ink"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=18, pady=(14, 0))
        tk.Label(sec, text="Funcional e senha para login no Painel BPM",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 8)).pack(
                     anchor="w", padx=18, pady=(2, 10))
        make_hairline(sec, bg=C["hair"]).pack(fill="x")

        cred_form = tk.Frame(sec, bg=C["surface"], padx=18, pady=14)
        cred_form.pack(fill="x")
        cred_form.columnconfigure(0, weight=1)
        cred_form.columnconfigure(1, weight=1)

        fc = tk.Frame(cred_form, bg=C["surface"])
        fc.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Label(fc, text="Funcional (somente números)", bg=C["surface"],
                 fg=C["ink_muted"], font=("Segoe UI", 8)).pack(anchor="w")
        self._ent_func = styled_entry(fc, textvariable=self._func_var)
        self._ent_func.pack(fill="x", pady=(4, 0))

        sc2 = tk.Frame(cred_form, bg=C["surface"])
        sc2.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        tk.Label(sc2, text="Senha (até 6 dígitos)", bg=C["surface"],
                 fg=C["ink_muted"], font=("Segoe UI", 8)).pack(anchor="w")
        self._ent_senha = styled_entry(sc2, textvariable=self._senha_var, show="•")
        self._ent_senha.pack(fill="x", pady=(4, 0))

        def _key_func(e):
            if e.keysym in ("BackSpace", "Delete", "Tab", "Return", "Left", "Right", "Home", "End"):
                return
            if e.state & 0x4:
                return
            if e.char and e.char.isdigit():
                return
            if e.char:
                return "break"

        def _key_senha(e):
            if e.keysym in ("BackSpace", "Delete", "Tab", "Return", "Left", "Right", "Home", "End"):
                return
            if e.state & 0x4:
                return
            if not e.char:
                return
            if not e.char.isdigit():
                return "break"
            if len(self._ent_senha.get()) >= 6 and not self._ent_senha.selection_present():
                return "break"

        self._ent_func.bind("<KeyPress>", _key_func)
        self._ent_senha.bind("<KeyPress>", _key_senha)

        sec2 = card_frame(body)
        sec2.pack(fill="x", padx=32, pady=(16, 0))
        tk.Label(sec2, text="Operações", bg=C["surface"],
                 fg=C["ink"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=18, pady=(14, 0))
        tk.Label(sec2, text="Informe CNPJ, agência, conta, plataforma e valor de cada operação",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 8)).pack(
                     anchor="w", padx=18, pady=(2, 10))
        make_hairline(sec2, bg=C["hair"]).pack(fill="x")

        self._rows_wrap = tk.Frame(sec2, bg=C["surface"], padx=18, pady=12)
        self._rows_wrap.pack(fill="x")

        add_row_f = tk.Frame(sec2, bg=C["surface"], padx=18, pady=10)
        add_row_f.pack(fill="x")
        styled_button(add_row_f, "+ Adicionar operação", self._add_row,
                      small=True).pack(anchor="w")

        make_hairline(body, bg=C["hair"]).pack(fill="x", padx=32, pady=(20, 0))

        foot = tk.Frame(body, bg=C["bg"])
        foot.pack(fill="x", padx=32, pady=16)
        styled_button(foot, "← Voltar",
                      lambda: self.controller.show_frame("BPM_HUB")).pack(side="left")
        styled_button(foot, "Trocar p/ BPM Invertido →",
                      lambda: self.controller.show_frame("BPM_CONFIG")).pack(side="left", padx=(8, 0))
        self._run_btn = styled_button(foot, "▶  Iniciar BPM",
                                      self._start_bpm, accent=True)
        self._run_btn.pack(side="right")

    def _add_row(self):
        row_f = tk.Frame(self._rows_wrap, bg=C["surface"], pady=6)
        row_f.pack(fill="x")
        for c in range(5):
            row_f.columnconfigure(c, weight=1, uniform="novarow")

        def _field(col, label, max_len=None, min_len=None, valid_lengths=None,
                   hint_text=None):
            wrap = tk.Frame(row_f, bg=C["surface"])
            wrap.grid(row=0, column=col, sticky="new", padx=(0 if col == 0 else 6, 0))
            tk.Label(wrap, text=label, bg=C["surface"], fg=C["ink_muted"],
                     font=("Segoe UI", 7, "bold")).pack(anchor="w")
            var = tk.StringVar()
            ent = styled_entry(wrap, textvariable=var, width=14)
            ent.pack(fill="x", pady=(3, 0))
            hint = tk.Label(wrap, text=" ", bg=C["surface"], fg=C["warn"],
                            font=("Segoe UI", 7), anchor="w", justify="left")
            hint.pack(anchor="w", fill="x", pady=(2, 0))
            if max_len is not None:
                _bind_digits_only(ent, var, max_len, min_len=min_len, hint_lbl=hint,
                                  valid_lengths=valid_lengths, hint_text=hint_text)
            return var, ent, hint

        cnpj_var, cnpj_ent, _ = _field(
            0, "CNPJ", max_len=14, valid_lengths={11, 14},
            hint_text="número irregular (CPF: 11 ou CNPJ: 14 dígitos)")
        ag_var, ag_ent, _ = _field(
            1, "AGÊNCIA", max_len=4, min_len=4, hint_text="agência deve ter 4 dígitos")
        conta_var, conta_ent, conta_hint = _field(2, "CONTA")
        plat_var, plat_ent, _ = _field(
            3, "PLATAFORMA", max_len=4, min_len=4, hint_text="plataforma deve ter 4 dígitos")

        _bind_conta_field(conta_ent, conta_var, conta_hint)

        val_wrap = tk.Frame(row_f, bg=C["surface"])
        val_wrap.grid(row=0, column=4, sticky="new", padx=(6, 0))
        tk.Label(val_wrap, text="VALOR", bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w")
        val_inner = tk.Frame(val_wrap, bg=C["surface"])
        val_inner.pack(fill="x", pady=(3, 0))
        val_var = tk.StringVar(value="R$ 0,00")
        val_digits = [""]
        val_ent = styled_entry(val_inner, textvariable=val_var, width=12)
        val_ent.pack(side="left", fill="x", expand=True)

        def fmt_val():
            if not val_digits[0]:
                val_var.set("R$ 0,00")
                return
            d = Decimal(int(val_digits[0])) / Decimal("100")
            val_var.set(_fmt_brl(d))

        def on_key(e):
            if e.keysym == "BackSpace":
                val_digits[0] = val_digits[0][:-1]; fmt_val(); return "break"
            if e.char and e.char.isdigit():
                val_digits[0] += e.char; fmt_val(); return "break"
            if e.keysym in {"Tab", "Left", "Right", "Home", "End"}:
                return
            return "break"

        def on_paste(_):
            try:
                clip = row_f.clipboard_get()
            except Exception:
                return "break"
            d = _parse_brl(clip)
            if d is None:
                return "break"
            val_digits[0] = str(max(int((d * 100).quantize(Decimal("1"))), 0))
            fmt_val(); return "break"

        val_ent.bind("<KeyPress>", on_key)
        val_ent.bind("<<Paste>>", on_paste)
        val_ent.bind("<Control-v>", on_paste)

        rm_btn = styled_button(val_inner, "✕", lambda: self._remove_row(entry),
                               danger=True, small=True)
        rm_btn.pack(side="left", padx=(6, 0))

        make_hairline(row_f, bg=C["hair"]).grid(row=1, column=0, columnspan=5,
                                                sticky="ew", pady=(8, 0))

        entry = {
            "frame": row_f, "cnpj": cnpj_var, "agencia": ag_var,
            "conta": conta_var, "plataforma": plat_var,
            "val_var": val_var, "val_digits": val_digits,
        }
        self._rows.append(entry)
        return entry

    def _remove_row(self, entry):
        if len(self._rows) <= 1:
            messagebox.showinfo("Operações", "É necessário ao menos uma operação.")
            return
        entry["frame"].destroy()
        self._rows.remove(entry)

    def _start_bpm(self):
        func = only_digits(self._func_var.get())
        senha = only_digits(self._senha_var.get())
        if not func:
            messagebox.showwarning("Funcional obrigatório", "Informe o funcional (somente números).")
            return
        if not (1 <= len(senha) <= 6):
            messagebox.showwarning("Senha inválida", "Senha deve ter 1 a 6 dígitos.")
            return

        selection = []
        for entry in self._rows:
            cnpj = only_digits(entry["cnpj"].get())
            agencia = only_digits(entry["agencia"].get())
            conta = entry["conta"].get().strip()
            plataforma = entry["plataforma"].get().strip()
            raw_val = entry["val_var"].get()
            d = _parse_brl(raw_val)
            if not cnpj or not agencia or not conta or not plataforma:
                messagebox.showwarning(
                    "Dados incompletos",
                    "Preencha CNPJ, agência, conta e plataforma de todas as operações.")
                return
            if len(cnpj) not in (11, 14):
                messagebox.showwarning(
                    "CNPJ/CPF irregular",
                    "Informe um CPF (11 dígitos) ou CNPJ (14 dígitos) válido.")
                return
            if len(agencia) != 4:
                messagebox.showwarning(
                    "Agência irregular", "A agência deve ter exatamente 4 dígitos.")
                return
            if len(plataforma) != 4:
                messagebox.showwarning(
                    "Plataforma irregular", "A plataforma deve ter exatamente 4 dígitos.")
                return
            if "-" not in conta or len(only_digits(conta)) < 6:
                messagebox.showwarning(
                    "Conta incompleta",
                    "Informe a conta com o dígito verificador (ex.: 99451-8).")
                return
            if d is None or d == Decimal("0.00"):
                messagebox.showwarning("Valor inválido", "Informe um valor válido para cada operação.")
                return
            selection.append({
                "cliente": f"CNPJ {cnpj}",
                "valor": _fmt_brl_from_raw(raw_val),
                "CNPJ": cnpj, "AG": agencia, "CONTA": conta, "PLATAFORMA": plataforma,
            })

        if not selection:
            messagebox.showwarning("Seleção vazia", "Adicione ao menos uma operação.")
            return

        self.controller.bpm_funcional    = func
        self.controller.bpm_password     = senha
        self.controller.bpm_run_selection = selection
        self.controller.bpm_run_mode      = "nova_plataforma"
        self.controller.show_frame("BPM")


class ShareFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self.vars = {k: tk.StringVar() for k in [
            "razao_social","cnpj","conta","plataforma","regiao",
            "trader_espec","valor_operacao","spread","prazo_min","prazo_max","modalidade"
        ]}
        self.hidden = {"premio":"sem prêmio","liquidacao":"Débito em CC","cnpj8":""}
        self.vars["modalidade"].trace_add("write", lambda *_: self._update_resumo())
        self.vars["regiao"].trace_add("write",    lambda *_: self._update_trader_espec())
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24,0))
        tk.Label(hdr, text="Cadastro Share", bg=C["bg"], fg=C["ink"],
                 font=("Georgia",18,"bold")).pack(side="left")
        styled_button(hdr, "📄  Abrir PDF…", self._on_open_pdf, accent=True).pack(side="right")
        styled_button(hdr, "← Voltar",
                      lambda: self.controller.show_frame("Home")).pack(side="right", padx=(0,6))
        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(14,0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        body = self._sf.inner
        body.configure(bg=C["bg"])

        fields_card = card_frame(body)
        fields_card.pack(fill="x", padx=32, pady=(20,0))
        tk.Label(fields_card, text="Campos extraídos", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI",10,"bold")).pack(anchor="w", padx=18, pady=(14,0))
        tk.Label(fields_card, text="Edite se necessário antes de gerar o resumo.",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI",8)).pack(anchor="w", padx=18, pady=(2,10))
        make_hairline(fields_card, bg=C["hair"]).pack(fill="x")

        grid = tk.Frame(fields_card, bg=C["surface"], padx=18, pady=14)
        grid.pack(fill="x")
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)

        def field(key, label, row, col, combo=False, width=28, readonly=False):
            b = tk.Frame(grid, bg=C["surface"])
            b.grid(row=row, column=col, sticky="ew", padx=(0 if col==0 else 10, 0), pady=6)
            tk.Label(b, text=label, bg=C["surface"], fg=C["ink_muted"],
                     font=("Segoe UI",8)).pack(anchor="w", pady=(0,4))
            row_f = tk.Frame(b, bg=C["surface"])
            row_f.pack(fill="x")
            row_f.columnconfigure(0, weight=1)
            if combo:
                w = ttk.Combobox(row_f, textvariable=self.vars[key],
                                 values=["Autorizadas CSV","Autorizadas SISPAG"],
                                 state="normal", width=width)
                w.grid(row=0, column=0, sticky="ew", padx=(0,6))
            else:
                e = styled_entry(row_f, textvariable=self.vars[key], width=width)
                if readonly: e.configure(state="readonly", readonlybackground=C["bg"])
                e.grid(row=0, column=0, sticky="ew", padx=(0,6))
            styled_button(row_f, "↗", lambda k=key: self._copy(k), small=True).grid(row=0,column=1)

        specs = [
            ("razao_social","Razão Social",0,0,False,28,False),
            ("cnpj","CNPJ",0,1,False,22,False),
            ("conta","Conta Corrente",1,0,False,22,False),
            ("valor_operacao","Valor da Operação",1,1,False,22,False),
            ("plataforma","Plataforma",2,0,False,14,False),
            ("regiao","Região da Plataforma",2,1,False,14,False),
            ("trader_espec","Trader / Espec",3,0,False,28,True),
            ("spread","Spread",3,1,False,14,False),
            ("prazo_min","Prazo Mínimo NF",4,0,False,14,False),
            ("prazo_max","Prazo Máximo NF",4,1,False,14,False),
            ("modalidade","Modalidade",5,0,True,28,False),
        ]
        for k,l,r,c,combo,w,ro in specs:
            field(k,l,r,c,combo,w,ro)

        act = tk.Frame(body, bg=C["bg"])
        act.pack(fill="x", padx=32, pady=(14,0))
        styled_button(act,"🔄  Gerar Resumo",  self._update_resumo, accent=True).pack(side="left")
        styled_button(act,"🧽  Limpar",          self._clear_all).pack(side="left", padx=(6,0))

        res_card = card_frame(body)
        res_card.pack(fill="x", padx=32, pady=(14,0))
        tk.Label(res_card, text="Resumo", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI",10,"bold")).pack(anchor="w", padx=18, pady=(14,0))
        make_hairline(res_card, bg=C["hair"]).pack(fill="x")

        txt_wrap = tk.Frame(res_card, bg=C["surface"], padx=18, pady=14)
        txt_wrap.pack(fill="x")
        txt_wrap.columnconfigure(0, weight=1)
        self.txt_resumo = tk.Text(txt_wrap, height=7, wrap="word", bd=0, relief="flat",
                                  bg=C["bg"], fg=C["ink"], insertbackground=C["accent"],
                                  highlightthickness=1, highlightbackground=C["hair"],
                                  font=("Segoe UI",9), padx=10, pady=8)
        self.txt_resumo.grid(row=0, column=0, sticky="ew", padx=(0,6))
        sc = MinimalScrollbar(txt_wrap, command=self.txt_resumo.yview, bg=C["surface"])
        sc.grid(row=0, column=1, sticky="ns")
        self.txt_resumo.configure(yscrollcommand=sc.set)
        bind_text_mousewheel(self.txt_resumo)

        foot = tk.Frame(body, bg=C["bg"])
        foot.pack(fill="x", padx=32, pady=(12,30))
        styled_button(foot,"📋  Copiar Resumo", self._copy_resumo, accent=True).pack(side="left")

    def on_show(self):
        if hasattr(self, "_sf"):
            self._sf.refresh_bindings()

    def _copy(self, k):
        v = self.vars[k].get()
        self.clipboard_clear(); self.clipboard_append(v)

    def _copy_resumo(self):
        v = self.txt_resumo.get("1.0","end-1c")
        self.clipboard_clear(); self.clipboard_append(v)

    def _update_trader_espec(self):
        reg = self.vars["regiao"].get().strip()
        tr, sp = trader_espec_from_regiao(reg)
        self.vars["trader_espec"].set(f"{tr} / {sp}".strip(" /") if (tr or sp) else "")
        self._update_resumo()

    def _on_open_pdf(self):
        p = filedialog.askopenfilename(title="Selecione um PDF",
                                       filetypes=[("PDF","*.pdf")])
        if not p: return
        try:
            tl, tp = extract_text_from_pdf(p)
            t = tl
            if not t.strip():
                messagebox.showwarning("PDF sem texto","Não foi possível extrair texto."); return
            lines = t.splitlines()
            tn, tc = normalize_text_variants(t)
            pmap = _plain_label_value_map(p)
            raz  = extract_razao_social(t,lines,tn,tc,t_plain=tp,pdf_path=p) or ""
            cn   = extract_cnpj(t,tn,tc,pmap) or ""
            conta= extract_conta_corrente(lines,tn,tc,pmap) or ""
            plat = extract_plataforma(t,tn,tc,pmap) or ""
            reg  = extract_regiao(t,tn,tc,pmap) or ""
            val  = extract_valor(t,tn,tc,pmap) or ""
            sp   = extract_spread(t,tn,tc,pmap) or ""
            pmin = extract_prazo_min(t,tn,tc,pmap) or ""
            pmax = extract_prazo_max(t,tn,tc,pmap) or ""
            mod  = extract_modalidade(t,tn,tc,pmap) or ""
            liq  = "Débito em CC" if not RE_LIQ_CRED.search(t) else "Crédito em CC"
            prem = "com prêmio" if RE_PREMIO.search(t) else "sem prêmio"
            self.vars["razao_social"].set(raz)
            self.vars["cnpj"].set(only_digits(cn) if cn else "")
            self.vars["conta"].set(conta)
            self.vars["plataforma"].set(plat)
            self.vars["regiao"].set(reg)
            self.vars["valor_operacao"].set(val)
            self.vars["spread"].set(normalize_percent_br(sp) if sp else "")
            self.vars["prazo_min"].set(pmin)
            self.vars["prazo_max"].set(pmax)
            self.vars["modalidade"].set(mod)
            tr, spe = trader_espec_from_regiao(reg)
            self.vars["trader_espec"].set(f"{tr} / {spe}".strip(" /") if (tr or spe) else "")
            self.hidden["premio"] = prem
            self.hidden["liquidacao"] = liq
            self._update_resumo()
        except Exception as e:
            messagebox.showerror("Erro ao ler PDF", str(e))

    def _update_resumo(self):
        spread = self.vars["spread"].get().strip()
        pmin   = self.vars["prazo_min"].get().strip()
        pmax   = self.vars["prazo_max"].get().strip()
        plat   = self.vars["plataforma"].get().strip()
        reg    = self.vars["regiao"].get().strip()
        mod    = normalize_modalidade(self.vars["modalidade"].get() or "")
        prem   = self.hidden.get("premio","sem prêmio")
        liq    = self.hidden.get("liquidacao","Débito em CC")
        te     = self.vars["trader_espec"].get().strip()
        if spread and "a.a" not in spread.lower(): spread = normalize_percent_br(spread)
        troca  = infer_troca(mod)
        prazo  = ""
        if pmin and pmax: prazo = f"Prazo {pmin} a {pmax} dias"
        elif pmin: prazo = f"Prazo mínimo {pmin} dias"
        elif pmax: prazo = f"Prazo máximo {pmax} dias"
        pr = ""
        if plat and reg: pr = f"Plataforma {plat} – Região {reg}"
        elif plat: pr = f"Plataforma {plat}"
        elif reg: pr = f"Região {reg}"
        L = [f"Spread mínimo {spread} – {prem} – Troca de Arquivo {troca}" if spread
             else f"Spread mínimo – {prem} – Troca de Arquivo {troca}"]
        if prazo: L.append(prazo)
        if liq:   L.append(f"Liquidação {liq}")
        if pr:    L.append(pr)
        if te:    L.append(f"Trader/Espec {te}")
        self.txt_resumo.delete("1.0","end")
        self.txt_resumo.insert("1.0", "\n".join(L).strip())

    def _clear_all(self):
        for v in self.vars.values(): v.set("")
        self.hidden["premio"] = "sem prêmio"
        self.hidden["liquidacao"] = "Débito em CC"
        self.txt_resumo.delete("1.0","end")

class OperacoesInvertidoFrame(tk.Frame):
    """Hub de Operações Invertido — escolhe entre análise de planilhas
    (.xlsx), a consulta de Limites Invertido / LTC e o histórico de
    operações (em breve)."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._xlsx_path = None
        self._overlay   = None
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=44, pady=(36, 0))
        eyebrow_label(hdr, "MESA DE OPERAÇÕES").pack(anchor="w")
        tk.Label(hdr, text="Operações Invertido", bg=C["bg"], fg=C["ink"],
                 font=("Segoe UI", 22, "bold")).pack(anchor="w", pady=(6, 0))
        tk.Label(hdr, text="Escolha uma ferramenta para continuar.",
                 bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", pady=(20, 0))

        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=44, pady=(32, 0))
        for c in range(3):
            body.columnconfigure(c, weight=1, uniform="opc")
        self._cards_body = body

        self._make_option_card(
            body, 0, 0, "▤", "Analisar Operações",
            "Importe uma planilha .xlsx para análise das operações.",
            self._open_analisar_overlay, "#5a9e72")

        self._make_option_card(
            body, 0, 1, "⬡", "Limites Invertido",
            "Consulta o LTC e o limite disponível de cada cliente.",
            lambda: self.controller.show_frame("LimitesInvertido"), C["accent"])

        self._make_option_card(
            body, 0, 2, "▦", "Histórico Operações",
            "Consulta o histórico de operações já realizadas.",
            self._open_historico_placeholder, "#8b72c9")

        self._taxas_card = None
        self._refresh_taxas_card()

    def _taxas_vencidas(self):
        try:
            return len(TaxasData.get().vencidas(LIMITE_INVERTIDO_CNPJS)) > 0
        except Exception:
            return False

    def on_show(self):
        self._refresh_taxas_card()

    def _refresh_taxas_card(self):
        try:
            if self._taxas_card is not None and self._taxas_card.winfo_exists():
                self._taxas_card.destroy()
        except Exception:
            pass
        self._taxas_card = self._make_option_card(
            self._cards_body, 1, 0, "⇄", "Taxas (Depara)",
            "Tabela de depara de taxas por cliente/modalidade.",
            lambda: self.controller.show_frame("TaxasInvertido"), "#c4a832",
            alert=self._taxas_vencidas())

    # ── Cards de opção ───────────────────────────────────────────────────────
    def _make_option_card(self, parent, row, col, icon, title, sub, command, color,
                           alert=False, alert_text="● Taxas vencidas"):
        return make_hub_option_card(parent, row, col, icon, title, sub, command, color,
                                    alert=alert, alert_text=alert_text)

    # ── Histórico Operações (placeholder) ───────────────────────────────────
    def _open_historico_placeholder(self):
        messagebox.showinfo(
            "Histórico Operações",
            "Histórico Operações estará disponível em breve.",
            parent=self.controller)

    # ── Overlay: Analisar Operações ─────────────────────────────────────────
    def _open_analisar_overlay(self):
        if self._overlay is not None:
            return
        overlay = tk.Frame(self, bg="#0c0c0c")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        overlay.bind("<Button-1>", lambda _e: self._close_analisar_overlay())
        self._overlay = overlay

        card = tk.Frame(overlay, bg=C["surface"],
                        highlightthickness=1, highlightbackground=C["hair"])
        card.place(relx=0.5, rely=0.5, anchor="center", width=460, height=320)
        card.bind("<Button-1>", lambda _e: "break")

        pad = tk.Frame(card, bg=C["surface"], padx=26, pady=22)
        pad.pack(fill="both", expand=True)

        top = tk.Frame(pad, bg=C["surface"])
        top.pack(fill="x")
        tk.Label(top, text="Analisar Operações", bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 13, "bold")).pack(side="left")
        styled_button(top, "✕", self._close_analisar_overlay,
                      small=True).pack(side="right")

        tk.Label(pad, text="Selecione uma planilha .xlsx com as operações para análise.",
                 bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9),
                 wraplength=400, justify="left").pack(anchor="w", pady=(10, 0))

        make_hairline(pad, bg=C["hair"]).pack(fill="x", pady=(16, 0))

        drop = tk.Frame(pad, bg=C["surface2"], highlightthickness=1,
                        highlightbackground=C["hair"])
        drop.pack(fill="x", pady=(16, 0))
        inner_drop = tk.Frame(drop, bg=C["surface2"], pady=24)
        inner_drop.pack(fill="x")

        has_file = bool(self._xlsx_path)
        icon_lbl = tk.Label(inner_drop, text=("✓" if has_file else "▤"),
                            bg=C["surface2"],
                            fg=(C["ok"] if has_file else C["ink_faint"]),
                            font=("Segoe UI", 20))
        icon_lbl.pack()
        file_lbl = tk.Label(inner_drop,
                            text=(os.path.basename(self._xlsx_path) if has_file
                                  else "Nenhum arquivo selecionado"),
                            bg=C["surface2"],
                            fg=(C["ink"] if has_file else C["ink_muted"]),
                            font=("Segoe UI", 9))
        file_lbl.pack(pady=(8, 0))

        self._overlay_file_lbl = file_lbl
        self._overlay_icon_lbl = icon_lbl

        foot = tk.Frame(pad, bg=C["surface"])
        foot.pack(fill="x", pady=(20, 0))
        self._overlay_action_btn = styled_button(
            foot, "Selecionar arquivo .xlsx…",
            self._overlay_action_click, accent=True)
        self._overlay_action_btn.pack(side="left")
        self._overlay_remove_btn = styled_button(
            foot, "Remover arquivo",
            self._remove_xlsx, danger=True, small=True)
        if has_file:
            self._overlay_remove_btn.pack(side="left", padx=(8, 0))
        self._overlay_ready = bool(self._xlsx_path)
        if self._overlay_ready:
            self._set_overlay_analyze_state(animate=False)
        styled_button(foot, "Cancelar",
                      self._close_analisar_overlay).pack(side="right")

    def _overlay_action_click(self):
        if self._overlay_ready:
            self._start_analyze()
        else:
            self._pick_xlsx()

    def _remove_xlsx(self):
        self._xlsx_path = None
        self._overlay_ready = False
        if getattr(self.controller, "invertido_xlsx_path", None):
            self.controller.invertido_xlsx_path = None
        if hasattr(self, "_overlay_file_lbl") and self._overlay_file_lbl.winfo_exists():
            self._overlay_file_lbl.configure(text="Nenhum arquivo selecionado", fg=C["ink_muted"])
            self._overlay_icon_lbl.configure(text="▤", fg=C["ink_faint"])
        if hasattr(self, "_overlay_remove_btn") and self._overlay_remove_btn.winfo_exists():
            self._overlay_remove_btn.pack_forget()
        btn = getattr(self, "_overlay_action_btn", None)
        if btn is not None and btn.winfo_exists():
            btn.configure(text="Selecionar arquivo .xlsx…",
                          bg=C["accent_dim"], fg=C["accent"],
                          command=self._overlay_action_click)
            btn.bind("<Enter>", lambda _: btn.configure(bg=C["accent"], fg=C["bg"]))
            btn.bind("<Leave>", lambda _: btn.configure(bg=C["accent_dim"], fg=C["accent"]))

    def _set_overlay_analyze_state(self, animate=True):
        btn = getattr(self, "_overlay_action_btn", None)
        if btn is None or not btn.winfo_exists():
            return
        self._overlay_ready = True
        btn.configure(command=self._start_analyze)
        if not animate:
            btn.configure(text="Analisar", bg=C["ok_dim"], fg=C["ok"],
                          activebackground=C["ok"], activeforeground=C["bg"])
            btn.bind("<Enter>", lambda _: btn.configure(bg=C["ok"], fg=C["bg"]))
            btn.bind("<Leave>", lambda _: btn.configure(bg=C["ok_dim"], fg=C["ok"]))
            return

        steps = [
            (C["accent_dim"], C["accent"], "Selecionar arquivo .xlsx…"),
            (C["surface3"],   C["ink"],    "Preparando…"),
            ("#2d4a38",       "#6fbf96",   "Analisar"),
            (C["ok_dim"],     C["ok"],     "Analisar"),
        ]
        def _step(i=0):
            if btn is None or not btn.winfo_exists():
                return
            bg, fg, text = steps[min(i, len(steps) - 1)]
            btn.configure(text=text, bg=bg, fg=fg,
                          activebackground=C["ok"], activeforeground=C["bg"])
            if i < len(steps) - 1:
                self.after(55, lambda: _step(i + 1))
            else:
                btn.bind("<Enter>", lambda _: btn.configure(bg=C["ok"], fg=C["bg"]))
                btn.bind("<Leave>", lambda _: btn.configure(bg=C["ok_dim"], fg=C["ok"]))
        _step()

    def _start_analyze(self):
        if not self._xlsx_path:
            return
        self.controller.invertido_xlsx_path = self._xlsx_path
        self._close_analisar_overlay()
        self.controller.show_frame("AnalisarOperacoes")

    def _pick_xlsx(self):
        p = filedialog.askopenfilename(
            title="Selecionar planilha de operações",
            filetypes=[("Planilha Excel", "*.xlsx")])
        if not p:
            return
        self._xlsx_path = p
        if hasattr(self, "_overlay_file_lbl") and self._overlay_file_lbl.winfo_exists():
            self._overlay_file_lbl.configure(text=os.path.basename(p), fg=C["ink"])
            self._overlay_icon_lbl.configure(text="✓", fg=C["ok"])
        if hasattr(self, "_overlay_remove_btn") and self._overlay_remove_btn.winfo_exists():
            self._overlay_remove_btn.pack(side="left", padx=(8, 0))
        self._set_overlay_analyze_state(animate=True)

    def _close_analisar_overlay(self):
        if self._overlay is not None:
            try: self._overlay.destroy()
            except Exception: pass
            self._overlay = None
        self._overlay_ready = bool(self._xlsx_path)


class TaxasInvertidoFrame(tk.Frame):
    """Tabela de depara de taxas (%) por cliente de Operações Invertido.
    Taxas têm validade mensal: ao virar o mês, ficam vencidas até serem
    atualizadas novamente."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._data = TaxasData.get()
        self._cards = {}
        self._build()
        self._data.on_reconnect(self._on_data_reconnect)

    def _on_data_reconnect(self):
        # Chamado pela thread de retry; agenda a atualização na thread da UI.
        try:
            self.after(0, self._refresh_network_state)
        except Exception:
            pass

    def _refresh_network_state(self):
        if not self.winfo_exists():
            return
        disponivel = self._data.is_available()
        if disponivel:
            self._net_banner.pack_forget()
        else:
            self._net_banner.pack(fill="x", padx=0, pady=(0, 0), after=self._hairline_top)
        if hasattr(self, "_grid"):
            self._setup_cards()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24, 0))
        styled_button(hdr, "← Voltar",
                      lambda: self.controller.show_frame("OperacoesInvertido")).pack(side="left")
        tk.Label(hdr, text="Taxas (Depara)", bg=C["bg"], fg=C["ink"],
                 font=("Georgia", 18, "bold")).pack(side="left", padx=(14, 0))

        sub = tk.Frame(self, bg=C["bg"])
        sub.pack(fill="x", padx=32)
        tk.Label(sub, text="Taxa (%) vigente por cliente. Validade mensal — "
                            "renove ao virar o mês.",
                 bg=C["bg"], fg=C["ink_muted"], font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))
        self._validade_lbl = tk.Label(sub, text="", bg=C["bg"], fg=C["ink_faint"],
                                      font=("Segoe UI", 8))
        self._validade_lbl.pack(anchor="w", pady=(2, 0))

        self._hairline_top = make_hairline(self, bg=C["hair"])
        self._hairline_top.pack(fill="x", padx=0, pady=(16, 0))

        self._net_banner = tk.Frame(self, bg=C["err_dim"])
        tk.Label(
            self._net_banner,
            text="⚠ Sem conexão com a rede — não é possível ler ou salvar as taxas "
                 "agora. Tentando reconectar automaticamente…",
            bg=C["err_dim"], fg=C["err"], font=("Segoe UI", 8, "bold"),
            anchor="w", justify="left", wraplength=900,
        ).pack(side="left", fill="x", expand=True, padx=18, pady=8)
        # _net_banner exibido condicionalmente em _refresh_network_state()

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        self._grid_outer = self._sf.inner
        self._grid_outer.configure(bg=C["bg"])

        self._grid = tk.Frame(self._grid_outer, bg=C["bg"])
        self._grid.pack(padx=32, pady=(16, 24), fill="x")
        for c in range(3):
            self._grid.columnconfigure(c, weight=1, uniform="tcards")

    def on_show(self):
        self._sf.refresh_bindings()
        self._validade_lbl.configure(
            text=f"Mês de referência: {self._mes_atual_label()}")
        self._refresh_network_state()

    @staticmethod
    def _mes_atual_label():
        MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
                  "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
        today = date.today()
        return f"{MESES[today.month - 1]}/{today.year}"

    def _setup_cards(self):
        for w in self._grid.winfo_children():
            w.destroy()
        self._cards = {}
        nomes = list(BPM_CLIENT_DATA.keys())
        for idx, nome in enumerate(nomes):
            row, col = divmod(idx, 3)
            self._make_card(nome, row, col)

    def _make_card(self, nome, row, col):
        cnpj = only_digits(BPM_CLIENT_DATA[nome].get("CNPJ", ""))
        disponivel = self._data.is_available()
        info = self._data.get_taxa(cnpj) if disponivel else None
        vigente = self._data.is_vigente(cnpj) if disponivel else False

        bg = C["surface"]
        outer = tk.Frame(self._grid, bg=C["bg"])
        outer.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)

        card = tk.Frame(outer, bg=bg, highlightthickness=1, highlightbackground=C["hair"], bd=0)
        card.pack(fill="both", expand=True)
        top_color = C["ok"] if vigente else (C["err"] if info else C["hair"])
        top_bar = tk.Frame(card, bg=top_color, height=2)
        top_bar.pack(fill="x")

        body = tk.Frame(card, bg=bg, padx=16, pady=14)
        body.pack(fill="both", expand=True)

        tk.Label(body, text=nome, bg=bg, fg=C["ink"],
                 font=("Segoe UI", 10, "bold"), wraplength=190,
                 justify="left", anchor="w").pack(fill="x")
        tk.Label(body, text=f"CNPJ {cnpj}" if cnpj else "CNPJ não cadastrado",
                 bg=bg, fg=C["ink_faint"], font=("Segoe UI", 7), anchor="w").pack(
                     fill="x", pady=(2, 10))

        taxa_row = tk.Frame(body, bg=bg)
        taxa_row.pack(fill="x")
        tk.Label(taxa_row, text="TAXA", bg=bg, fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold"), width=6, anchor="w").pack(side="left")
        taxa_txt = f"{info['taxa']}%" if info else "—"
        taxa_lbl = tk.Label(taxa_row, text=taxa_txt, bg=bg,
                            fg=C["ink"] if info else C["ink_faint"],
                            font=("Segoe UI", 11, "bold"), anchor="w")
        taxa_lbl.pack(side="left", fill="x", expand=True)

        status_row = tk.Frame(body, bg=bg)
        status_row.pack(fill="x", pady=(6, 0))
        if not disponivel:
            status_text, status_color = "Indisponível (sem rede)", C["ink_faint"]
        elif vigente:
            status_text, status_color = "Vigente", C["ok"]
        elif info:
            status_text, status_color = "Vencida — renovar", C["err"]
        else:
            status_text, status_color = "Sem taxa cadastrada", C["ink_faint"]
        status_lbl = tk.Label(status_row, text=status_text, bg=bg, fg=status_color,
                              font=("Segoe UI", 7, "bold"))
        status_lbl.pack(side="left")

        btn_row = tk.Frame(body, bg=bg)
        btn_row.pack(fill="x", pady=(12, 0))
        edit_btn = styled_button(
            btn_row, "Editar" if info else "Definir taxa",
            lambda n=nome, c=cnpj: self._open_edit_dialog(n, c),
            accent=not vigente, small=True)
        edit_btn.pack(side="left")
        if not disponivel:
            edit_btn.configure(state="disabled")

        self._cards[cnpj] = {
            "outer": outer, "taxa_lbl": taxa_lbl, "status_lbl": status_lbl,
            "top_bar": top_bar,
        }

    def _open_edit_dialog(self, nome, cnpj):
        if not self._data.is_available():
            return
        info = self._data.get_taxa(cnpj)
        dlg = tk.Toplevel(self)
        dlg.title(f"Taxa — {nome}")
        dlg.configure(bg=C["surface"])
        dlg.geometry("380x260")
        dlg.resizable(False, False)
        dlg.grab_set()

        pad = tk.Frame(dlg, bg=C["surface"], padx=24, pady=20)
        pad.pack(fill="both", expand=True)

        tk.Label(pad, text=nome, bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(pad, text=f"CNPJ {cnpj}", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))

        make_hairline(pad, bg=C["hair"]).pack(fill="x", pady=(14, 14))

        tk.Label(pad, text="TAXA (%)", bg=C["surface"], fg=C["ink_faint"],
                 font=("Segoe UI", 7, "bold")).pack(anchor="w")
        taxa_var = tk.StringVar(value=(info["taxa"] if info else ""))
        entry = styled_entry(pad, textvariable=taxa_var, width=14)
        entry.configure(font=("Segoe UI", 14, "bold"))
        entry.pack(anchor="w", pady=(6, 0))
        entry.focus_set()

        hint_lbl = tk.Label(pad, text="Formato: 1,3950  •  2,0000  •  1,4300",
                            bg=C["surface"], fg=C["ink_faint"], font=("Segoe UI", 7))
        hint_lbl.pack(anchor="w", pady=(6, 0))

        err_lbl = tk.Label(pad, text="", bg=C["surface"], fg=C["err"],
                           font=("Segoe UI", 8))
        err_lbl.pack(anchor="w", pady=(4, 0))

        def _valido(txt):
            txt = (txt or "").strip()
            return bool(re.fullmatch(r"\d{1,3},\d{1,4}", txt))

        def _normalizar(txt):
            inteiro, dec = txt.strip().split(",")
            dec = dec.ljust(4, "0")
            return f"{inteiro},{dec}"

        def _salvar():
            txt = taxa_var.get().strip()
            if not _valido(txt):
                err_lbl.configure(
                    text="Formato inválido. Use vírgula, ex.: 1,3950")
                return
            txt = _normalizar(txt)
            ok = self._data.set_taxa(cnpj, txt)
            if not ok:
                err_lbl.configure(
                    text="Sem conexão com a rede. Não foi possível salvar agora.")
                return
            dlg.destroy()
            self._refresh_network_state()

        btn_row = tk.Frame(pad, bg=C["surface"])
        btn_row.pack(fill="x", pady=(18, 0))
        styled_button(btn_row, "Cancelar", dlg.destroy).pack(side="left")
        styled_button(btn_row, "Salvar", _salvar, accent=True).pack(side="right")

        entry.bind("<Return>", lambda _e: _salvar())


class AnalisarOperacoesFrame(tk.Frame, ThreadSafeUIMixin):
    """Exibe as operações importadas de uma planilha .xlsx, agrupadas por sacado."""

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._groups = []
        self._cards = []
        self._limit_btns = {}
        self._worker_running = False
        self._last_path = None
        self._all_ops = []
        self._excluded_uids = set()
        self._detail_overlay = None
        self._detail_key = None
        self._limite_overlay = None
        self._alert_overlay = None
        self._alert_decisions = {}
        self._alert_items = []
        self._pending_ops = []
        self._init_ui_queue()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24, 0))
        styled_button(hdr, "← Voltar",
                      lambda: (self._close_detalhes(),
                               self.controller.show_frame("OperacoesInvertido"))).pack(side="left")
        tk.Label(hdr, text="Analisar Operações", bg=C["bg"], fg=C["ink"],
                 font=("Georgia", 18, "bold")).pack(side="left", padx=(14, 0))

        sub = tk.Frame(self, bg=C["bg"])
        sub.pack(fill="x", padx=32)
        self._sub_lbl = tk.Label(
            sub, text="Importe e analise as operações da planilha selecionada.",
            bg=C["bg"], fg=C["ink_muted"], font=("Segoe UI", 9))
        self._sub_lbl.pack(anchor="w", pady=(4, 0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(16, 0))

        self._taxas_banner = tk.Frame(self, bg=C["err_dim"])
        self._taxas_banner_lbl = tk.Label(
            self._taxas_banner, text="", bg=C["err_dim"], fg=C["err"],
            font=("Segoe UI", 9, "bold"), anchor="w", justify="left", wraplength=900)
        self._taxas_banner_lbl.pack(side="left", fill="x", expand=True,
                                    padx=18, pady=10)
        styled_button(
            self._taxas_banner, "Atualizar taxas →",
            lambda: self.controller.show_frame("TaxasInvertido"),
            danger=True, small=True).pack(side="right", padx=18)
        # _taxas_banner é exibido condicionalmente via _refresh_taxas_banner()

        self._loading_outer = tk.Frame(self, bg=C["bg"])
        self._loading_outer.pack(fill="x", padx=32, pady=(28, 0))
        loading_card = tk.Frame(self._loading_outer, bg=C["surface"],
                                highlightthickness=1, highlightbackground=C["hair"])
        loading_card.pack(fill="x")
        loading_body = tk.Frame(loading_card, bg=C["surface"], padx=18, pady=16)
        loading_body.pack(fill="x")
        self._loading_icon = tk.Label(loading_body, text="◐", bg=C["surface"],
                                      fg=C["ok"], font=("Segoe UI", 16, "bold"))
        self._loading_icon.pack()
        self._loading_lbl = tk.Label(
            loading_body, text="Analisando planilha…", bg=C["surface"],
            fg=C["ink_muted"], font=("Segoe UI", 9))
        self._loading_lbl.pack(pady=(8, 0))
        self._loading_spin = [None]
        self._loading_angle = [0]

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        self._grid_outer = self._sf.inner
        self._grid_outer.configure(bg=C["bg"])

        self._grid = tk.Frame(self._grid_outer, bg=C["bg"])
        self._grid.pack(padx=32, pady=(16, 24), fill="x")
        for c in range(3):
            self._grid.columnconfigure(c, weight=1, uniform="acards")

    def on_show(self):
        self._sf.refresh_bindings()
        self._refresh_taxas_banner()
        path = getattr(self.controller, "invertido_xlsx_path", None)
        if not path:
            self.controller.show_frame("OperacoesInvertido")
            return
        fname = os.path.basename(path)
        self._sub_lbl.configure(text=f"Planilha: {fname}")
        if path == self._last_path and self._groups and not self._worker_running:
            self._refresh_all_limit_buttons()
            return
        self._last_path = path
        self._reset_view()
        self._start_worker(path)
        self.controller.register_limites_listener(self._on_limites_update)

    def _refresh_taxas_banner(self):
        try:
            vencidas = TaxasData.get().vencidas(LIMITE_INVERTIDO_CNPJS)
        except Exception:
            vencidas = []
        if vencidas:
            nomes_vencidos = sorted({
                nome for nome, info in BPM_CLIENT_DATA.items()
                if only_digits(info.get("CNPJ", "")) in vencidas
            })
            qtd = len(nomes_vencidos)
            texto = (f"⚠ {qtd} taxa(s) vencida(s) este mês "
                     f"({self._mes_atual_label()}): " + ", ".join(nomes_vencidos[:6]) +
                     ("…" if qtd > 6 else "") +
                     ". Atualize em Taxas (Depara) antes de seguir com as operações.")
            self._taxas_banner_lbl.configure(text=texto)
            self._taxas_banner.pack(fill="x", padx=0, pady=(0, 0), before=self._loading_outer)
        else:
            self._taxas_banner.pack_forget()

    @staticmethod
    def _mes_atual_label():
        MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
                  "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
        today = date.today()
        return f"{MESES[today.month - 1]}/{today.year}"

    def _on_limites_update(self, _cnpj=None):
        self._ui(self._refresh_all_limit_buttons)

    def _reset_view(self):
        self._groups = []
        self._cards = []
        for w in self._grid.winfo_children():
            w.destroy()
        self._loading_lbl.configure(text="Analisando planilha…", fg=C["ink_muted"])
        self._loading_icon.configure(text="◐", fg=C["ok"])
        self._loading_outer.pack(fill="x", padx=32, pady=(28, 0))
        self._start_loading_anim()

    def _start_loading_anim(self):
        if self._loading_spin[0]:
            try:
                self.after_cancel(self._loading_spin[0])
            except Exception:
                pass
        self._loading_angle[0] = 0

        def tick():
            self._loading_angle[0] = (self._loading_angle[0] + 1) % 4
            if self._loading_icon.winfo_exists():
                self._loading_icon.configure(
                    text=["◐", "◓", "◑", "◒"][self._loading_angle[0]])
            self._loading_spin[0] = self.after(120, tick)

        tick()

    def _stop_loading_anim(self):
        if self._loading_spin[0]:
            try:
                self.after_cancel(self._loading_spin[0])
            except Exception:
                pass
            self._loading_spin[0] = None

    def _make_group_card(self, group, row, col):
        bg = C["surface"]
        outer = tk.Frame(self._grid, bg=C["bg"])
        outer.grid(row=row, column=col, sticky="new", padx=5, pady=5)

        card = tk.Frame(outer, bg=bg, highlightthickness=1,
                        highlightbackground=C["hair"], bd=0)
        card.pack(fill="x")
        top_bar = tk.Frame(card, bg=C["ok"], height=2)
        top_bar.pack(fill="x")

        body = tk.Frame(card, bg=bg, padx=14, pady=12)
        body.pack(fill="x")

        tk.Label(body, text=group["nome_sacado"], bg=bg, fg=C["ink"],
                 font=("Segoe UI", 9, "bold"), wraplength=170,
                 justify="center").pack()

        info = tk.Frame(body, bg=bg)
        info.pack(pady=(10, 0), fill="x")
        liquido, completo = self._calc_valor_liquido_group(group)
        linhas = [
            ("Notas", str(group["count"]), False),
            ("Montante", group["valor_total"], True),
        ]
        if liquido is not None:
            liq_txt = _fmt_brl(liquido) + ("" if completo else " *")
            linhas.append(("Líquido", liq_txt, True))
        for label, value, accent in linhas:
            row_f = tk.Frame(info, bg=bg)
            row_f.pack(fill="x", pady=(0, 3))
            tk.Label(row_f, text=label, bg=bg, fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold"), width=9, anchor="w").pack(side="left")
            fg = C["ok"] if accent else C["ink_muted"]
            tk.Label(row_f, text=value, bg=bg, fg=fg,
                     font=("Segoe UI", 8, "bold" if accent else "normal"),
                     anchor="w").pack(side="left", fill="x", expand=True)

        btn_row = tk.Frame(body, bg=bg)
        btn_row.pack(fill="x", pady=(12, 0))
        status, label = self._evaluate_group_limite(group)
        lim_btn = styled_button_limite(
            btn_row, label,
            lambda g=group: self._open_limite_modal(g),
            variant=self._limite_variant(status),
            small=True)
        lim_btn.pack(side="left")
        styled_button(btn_row, "Detalhes →",
                      lambda k=group["nome_sacado"]: self._open_detalhes(k),
                      accent=True, small=True).pack(side="right")

        cnpj = group.get("doc_sacado") or ""
        self._limit_btns[self._group_limite_key(group)] = lim_btn

        return {"outer": outer, "lim_btn": lim_btn, "group": group}

    def _limite_variant(self, status):
        return {"ok": "ok", "quase": "warn", "nao_validado": "warn",
                "nao_encontrado": "warn", "validando": "idle",
                "insuficiente": "err", "sem_taxa": "err"}.get(status, "warn")

    def _group_limite_key(self, group):
        return group.get("doc_sacado") or _normalize_sacado_key(group["nome_sacado"])

    def _calc_valor_liquido_group(self, group):
        """Recalcula o valor líquido do grupo nota a nota (cada nota pode
        ter vencimento/prazo diferente), usando a taxa vigente do cliente
        no Depara. Retorna (valor_liquido: Decimal|None, completo: bool).
        completo=False indica que ao menos uma nota não pôde ser calculada
        (sem taxa vigente ou sem data de vencimento válida)."""
        cnpj = group.get("doc_sacado") or ""
        taxa_info = TaxasData.get().get_taxa(cnpj) if cnpj else None
        vigente = TaxasData.get().is_vigente(cnpj) if cnpj else False
        if not taxa_info or not vigente:
            return None, False
        taxa_str = taxa_info.get("taxa")
        hoje = date.today()
        total_liq = Decimal("0")
        completo = True
        for op in group.get("notas", []):
            venc = _parse_data_curta(op.get("data_vencimento"))
            if venc is None:
                completo = False
                continue
            prazo = (venc - hoje).days
            vl = calcular_valor_liquido(op.get("valor_raw", Decimal("0")), taxa_str, prazo)
            if vl is None:
                completo = False
                continue
            total_liq += vl
        if not group.get("notas"):
            completo = False
        return total_liq, completo

    def _get_limite_data(self, group):
        cnpj = group.get("doc_sacado") or ""
        if not cnpj:
            return None
        return getattr(self.controller, "invertido_limites_cache", {}).get(cnpj)

    def _evaluate_group_limite(self, group):
        cnpj = group.get("doc_sacado") or ""
        if not cnpj or cnpj not in LIMITE_INVERTIDO_CNPJS:
            tem_taxa = bool(cnpj) and TaxasData.get().get_taxa(cnpj) is not None
            if not tem_taxa:
                return "sem_taxa", "Sem taxa Parametrizada"
            return "nao_encontrado", "Limite não encontrado"
        limite_data = self._get_limite_data(group)
        lf = self.controller.frames.get("LimitesInvertido")
        if lf and lf._worker_running and not limite_data:
            return "validando", "Validando…"
        if not limite_data:
            return "nao_validado", "Limites não validados"
        return _evaluate_limite_operacao(group["total"], limite_data)

    def _refresh_all_limit_buttons(self):
        for key, btn in list(self._limit_btns.items()):
            if not btn.winfo_exists():
                continue
            group = next((g for g in self._groups if self._group_limite_key(g) == key), None)
            if not group:
                continue
            status, label = self._evaluate_group_limite(group)
            _set_limite_button(btn, label, self._limite_variant(status))

    def _close_limite_modal(self):
        if self._limite_overlay is not None:
            try:
                self._limite_overlay.destroy()
            except Exception:
                pass
            self._limite_overlay = None

    def _start_limites_validation(self):
        lf = self.controller.frames.get("LimitesInvertido")
        if lf is None or lf._worker_running:
            return
        self._close_limite_modal()
        if lf._started and not lf._worker_running:
            lf._restart_limites()
        else:
            lf._started = True
            lf._cancel_requested = False
            lf._setup_cards()
            lf._start_worker()
        self._refresh_all_limit_buttons()

    def _open_limite_modal(self, group):
        if self._limite_overlay is not None:
            return
        status, label = self._evaluate_group_limite(group)
        limite_data = self._get_limite_data(group)

        overlay = tk.Frame(self, bg="#0c0c0c")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._limite_overlay = overlay

        card = tk.Frame(overlay, bg=C["surface"],
                        highlightthickness=1, highlightbackground=C["hair"])
        card.place(relx=0.5, rely=0.5, anchor="center", width=420, height=300)
        card.bind("<Button-1>", lambda _e: "break")

        pad = tk.Frame(card, bg=C["surface"], padx=24, pady=20)
        pad.pack(fill="both", expand=True)

        top = tk.Frame(pad, bg=C["surface"])
        top.pack(fill="x")
        title_fg = {"ok": C["ok"], "quase": C["warn"], "insuficiente": C["err"],
                    "nao_encontrado": C["warn"], "sem_taxa": C["err"]}.get(status, C["warn"])
        tk.Label(top, text=label, bg=C["surface"], fg=title_fg,
                 font=("Segoe UI", 12, "bold")).pack(side="left")
        styled_button(top, "✕", self._close_limite_modal, small=True).pack(side="right")

        tk.Label(pad, text=group["nome_sacado"], bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 9), wraplength=360, justify="left").pack(
            anchor="w", pady=(10, 0))

        make_hairline(pad, bg=C["hair"]).pack(fill="x", pady=(14, 0))

        body = tk.Frame(pad, bg=C["surface"])
        body.pack(fill="both", expand=True, pady=(12, 0))

        if status == "sem_taxa":
            tk.Label(
                body,
                text="Sem taxa parametrizada ou convênio ativo para operações invertidos.",
                bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9),
                wraplength=360, justify="left",
            ).pack(anchor="w")
            foot = tk.Frame(pad, bg=C["surface"])
            foot.pack(fill="x", pady=(16, 0))
            styled_button(foot, "Fechar", self._close_limite_modal, accent=True, small=True).pack(
                side="right")
            return

        if status == "nao_validado":
            tk.Label(
                body,
                text=("Os limites ainda não foram consultados para este sacado.\n"
                      "Deseja iniciar a validação agora?"),
                bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9),
                wraplength=360, justify="left",
            ).pack(anchor="w")
            foot = tk.Frame(pad, bg=C["surface"])
            foot.pack(fill="x", pady=(16, 0))
            styled_button(foot, "Cancelar", self._close_limite_modal, small=True).pack(side="right")
            styled_button(foot, "Iniciar validação",
                          self._start_limites_validation, accent=True, small=True).pack(
                side="right", padx=(0, 8))
            return

        if status == "nao_encontrado":
            doc = group.get("doc_sacado") or "—"
            tk.Label(
                body,
                text=(f"CNPJ {doc} não consta na lista de Limites Invertido.\n"
                      "Este sacado não possui consulta de limite cadastrada no app."),
                bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9),
                wraplength=360, justify="left",
            ).pack(anchor="w")
            foot = tk.Frame(pad, bg=C["surface"])
            foot.pack(fill="x", pady=(16, 0))
            styled_button(foot, "Fechar", self._close_limite_modal, accent=True, small=True).pack(
                side="right")
            return

        mont = group["valor_total"]
        lines = [f"Montante do grupo: {mont}"]
        if limite_data:
            if limite_data.get("ltc_str"):
                lines.append(f"LTC ativo · vence {limite_data['ltc_str']}")
            lines.append(f"Limite disp. (fornecedor): {_fmt_limite_int(limite_data.get('limite_disp'))}")
            limite = limite_data.get("limite_disp")
            if limite is not None:
                slack = limite - int(group["total"].quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                lines.append(f"Sobra: {_fmt_limite_int(slack)}")
            if limite_data.get("via"):
                lines.append(f"Consulta via {limite_data['via']}")
            if status == "quase":
                lines.append(f"\nSobra abaixo de R$ {LIMITE_SOBRA_MIN:,}".replace(",", "."))
            elif status == "insuficiente":
                if limite_data.get("state") == "ltc_expired":
                    lines.append("\nLTC vencido ou indisponível.")
                else:
                    lines.append("\nMontante superior ao limite disponível.")

        tk.Label(body, text="\n".join(lines), bg=C["surface"], fg=C["ink_muted"],
                 font=("Segoe UI", 9), wraplength=360, justify="left").pack(anchor="w")

        foot = tk.Frame(pad, bg=C["surface"])
        foot.pack(fill="x", pady=(16, 0))
        styled_button(foot, "Fechar", self._close_limite_modal, accent=True, small=True).pack(
            side="right")

    def _find_group(self, nome_sacado):
        key = _normalize_sacado_key(nome_sacado)
        for g in self._groups:
            if _normalize_sacado_key(g["nome_sacado"]) == key:
                return g
        return None

    def _bind_modal_scroll(self, widgets, sf):
        def _mw(event):
            sf._scroll_mousewheel(event)
            return "break"
        for widget in widgets:
            widget.bind("<MouseWheel>", _mw)
            widget.bind("<Button-4>", _mw)
            widget.bind("<Button-5>", _mw)

    def _open_detalhes(self, nome_sacado):
        group = self._find_group(nome_sacado)
        if group is None or self._detail_overlay is not None:
            return

        overlay = tk.Frame(self, bg="#0c0c0c")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._detail_overlay = overlay
        self._detail_key = self._group_limite_key(group)
        self._detail_nome = group["nome_sacado"]

        card = tk.Frame(overlay, bg=C["surface"],
                        highlightthickness=1, highlightbackground=C["hair"])
        card.place(relx=0.5, rely=0.5, anchor="center", width=720, height=620)
        card.bind("<Button-1>", lambda _e: "break")

        pad = tk.Frame(card, bg=C["surface"], padx=28, pady=22)
        pad.pack(fill="both", expand=True)

        top = tk.Frame(pad, bg=C["surface"])
        top.pack(fill="x")
        tk.Label(top, text=group["nome_sacado"], bg=C["surface"], fg=C["ink"],
                 font=("Segoe UI", 13, "bold"), wraplength=380,
                 justify="left").pack(side="left", fill="x", expand=True)
        styled_button(top, "✕", self._close_detalhes, small=True).pack(side="right")
        styled_button(top, "Enviar e-mail →", self._enviar_email_risco_sacado,
                      accent=True, small=True).pack(side="right", padx=(0, 8))

        self._detail_sub_lbl = tk.Label(
            pad, text="", bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9))
        self._detail_sub_lbl.pack(anchor="w", pady=(8, 0))

        tk.Label(
            pad,
            text=("Inclua notas excluídas na triagem ou exclua notas que não devem "
                  "compor o montante. Apenas notas já existentes nesta planilha."),
            bg=C["surface"], fg=C["ink_faint"], font=("Segoe UI", 8),
            wraplength=660, justify="left",
        ).pack(anchor="w", pady=(2, 0))

        make_hairline(pad, bg=C["hair"]).pack(fill="x", pady=(14, 0))

        scroll_wrap = tk.Frame(pad, bg=C["surface"], height=420)
        scroll_wrap.pack(fill="both", expand=True, pady=(12, 0))
        scroll_wrap.pack_propagate(False)
        self._detail_scroll_wrap = scroll_wrap

        sf = ScrollableFrame(scroll_wrap, bg=C["surface"])
        sf.pack(fill="both", expand=True)
        sf.link_wheel(scroll_wrap)
        sf.link_wheel(pad)
        sf.link_wheel(card)
        list_outer = sf.inner
        list_outer.configure(bg=C["surface"])
        self._detail_sf = sf
        self._detail_list_outer = list_outer
        self._detail_card = card
        self._detail_pad = pad

        self._render_detalhes_list()

    def _detail_ops_do_cliente(self):
        """Todas as ops da planilha (incluídas ou não) para o cliente em detalhe."""
        key = self._detail_key

        def _op_key(op):
            doc = only_digits(op.get("doc_sacado") or "")
            return doc or _normalize_sacado_key(op.get("nome_sacado") or "")

        return [op for op in self._all_ops if _op_key(op) == key]

    def _detail_doc_sacado(self):
        for op in self._detail_ops_do_cliente():
            doc = only_digits(op.get("doc_sacado") or "")
            if doc:
                return doc
        return ""

    def _detail_doc_cedente(self):
        for op in self._detail_ops_do_cliente():
            doc = only_digits(op.get("doc_cedente") or "")
            if doc:
                return doc
        return ""

    def _enviar_email_risco_sacado(self):
        doc_sacado = self._detail_doc_sacado()
        doc_cedente = self._detail_doc_cedente()
        todas = self._detail_ops_do_cliente()
        incluidas = [op for op in todas if op["uid"] not in self._excluded_uids]
        if not incluidas:
            messagebox.showwarning(
                "Sem notas", "Não há notas incluídas no montante para este sacado.")
            return

        taxa_info = TaxasData.get().get_taxa(doc_sacado) if doc_sacado else None
        vigente = TaxasData.get().is_vigente(doc_sacado) if doc_sacado else False
        if not taxa_info or not vigente:
            messagebox.showwarning(
                "Taxa não vigente",
                "Não há taxa vigente cadastrada em Taxas (Depara) para este cliente. "
                "Atualize a taxa antes de enviar o e-mail.")
            return
        taxa_str = taxa_info.get("taxa")

        hoje = date.today()
        notas_calc = []
        for op in incluidas:
            venc = _parse_data_curta(op.get("data_vencimento"))
            prazo = (venc - hoje).days if venc else 0
            vl = calcular_valor_liquido(op.get("valor_raw", Decimal("0")), taxa_str, prazo)
            notas_calc.append({
                "nf": op.get("nf"),
                "data_vencimento": op.get("data_vencimento"),
                "valor_raw": op.get("valor_raw", Decimal("0")),
                "valor_liquido": vl,
            })

        nome_sacado = self._detail_nome
        subject = f"RISCO SACADO INVERTIDO - {nome_sacado.upper()}"
        html = build_risco_sacado_email_html(
            sacado_nome=nome_sacado.upper(),
            sacado_cnpj=doc_sacado,
            cedente_cnpj=doc_cedente,
            notas=notas_calc,
            taxa_str=taxa_str)

        try:
            enviar_email_outlook_risco_sacado(subject, html)
        except Exception as e:
            messagebox.showerror("Erro ao abrir e-mail", str(e))

    def _render_detalhes_list(self):
        list_outer = self._detail_list_outer
        sf = self._detail_sf
        for w in list_outer.winfo_children():
            w.destroy()

        scroll_targets = [self._detail_overlay, self._detail_card, self._detail_pad,
                          self._detail_scroll_wrap, sf, list_outer]

        todas = self._detail_ops_do_cliente()
        incluidas = [op for op in todas if op["uid"] not in self._excluded_uids]
        excluidas = [op for op in todas if op["uid"] in self._excluded_uids]
        total = sum((op.get("valor_raw", Decimal("0")) for op in incluidas), Decimal("0"))
        liquido, completo = self._calc_valor_liquido_group(
            {"doc_sacado": self._detail_doc_sacado(), "notas": incluidas})
        resumo = f"{len(incluidas)} nota(s) no montante · {_fmt_brl(total)}"
        if liquido is not None:
            resumo += f" · Líquido {_fmt_brl(liquido)}" + ("" if completo else " *")
        if excluidas:
            resumo += f" · {len(excluidas)} excluída(s)"
        self._detail_sub_lbl.configure(text=resumo)

        def _nota_card(op, included):
            item = tk.Frame(list_outer, bg=C["surface2"],
                            highlightthickness=1,
                            highlightbackground=C["hair"] if included else C["err"])
            item.pack(fill="x", pady=(0, 8))
            scroll_targets.append(item)

            head = tk.Frame(item, bg=C["surface2"], padx=14, pady=10)
            head.pack(fill="x")
            scroll_targets.append(head)
            badge_txt = "No montante" if included else "Excluída"
            badge_fg = C["ok"] if included else C["err"]
            tk.Label(head, text=badge_txt, bg=C["surface2"],
                     fg=badge_fg, font=("Segoe UI", 7, "bold")).pack(side="left")
            if included:
                styled_button(
                    head, "Excluir", lambda o=op: self._toggle_nota(o, exclude=True),
                    danger=True, small=True).pack(side="right")
            else:
                styled_button(
                    head, "Incluir", lambda o=op: self._toggle_nota(o, exclude=False),
                    accent=True, small=True).pack(side="right")

            body = tk.Frame(item, bg=C["surface2"], padx=14, pady=12)
            body.pack(fill="x")
            scroll_targets.append(body)
            fields = [
                ("NF", op["nf"] or "—", False),
                ("Valor", op["valor"], True),
                ("Inclusão", op["data_inclusao"] or "—", False),
                ("Vencimento", op["data_vencimento"] or "—", False),
                ("Prazo", f"{op['prazo']} dias" if op.get("prazo") else "—", False),
            ]
            for label, value, accent in fields:
                row_f = tk.Frame(body, bg=C["surface2"])
                row_f.pack(fill="x", pady=(0, 3))
                scroll_targets.append(row_f)
                tk.Label(row_f, text=label, bg=C["surface2"], fg=C["ink_faint"],
                         font=("Segoe UI", 7, "bold"), width=10, anchor="w").pack(side="left")
                fg = C["ok"] if accent else C["ink"]
                tk.Label(row_f, text=value, bg=C["surface2"], fg=fg,
                         font=("Segoe UI", 8, "bold" if accent else "normal"),
                         anchor="w").pack(side="left", fill="x", expand=True)

        if incluidas:
            tk.Label(list_outer, text=f"NO MONTANTE ({len(incluidas)})", bg=C["surface"],
                     fg=C["ink_faint"], font=("Segoe UI", 7, "bold")).pack(
                         anchor="w", pady=(0, 6))
            scroll_targets.append(list_outer)
        for op in incluidas:
            _nota_card(op, included=True)

        if excluidas:
            tk.Label(list_outer, text=f"EXCLUÍDAS / DISPONÍVEIS ({len(excluidas)})",
                     bg=C["surface"], fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(14, 6))
            for op in excluidas:
                _nota_card(op, included=False)

        if not todas:
            empty = tk.Frame(list_outer, bg=C["surface"])
            empty.pack(fill="x", pady=24)
            scroll_targets.append(empty)
            tk.Label(empty, text="Nenhuma nota encontrada para este cliente na planilha.",
                     bg=C["surface"], fg=C["ink_muted"],
                     font=("Segoe UI", 9)).pack()

        self._bind_modal_scroll(scroll_targets, sf)

        def _sync_modal():
            sf.update_idletasks()
            sf._sync_scrollregion()
            sf.refresh_bindings()
        self.after_idle(_sync_modal)

    def _toggle_nota(self, op, exclude):
        uid = op["uid"]
        if exclude:
            if not messagebox.askyesno(
                "Excluir nota",
                f"Excluir a NF {op.get('nf') or '—'} ({op.get('valor')}) do montante "
                "deste cliente?",
                parent=self.controller):
                return
            self._excluded_uids.add(uid)
        else:
            self._excluded_uids.discard(uid)
        self._render_detalhes_list()
        self._rebuild_group_cards()

    def _rebuild_group_cards(self):
        excluded = self._excluded_uids
        visible_ops = [op for op in self._all_ops if op["uid"] not in excluded]
        groups = _group_invertido_ops(visible_ops)
        self._groups = groups
        for w in self._grid.winfo_children():
            w.destroy()
        self._cards = []
        self._limit_btns = {}
        if not groups:
            empty = tk.Frame(self._grid, bg=C["bg"])
            empty.grid(row=0, column=0, columnspan=3, sticky="ew", pady=24)
            tk.Label(empty, text="Nenhuma operação encontrada na planilha.",
                     bg=C["bg"], fg=C["ink_muted"],
                     font=("Segoe UI", 10)).pack()
            self._sub_lbl.configure(text="0 grupo(s) · 0 nota(s)")
            return
        for idx, group in enumerate(groups):
            row, col = divmod(idx, 3)
            self._cards.append(self._make_group_card(group, row, col))
        total_notas = sum(g["count"] for g in groups)
        self._sub_lbl.configure(
            text=(f"{len(groups)} grupo(s) · {total_notas} nota(s) · "
                  f"{os.path.basename(self._last_path or '')}"))

    def _close_alerts_modal(self):
        if self._alert_overlay is not None:
            try:
                self._alert_overlay.destroy()
            except Exception:
                pass
            self._alert_overlay = None

    def _update_alerts_footer(self):
        if not hasattr(self, "_alert_footer_lbl") or not self._alert_footer_lbl.winfo_exists():
            return
        pending = sum(
            1 for item in self._alert_items
            if self._alert_decisions.get(item["index"]) is None
        )
        if pending:
            self._alert_footer_lbl.configure(
                text=f"{pending} nota(s) aguardando decisão",
                fg=C["warn"])
            self._alert_continue_btn.configure(state="disabled")
        else:
            accepted = sum(
                1 for item in self._alert_items
                if self._alert_decisions.get(item["index"]) == "accept"
            )
            rejected = len(self._alert_items) - accepted
            self._alert_footer_lbl.configure(
                text=(f"{accepted} aceita(s) · {rejected} rejeitada(s) — "
                      "clique em Continuar para prosseguir"),
                fg=C["ok"])
            self._alert_continue_btn.configure(state="normal")

    def _set_alert_decision(self, item, decision, card_widgets):
        idx = item["index"]
        self._alert_decisions[idx] = decision
        badge = card_widgets["badge"]
        card = card_widgets["card"]
        if decision == "accept":
            badge.configure(text="Aceita", fg=C["ok"])
            card.configure(highlightbackground=C["ok"])
        else:
            badge.configure(text="Rejeitada", fg=C["err"])
            card.configure(highlightbackground=C["err"])
        self._update_alerts_footer()

    def _accept_all_alerts(self):
        for item, widgets in getattr(self, "_alert_item_widgets", []):
            self._set_alert_decision(item, "accept", widgets)

    def _reject_all_alerts(self):
        for item, widgets in getattr(self, "_alert_item_widgets", []):
            self._set_alert_decision(item, "reject", widgets)

    def _cancel_alerts_review(self):
        self._close_alerts_modal()
        self._pending_ops = []
        self._alert_items = []
        self._alert_decisions = {}
        self.controller.show_frame("OperacoesInvertido")

    def _finish_alerts_review(self):
        pending = any(
            self._alert_decisions.get(item["index"]) is None
            for item in self._alert_items
        )
        if pending:
            return
        rejected_uids = {
            item["op"]["uid"] for item in self._alert_items
            if self._alert_decisions.get(item["index"]) == "reject"
        }
        self._excluded_uids = set(rejected_uids)
        self._close_alerts_modal()
        self._pending_ops = []
        self._alert_items = []
        self._alert_decisions = {}
        self._render_results(self._all_ops)

    def _show_alerts_modal(self, ops, alerts):
        self._close_alerts_modal()
        self._pending_ops = ops
        self._alert_items = alerts
        self._alert_decisions = {}

        overlay = tk.Frame(self, bg="#0c0c0c")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._alert_overlay = overlay
        self._alert_item_widgets = []

        card = tk.Frame(overlay, bg=C["surface"],
                        highlightthickness=1, highlightbackground=C["hair"])
        card.place(relx=0.5, rely=0.5, anchor="center", width=780, height=620)
        card.bind("<Button-1>", lambda _e: "break")

        pad = tk.Frame(card, bg=C["surface"], padx=28, pady=22)
        pad.pack(fill="both", expand=True)

        top = tk.Frame(pad, bg=C["surface"])
        top.pack(fill="x")
        tk.Label(top, text="Alertas na análise", bg=C["surface"], fg=C["warn"],
                 font=("Segoe UI", 13, "bold")).pack(side="left")
        styled_button(top, "✕", self._cancel_alerts_review, small=True).pack(side="right")
        styled_button(top, "✕ Recusar todas", self._reject_all_alerts,
                      danger=True, small=True).pack(side="right", padx=(0, 8))
        styled_button(top, "✓ Aprovar todas", self._accept_all_alerts,
                      accent=True, small=True).pack(side="right", padx=(0, 8))

        tk.Label(
            pad,
            text=(f"{len(alerts)} nota(s) com pendências. Revise e decida se cada "
                  "uma deve seguir na operação."),
            bg=C["surface"], fg=C["ink_muted"], font=("Segoe UI", 9),
            wraplength=700, justify="left",
        ).pack(anchor="w", pady=(8, 0))

        make_hairline(pad, bg=C["hair"]).pack(fill="x", pady=(16, 0))

        scroll_wrap = tk.Frame(pad, bg=C["surface"], height=430)
        scroll_wrap.pack(fill="both", expand=True, pady=(12, 0))
        scroll_wrap.pack_propagate(False)

        sf = ScrollableFrame(scroll_wrap, bg=C["surface"])
        sf.pack(fill="both", expand=True)
        sf.link_wheel(scroll_wrap)
        sf.link_wheel(pad)
        sf.link_wheel(card)
        list_outer = sf.inner
        list_outer.configure(bg=C["surface"])

        all_scroll_targets = [overlay, card, pad, scroll_wrap, sf, list_outer]

        for item in alerts:
            op = item["op"]
            item_card = tk.Frame(list_outer, bg=C["surface2"],
                                 highlightthickness=1, highlightbackground=C["hair"])
            item_card.pack(fill="x", pady=(0, 10))
            all_scroll_targets.append(item_card)

            head = tk.Frame(item_card, bg=C["surface2"], padx=16, pady=10)
            head.pack(fill="x")
            all_scroll_targets.append(head)
            tk.Label(head, text=op["nome_sacado"], bg=C["surface2"], fg=C["ink"],
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(side="left", fill="x", expand=True)
            badge = tk.Label(head, text="Pendente", bg=C["surface2"], fg=C["warn"],
                             font=("Segoe UI", 8, "bold"))
            badge.pack(side="right")

            body = tk.Frame(item_card, bg=C["surface2"], padx=16)
            body.pack(fill="x", pady=(0, 10))
            all_scroll_targets.append(body)

            prazo_txt = f"{op['prazo']} dias" if op.get("prazo") else "—"
            for label, value, accent in (
                ("Nome Sacado", op.get("nome_sacado") or "—", False),
                ("NF", op.get("nf") or "—", False),
                ("Valor", op.get("valor") or "—", True),
                ("Vencimento", op.get("data_vencimento") or "—", False),
                ("Prazo", prazo_txt, False),
            ):
                row_f = tk.Frame(body, bg=C["surface2"])
                row_f.pack(fill="x", pady=(0, 2))
                all_scroll_targets.append(row_f)
                tk.Label(row_f, text=label, bg=C["surface2"], fg=C["ink_faint"],
                         font=("Segoe UI", 7, "bold"), width=12, anchor="w").pack(side="left")
                fg = C["ok"] if accent else C["ink"]
                tk.Label(row_f, text=value, bg=C["surface2"], fg=fg,
                         font=("Segoe UI", 8, "bold" if accent else "normal"),
                         anchor="w").pack(side="left", fill="x", expand=True)

            motivo_f = tk.Frame(body, bg=C["surface2"])
            motivo_f.pack(fill="x", pady=(6, 0))
            all_scroll_targets.append(motivo_f)
            tk.Label(motivo_f, text="Motivo", bg=C["surface2"], fg=C["ink_faint"],
                     font=("Segoe UI", 7, "bold"), width=12, anchor="w").pack(side="left", anchor="n")
            tk.Label(
                motivo_f,
                text=" · ".join(item["motivos"]),
                bg=C["surface2"], fg=C["warn"],
                font=("Segoe UI", 8), wraplength=560, justify="left",
            ).pack(side="left", fill="x", expand=True)

            btn_row = tk.Frame(item_card, bg=C["surface2"], padx=16)
            btn_row.pack(fill="x", pady=(6, 12))
            all_scroll_targets.append(btn_row)
            widgets = {"card": item_card, "badge": badge}
            self._alert_item_widgets.append((item, widgets))
            styled_button(
                btn_row, "✓ Aceitar",
                lambda it=item, w=widgets: self._set_alert_decision(it, "accept", w),
                accent=True, small=True,
            ).pack(side="left", padx=(0, 8))
            styled_button(
                btn_row, "✕ Rejeitar",
                lambda it=item, w=widgets: self._set_alert_decision(it, "reject", w),
                danger=True, small=True,
            ).pack(side="left")

        foot = tk.Frame(pad, bg=C["surface"])
        foot.pack(fill="x", pady=(14, 0))
        self._alert_footer_lbl = tk.Label(
            foot, text=f"{len(alerts)} nota(s) aguardando decisão",
            bg=C["surface"], fg=C["warn"], font=("Segoe UI", 9))
        self._alert_footer_lbl.pack(side="left")
        self._alert_continue_btn = styled_button(
            foot, "Continuar →", self._finish_alerts_review, accent=True, small=True)
        self._alert_continue_btn.configure(state="disabled")
        self._alert_continue_btn.pack(side="right")

        self._bind_modal_scroll(all_scroll_targets, sf)

        def _sync_modal():
            sf.update_idletasks()
            sf._sync_scrollregion()
            sf.refresh_bindings()
        self.after_idle(_sync_modal)

    def _on_parse_complete(self, ops):
        self._stop_loading_anim()
        self._loading_outer.pack_forget()
        self._all_ops = ops
        self._excluded_uids = set()
        alerts = _invertido_collect_alerts(ops)
        if alerts:
            self._show_alerts_modal(ops, alerts)
        else:
            self._render_results(ops)

    def _close_detalhes(self):
        if self._detail_overlay is not None:
            try:
                self._detail_overlay.destroy()
            except Exception:
                pass
            self._detail_overlay = None
        self._detail_key = None

    def _render_results(self, ops):
        self._stop_loading_anim()
        self._loading_outer.pack_forget()
        for w in self._grid.winfo_children():
            w.destroy()
        self._cards = []
        excluded = getattr(self, "_excluded_uids", set())
        visible_ops = [op for op in ops if op.get("uid") not in excluded]
        groups = _group_invertido_ops(visible_ops)
        self._groups = groups
        if not groups:
            empty = tk.Frame(self._grid, bg=C["bg"])
            empty.grid(row=0, column=0, columnspan=3, sticky="ew", pady=24)
            tk.Label(empty, text="Nenhuma operação encontrada na planilha.",
                     bg=C["bg"], fg=C["ink_muted"],
                     font=("Segoe UI", 10)).pack()
            return
        for idx, group in enumerate(groups):
            row, col = divmod(idx, 3)
            self._cards.append(self._make_group_card(group, row, col))
        total_notas = sum(g["count"] for g in groups)
        self._sub_lbl.configure(
            text=(f"{len(groups)} grupo(s) · {total_notas} nota(s) · "
                  f"{os.path.basename(self._last_path or '')}"))
        self._limit_btns = {}
        for card in self._cards:
            if card.get("lim_btn") and card.get("group"):
                self._limit_btns[self._group_limite_key(card["group"])] = card["lim_btn"]
        self.update_idletasks()
        self._sf._sync_scrollregion()
        self._sf.refresh_bindings()
        self._refresh_all_limit_buttons()

    def _show_error(self, msg):
        self._stop_loading_anim()
        self._loading_lbl.configure(text=msg, fg=C["err"])
        self._loading_icon.configure(text="✗", fg=C["err"])

    def _start_worker(self, path):
        if self._worker_running:
            return
        self._worker_running = True
        threading.Thread(target=self._worker, args=(path,), daemon=True).start()

    def _worker(self, path):
        try:
            ops = _parse_invertido_xlsx(path)
            self._ui(lambda o=ops: self._on_parse_complete(o))
        except Exception as e:
            self._ui(lambda m=str(e): self._show_error(f"Erro ao analisar: {m}"))
        finally:
            self._worker_running = False


class LimitesInvertidoFrame(tk.Frame, ThreadSafeUIMixin):
    COL_OK   = C["ok"]
    COL_WARN = C["warn"]
    COL_ERR  = C["err"]

    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller  = controller
        self._cards      = []
        self._worker_running  = False
        self._cancel_requested= False
        self._browser    = None
        self._started    = False
        self._restart_pending = False
        self._init_ui_queue()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24,0))
        styled_button(hdr, "← Voltar",
                      lambda: self.controller.show_frame("OperacoesInvertido")).pack(side="left")
        self._restart_btn = styled_button(hdr, "↻  Atualizar",
                                          self._restart_limites)
        self._restart_btn.pack(side="right")

        tk.Label(hdr, text="Limites Invertido", bg=C["bg"], fg=C["ink"],
                 font=("Georgia",18,"bold")).pack(side="left", padx=(14,0))

        sub = tk.Frame(self, bg=C["bg"])
        sub.pack(fill="x", padx=32)
        tk.Label(sub, text="Consulta o LTC e o limite disponível de cada cliente.",
                 bg=C["bg"], fg=C["ink_muted"], font=("Segoe UI",9)).pack(anchor="w", pady=(4,0))

        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(16,0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        self._grid_outer = self._sf.inner
        self._grid_outer.configure(bg=C["bg"])

        self._grid = tk.Frame(self._grid_outer, bg=C["bg"])
        self._grid.pack(padx=32, pady=(16,24), fill="x")
        for c in range(3):
            self._grid.columnconfigure(c, weight=1, uniform="lcards")

    def on_show(self):
        self._sf.refresh_bindings()
        if self._started: return
        self._started = True
        self._cancel_requested = False
        self._setup_cards()
        self._start_worker()

    def _restart_limites(self):
        self._restart_pending  = True
        self._cancel_requested = True
        br = getattr(self,"_browser",None)
        if br:
            try: br.close()
            except: pass
        self.after(200, self._check_restart)

    def _check_restart(self):
        if not self._restart_pending: return
        if self._worker_running: self.after(220, self._check_restart); return
        self._restart_pending = False
        self._cards = []
        for w in self._grid.winfo_children(): w.destroy()
        self._started = False
        self._cancel_requested = False
        self._setup_cards()
        self._started = True
        self._start_worker()

    def _setup_cards(self):
        for w in self._grid.winfo_children(): w.destroy()
        self._cards = []
        all_clients = list(BPM_CLIENT_DATA.keys())
        for idx, name in enumerate(all_clients):
            is_mapped = name in MAPPED_CLIENTS
            is_mirror = name in MIRROR_CLIENTS
            row, col  = divmod(idx, 3)
            c = self._make_card(name, row, col, is_mapped, is_mirror)
            self._cards.append(c)
        for i in range(len(self._cards)):
            self.after(50*i, lambda i=i: self._reveal_card(i))

    def _make_card(self, name, row, col, is_mapped, is_mirror):
        bg = C["surface"]; bord = C["hair"]
        outer = tk.Frame(self._grid, bg=C["bg"])
        outer.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)
        outer.grid_remove()

        card = tk.Frame(outer, bg=bg, highlightthickness=1, highlightbackground=bord, bd=0)
        card.pack(fill="both", expand=True)
        top_bar = tk.Frame(card, bg=bord, height=2)
        top_bar.pack(fill="x")

        body = tk.Frame(card, bg=bg, padx=14, pady=12)
        body.pack(fill="both", expand=True)

        icon_lbl = tk.Label(body, text="–", bg=bg, fg=C["ink_faint"],
                            font=("Segoe UI",14,"bold"))
        icon_lbl.pack()
        tk.Label(body, text=name, bg=bg, fg=C["ink_muted"],
                 font=("Segoe UI",9,"bold")).pack(pady=(6,0))
        if is_mirror:
            src = MIRROR_CLIENTS[name]
            tk.Label(body, text=f"via {src}", bg=bg, fg=C["ink_faint"],
                     font=("Segoe UI",7,"italic")).pack()

        info_lbl = tk.Label(body, text="", bg=bg, fg=C["ink_muted"],
                            font=("Segoe UI",8), wraplength=160, justify="center")
        info_lbl.pack(pady=(6,0))

        status_row = tk.Frame(body, bg=bg)
        status_row.pack(pady=(8,0))
        status_lbl = tk.Label(status_row, text="EM ESPERA" if is_mapped or is_mirror else "A MAPEAR",
                              bg=bg, fg=C["ink_faint"], font=("Segoe UI",7,"bold"))
        status_lbl.pack()

        spin_id = [None]
        angle   = [0]

        def tick():
            angle[0] = (angle[0]+20)%360
            if spin_id[0] is not None:
                status_lbl.configure(text=["◐","◓","◑","◒"][angle[0]//90])
            spin_id[0] = outer.after(120, tick) if spin_id[0] else None

        return {
            "outer":outer,"card":card,"top_bar":top_bar,"body":body,
            "icon_lbl":icon_lbl,"info_lbl":info_lbl,"status_lbl":status_lbl,
            "status_row":status_row,
            "spin_id":spin_id,"angle":angle,"tick":tick,
            "is_mapped":is_mapped,"is_mirror":is_mirror,"name":name,"state":"init"
        }

    def _reveal_card(self, idx):
        if idx >= len(self._cards): return
        c = self._cards[idx]
        c["outer"].grid()
        state = "waiting" if c["is_mapped"] else ("mirror_waiting" if c["is_mirror"] else "pending")
        self._set_state(idx, state)

    def _set_state(self, idx, state, info=""):
        if idx >= len(self._cards): return
        c = self._cards[idx]
        if c["spin_id"][0]:
            try: c["outer"].after_cancel(c["spin_id"][0])
            except: pass
            c["spin_id"][0] = None
        c["state"] = state

        cfg = {
            "waiting":       (C["surface"],  C["hair"],     "–",  C["ink_faint"],  "EM ESPERA",   C["ink_faint"]),
            "mirror_waiting":(C["surface"],  C["hair"],     "–",  C["ink_faint"],  "AGUARDANDO",  C["ink_faint"]),
            "pending":       (C["surface2"], C["hair"],     "–",  C["ink_faint"],  "A MAPEAR",    C["ink_faint"]),
            "processing":    (C["surface"],  C["ok"],       "…",  C["ok"],         "CONSULTANDO", C["ok"]),
            "ok":            (C["surface"],  C["ok"],       "✓",  C["ok"],         "LIMITE OK",   C["ok"]),
            "warn":          (C["surface"],  C["warn"],     "!",  C["warn"],        "ATENÇÃO",     C["warn"]),
            "error":         (C["surface"],  C["err"],      "✗",  C["err"],         "ERRO",        C["err"]),
            "ltc_expired":   (C["surface"],  C["err"],      "✗",  C["err"],         "LTC VENCIDO", C["err"]),
        }.get(state, (C["surface"], C["hair"], "–", C["ink_faint"], state.upper(), C["ink_faint"]))

        bg, bord, icon_t, icon_fg, status_t, status_fg = cfg
        c["card"].configure(bg=bg, highlightbackground=bord)
        c["top_bar"].configure(bg=bord)
        c["body"].configure(bg=bg)
        c["icon_lbl"].configure(bg=bg, text=icon_t, fg=icon_fg)
        c["info_lbl"].configure(bg=bg, text=info, fg=C["ink_muted"])
        c["status_lbl"].configure(bg=bg, text=status_t, fg=status_fg)
        c["status_row"].configure(bg=bg)
        for w in c["body"].winfo_children():
            try: w.configure(bg=bg)
            except: pass

        if state == "processing":
            c["spin_id"][0] = True
            c["tick"]()

    def _publish_limite(self, name, ltc_str=None, ltc_date=None, limite_disp=None,
                        state="processing", info=""):
        cnpj = only_digits(BPM_CLIENT_DATA.get(name, {}).get("CNPJ", ""))
        if not cnpj:
            return
        data = {
            "client_name": name,
            "cnpj": cnpj,
            "ltc_str": ltc_str,
            "ltc_date": ltc_date,
            "limite_disp": limite_disp,
            "state": state,
            "info": info,
        }
        self.controller.publish_limite_result(cnpj, data)
        for mn in LIMITE_SHARED_RESULTS.get(name, []):
            mcnpj = only_digits(BPM_CLIENT_DATA.get(mn, {}).get("CNPJ", ""))
            if not mcnpj:
                continue
            mdata = dict(data)
            mdata["client_name"] = mn
            mdata["via"] = name
            self.controller.publish_limite_result(mcnpj, mdata)

    def _start_worker(self):
        if self._worker_running: return
        self._worker_running = True
        threading.Thread(target=self._worker, daemon=True).start()

    def _find_idx(self, name):
        for i, c in enumerate(self._cards):
            if c["name"] == name: return i
        return -1

    def _worker(self):
        if not PLAYWRIGHT_OK:
            self._ui(lambda: messagebox.showerror("Erro","Playwright não disponível."))
            self._worker_running = False; return
        today = date.today()
        with sync_playwright() as p:
            browser = None
            try:
                for ch in ("chrome","msedge"):
                    try: browser = p.chromium.launch(channel=ch, headless=False); break
                    except: pass
                if browser is None:
                    try: browser = p.chromium.launch(headless=False)
                    except Exception as e: raise RuntimeError(f"Não foi possível iniciar navegador: {e}")
                self._browser = browser
                page = browser.new_page()
                page.set_default_timeout(60_000)
                RE_DATE = re.compile(r"\d{2}/\d{2}/\d{4}")

                for name in list(BPM_CLIENT_DATA.keys()):
                    if self._cancel_requested: break
                    if name not in MAPPED_CLIENTS: continue
                    idx = self._find_idx(name)
                    if idx == -1: continue
                    self._publish_limite(name, state="processing")
                    self._ui(lambda i=idx: self._set_state(i, "processing"))
                    url = LIMITE_CLIENT_URLS.get(name)
                    if not url:
                        self._ui(lambda i=idx: self._set_state(i,"error","URL não mapeada")); continue
                    try:
                        # ── PATCH 2: espera inteligente de carregamento ──────────
                        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
                        if self._cancel_requested: break

                        # Espera networkidle (max 30s) — sem sleep fixo
                        try:
                            page.wait_for_load_state("networkidle", timeout=30_000)
                        except Exception:
                            pass

                        # Polling JS: aguarda spinners/loaders sumirem (max 25s)
                        _deadline = time.time() + 25
                        while time.time() < _deadline:
                            if self._cancel_requested:
                                break
                            try:
                                _loading = page.evaluate(
                                    "(function(){"
                                    "var sels=['.loading','.spinner','[class*=\"loading\"]',"
                                    "'[class*=\"spinner\"]','.u-loading',"
                                    "'[aria-busy=\"true\"]','.itau-spinner'];"
                                    "for(var i=0;i<sels.length;i++){"
                                    "var els=document.querySelectorAll(sels[i]);"
                                    "for(var j=0;j<els.length;j++){"
                                    "var s=window.getComputedStyle(els[j]);"
                                    "if(s.display!=='none'&&s.visibility!=='hidden'"
                                    "&&s.opacity!=='0') return true;}}"
                                    "return false;})()"
                                )
                                if not _loading:
                                    break
                            except Exception:
                                break
                            time.sleep(0.5)

                        # Pausa mínima de segurança para renderização final
                        time.sleep(2)
                        if self._cancel_requested: break

                        # ler LTC — 3 tentativas com pausa entre elas
                        ltc_str = None
                        for _ltc_try in range(3):
                            try:
                                all_spans = page.locator("span.u-font-size--14.u-ml--8.u-mt--8.u-block")
                                for si in range(all_spans.count()):
                                    txt = (all_spans.nth(si).inner_text() or "").strip()
                                    if RE_DATE.match(txt):
                                        ltc_str = txt
                                        break
                            except Exception:
                                pass
                            if not ltc_str:
                                try:
                                    body_text = page.locator("body").inner_text()
                                    m = RE_DATE.search(body_text)
                                    if m:
                                        ltc_str = m.group(0)
                                except Exception:
                                    pass
                            if ltc_str:
                                break
                            time.sleep(2)
                        # ────────────────────────────────────────────────────────

                        try:
                            ltc_date = datetime.strptime(ltc_str,"%d/%m/%Y").date() if ltc_str else None
                        except: ltc_date = None

                        if ltc_date is None:
                            inf = "Não foi possível ler\ndata do LTC"
                            self._publish_limite(name, ltc_str=ltc_str, state="error", info=inf)
                            self._ui(lambda i=idx,inf=inf: self._set_state(i,"error",inf))
                            for mn in LIMITE_SHARED_RESULTS.get(name,[]):
                                mi = self._find_idx(mn)
                                if mi!=-1: self._ui(lambda i=mi,inf=inf: self._set_state(i,"error",inf))
                            continue

                        if ltc_date <= today:
                            inf = f"LTC vencido em {ltc_str}"
                            self._publish_limite(name, ltc_str=ltc_str, ltc_date=ltc_date,
                                                 state="ltc_expired", info=inf)
                            self._ui(lambda i=idx,inf=inf: self._set_state(i,"ltc_expired",inf))
                            for mn in LIMITE_SHARED_RESULTS.get(name,[]):
                                mi = self._find_idx(mn)
                                if mi!=-1: self._ui(lambda i=mi,inf=inf: self._set_state(i,"ltc_expired",inf))
                            continue

                        # ler limite
                        limite_disp = None
                        BAIXO = 1_000_000
                        try:
                            val_js = page.evaluate("""
                                (function(){
                                    var rows=document.querySelectorAll('table.atual tbody tr');
                                    for(var i=0;i<rows.length;i++){
                                        var n=rows[i].querySelector('td.tdNomeFinalidade');
                                        if(!n||String(n.textContent).trim().toLowerCase()!=='fornecedor') continue;
                                        var d=rows[i].querySelector('td[id^="valorDisponibilidade_"]');
                                        if(!d) continue;
                                        return String(d.textContent).trim();
                                    }
                                    return null;
                                })()
                            """)
                            if val_js:
                                s = val_js.replace(".","").replace(",","")
                                try: limite_disp = int(s)
                                except: pass
                            if limite_disp is None:
                                tds = page.locator("td.tdValorDisp")
                                for ti in range(tds.count()):
                                    s = re.sub(r"[^\d]","", tds.nth(ti).inner_text() or "")
                                    try:
                                        v = int(s)
                                        if limite_disp is None or v > limite_disp: limite_disp = v
                                    except: pass
                        except: pass

                        disp_fmt = f"R$ {limite_disp:,}".replace(",",".") if limite_disp is not None else "N/D"
                        inf = f"LTC ativo · vence {ltc_str}\nLimite Disp. (fornecedor): {disp_fmt}"
                        warn = limite_disp is not None and limite_disp < BAIXO
                        final_state = "warn" if warn else "ok"
                        if warn: inf += f"\n⚠ Limite abaixo de R$ {BAIXO:,}".replace(",",".")

                        self._publish_limite(name, ltc_str=ltc_str, ltc_date=ltc_date,
                                             limite_disp=limite_disp, state=final_state, info=inf)
                        self._ui(lambda i=idx,s=final_state,inf=inf: self._set_state(i,s,inf))
                        for mn in LIMITE_SHARED_RESULTS.get(name,[]):
                            mi = self._find_idx(mn)
                            if mi!=-1:
                                minf = inf + f"\n(via {name})"
                                self._ui(lambda i=mi,s=final_state,inf=minf: self._set_state(i,s,inf))

                    except BPMUserCancelled: break
                    except Exception as e:
                        if self._cancel_requested: break
                        es = str(e)[:80]
                        self._publish_limite(name, state="error", info=es)
                        self._ui(lambda i=idx,es=es: self._set_state(i,"error",es))

            except Exception as e:
                if not self._cancel_requested:
                    em = str(e)
                    self._ui(lambda: messagebox.showerror("Erro na consulta",em))
            finally:
                self._browser = None
                if browser:
                    try: browser.close()
                    except: pass
                self._worker_running = False


class BPMFrame(tk.Frame, ThreadSafeUIMixin):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=C["bg"])
        self.controller = controller
        self._cards   = []
        self._selected= []
        self._worker_running   = False
        self._cancel_requested = False
        self._browser = None
        self._started = False
        self._mode = "invertido"
        self._init_ui_queue()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill="x", padx=32, pady=(24,0))
        tk.Label(hdr, text="BPM — Operações em andamento", bg=C["bg"], fg=C["ink"],
                 font=("Georgia",18,"bold")).pack(side="left")
        self._cancel_btn = styled_button(hdr, "✕  Cancelar",
                                         self._on_cancel, danger=True)
        self._cancel_btn.pack(side="right")
        make_hairline(self, bg=C["hair"]).pack(fill="x", padx=0, pady=(14,0))

        self._sf = ScrollableFrame(self, bg=C["bg"])
        self._sf.pack(fill="both", expand=True)
        self._sf.link_wheel(self)
        self._grid_wrap = self._sf.inner
        self._grid_wrap.configure(bg=C["bg"])

        self._grid = tk.Frame(self._grid_wrap, bg=C["bg"])
        self._grid.pack(padx=32, pady=(16,8), fill="x")
        for c in range(3):
            self._grid.columnconfigure(c, weight=1, uniform="bcards")

        make_hairline(self._grid_wrap, bg=C["hair"]).pack(fill="x", padx=32, pady=(8,0))
        log_hdr = tk.Frame(self._grid_wrap, bg=C["bg"])
        log_hdr.pack(fill="x", padx=32, pady=(10,4))
        tk.Label(log_hdr, text="Log de execução", bg=C["bg"], fg=C["ink_muted"],
                 font=("Segoe UI",8,"bold")).pack(side="left")
        styled_button(log_hdr, "Limpar log",
                      self._clear_log, small=True).pack(side="right")

        log_frame = tk.Frame(self._grid_wrap, bg=C["bg"])
        log_frame.pack(fill="x", padx=32, pady=(0,24))
        log_frame.columnconfigure(0, weight=1)
        self._log = tk.Text(log_frame, height=10, wrap="word", bd=0, relief="flat",
                            bg=C["surface"], fg=C["log_step"],
                            font=("Consolas",8), padx=10, pady=8,
                            state="disabled")
        self._log.grid(row=0, column=0, sticky="ew")
        lsb = MinimalScrollbar(log_frame, command=self._log.yview, bg=C["bg"])
        lsb.grid(row=0, column=1, sticky="ns")
        self._log.configure(yscrollcommand=lsb.set)
        bind_text_mousewheel(self._log)
        self._log.tag_configure("step",    foreground=C["log_step"])
        self._log.tag_configure("heading", foreground=C["ink"])
        self._log.tag_configure("ok",      foreground=C["log_ok"])
        self._log.tag_configure("warn",    foreground=C["log_warn"])
        self._log.tag_configure("err",     foreground=C["log_err"])

    def _log_line(self, msg, tag="step"):
        def _do():
            self._log.configure(state="normal")
            self._log.insert("end", f"{datetime.now().strftime('%H:%M:%S')}  {msg}\n", tag)
            self._log.see("end")
            self._log.configure(state="disabled")
        self._ui(_do)

    def _clear_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0","end")
        self._log.configure(state="disabled")

    def on_show(self):
        self._sf.refresh_bindings()
        sel = getattr(self.controller,"bpm_run_selection",None) or []
        if not sel: return
        self._cancel_requested = False
        if self._started: return
        self._started = True
        self._selected = sel
        self._mode = getattr(self.controller, "bpm_run_mode", "invertido")
        self._setup_cards()
        self._start_worker()

    def _on_cancel(self):
        self._cancel_requested = True
        self.controller.bpm_run_selection = []
        br = getattr(self,"_browser",None)
        if br:
            try: br.close()
            except: pass
        self._reset()
        self._started = False
        dest = "BPM_CONFIG_NOVA" if self._mode == "nova_plataforma" else "BPM_CONFIG"
        self.controller.show_frame(dest)

    def _reset(self):
        self._cards = []
        for w in self._grid.winfo_children(): w.destroy()

    def _setup_cards(self):
        self._reset()
        for idx, item in enumerate(self._selected):
            cli = item.get("cliente","")
            raw = (item.get("valor","")).strip()
            display = raw[2:].strip() if raw.lower().startswith("r$") else raw
            row, col = divmod(idx,3)
            c = self._make_card(cli, display, row, col)
            self._cards.append(c)
        for i in range(1, len(self._cards)):
            self.after(60*i, lambda i=i: self._set_card_state(i,"waiting",show=True))
        if self._cards:
            self.after(0, lambda: self._set_card_state(0,"processing",show=True))

    def _make_card(self, name, val_display, row, col):
        bg = C["surface"]; bord = C["hair"]
        outer = tk.Frame(self._grid, bg=C["bg"])
        outer.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)
        outer.grid_remove()
        card = tk.Frame(outer, bg=bg, highlightthickness=1, highlightbackground=bord)
        card.pack(fill="both", expand=True)
        top_bar = tk.Frame(card, bg=bord, height=2)
        top_bar.pack(fill="x")
        body = tk.Frame(card, bg=bg, padx=14, pady=12)
        body.pack(fill="both", expand=True)
        icon_lbl = tk.Label(body, text="–", bg=bg, fg=C["ink_faint"],
                            font=("Segoe UI",14,"bold"))
        icon_lbl.pack()
        tk.Label(body, text=name, bg=bg, fg=C["ink"],
                 font=("Segoe UI",10,"bold")).pack(pady=(6,0))
        tk.Label(body, text=f"R$ {val_display}", bg=bg, fg=C["ink_muted"],
                 font=("Segoe UI",8)).pack(pady=(2,0))
        status_lbl = tk.Label(body, text="EM ESPERA", bg=bg, fg=C["ink_faint"],
                              font=("Segoe UI",7,"bold"))
        status_lbl.pack(pady=(8,0))
        angle = [0]; spin_id = [None]
        def tick():
            angle[0]=(angle[0]+90)%360
            status_lbl.configure(text=["◐","◓","◑","◒"][angle[0]//90])
            spin_id[0] = outer.after(180, tick) if spin_id[0] else None
        return {"outer":outer,"card":card,"top_bar":top_bar,"body":body,
                "icon_lbl":icon_lbl,"status_lbl":status_lbl,
                "spin_id":spin_id,"angle":angle,"tick":tick,"state":"init"}

    def _set_card_state(self, idx, state, show=False):
        if idx >= len(self._cards): return
        c = self._cards[idx]
        if c["spin_id"][0]:
            try: c["outer"].after_cancel(c["spin_id"][0])
            except: pass
            c["spin_id"][0] = None
        c["state"] = state

        cfg = {
            "waiting":    (C["surface"],  C["hair"],  "–",  C["ink_faint"], "EM ESPERA",  C["ink_faint"]),
            "processing": (C["surface"],  C["ok"],    "…",  C["ok"],        "EXECUTANDO", C["ok"]),
            "done":       (C["surface"],  C["ok"],    "✓",  C["ok"],        "CONCLUÍDO",  C["ok"]),
            "error":      (C["surface"],  C["err"],   "✗",  C["err"],       "ERRO",       C["err"]),
        }.get(state, (C["surface"], C["hair"], "–", C["ink_faint"], state, C["ink_faint"]))

        bg,bord,ic,ic_fg,st,st_fg = cfg
        c["card"].configure(bg=bg, highlightbackground=bord)
        c["top_bar"].configure(bg=bord)
        c["body"].configure(bg=bg)
        c["icon_lbl"].configure(bg=bg, text=ic, fg=ic_fg)
        c["status_lbl"].configure(bg=bg, text=st, fg=st_fg)
        for w in c["body"].winfo_children():
            try: w.configure(bg=bg)
            except: pass
        if show: c["outer"].grid()
        if state=="processing":
            c["spin_id"][0] = True
            c["tick"]()

    def _update_card_status(self, idx, text):
        if idx >= len(self._cards): return
        self._cards[idx]["status_lbl"].configure(text=text)

    def _start_worker(self):
        if self._worker_running: return
        self._worker_running = True
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        if not PLAYWRIGHT_OK:
            self._ui(lambda: messagebox.showerror("Erro","Playwright não disponível."))
            self._worker_running = False; return

        funcional = (getattr(self.controller,"bpm_funcional","") or "").strip()
        senha     = (getattr(self.controller,"bpm_password","") or "").strip()
        if not funcional or not senha:
            self._ui(lambda: messagebox.showerror("BPM","Credenciais não informadas."))
            self._worker_running = False; return

        PAINEL = "https://painelservicos.cloud.ihf/AplicAutSolicitacoesMiddle"
        PACE   = 0.85

        def pace(a=1.15, b=1.75):
            time.sleep(random.uniform(a*PACE, b*PACE))

        def pick_browser(p):
            for ch in ("chrome","msedge"):
                try: return p.chromium.launch(channel=ch, headless=False)
                except: pass
            return p.chromium.launch(headless=False)

        def wait_enabled(loc, timeout_ms=120_000):
            end = time.time() + timeout_ms/1000
            loc.wait_for(state="visible", timeout=timeout_ms)
            while time.time() < end:
                if self._cancel_requested: raise BPMUserCancelled()
                try:
                    if not loc.is_disabled(): return
                except: pass
                time.sleep(0.2)
            raise TimeoutError("Timeout aguardando campo habilitar.")

        def _to_dec(s):
            s = re.sub(r"[^\d,.\-]","",s or "")
            if not s: return None
            ld,lc = s.rfind("."),s.rfind(",")
            si = max(ld,lc)
            try:
                if si==-1: d=Decimal(re.sub(r"[^\d\-]","",s))
                else:
                    ip=re.sub(r"[^\d\-]","",s[:si]); dp=re.sub(r"[^\d]","",s[si+1:])
                    if not dp: return None
                    d=Decimal((ip if ip not in {"","-"} else "0")+"."+dp)
            except: return None
            return d.quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)

        def fill_currency(ctx, loc, amount, timeout_ms=120_000):
            wait_enabled(loc, timeout_ms)
            loc.scroll_into_view_if_needed(); loc.click()
            pace(0.18,0.35); loc.press("Control+A"); loc.press("Backspace")
            pace(0.12,0.25); loc.press("Delete"); pace(0.15,0.30)
            digits = re.sub(r"\D","",amount or "")
            if not digits: raise RuntimeError(f"Valor inválido: {amount}")
            for ch in digits:
                loc.type(ch, delay=random.randint(int(85*PACE),int(150*PACE)))
            loc.press("Tab"); pace(0.45,0.85)
            exp = _to_dec(amount); got = _to_dec(loc.input_value())
            if exp is None or got is None or got != exp:
                raise RuntimeError(f"Valor mascarado divergente. Esperado {amount}, obtido '{loc.input_value()}'.")

        def resolve_frame(timeout_ms=120_000):
            deadline = time.time()+timeout_ms/1000
            while time.time()<deadline:
                if self._cancel_requested: raise BPMUserCancelled()
                for frame in list(page.frames):
                    try:
                        if frame.is_detached(): continue
                        if frame.locator("#nova-solicitacao").count()>0: return frame
                    except: pass
                time.sleep(0.35*PACE)
            raise TimeoutError("Timeout: #nova-solicitacao não encontrado.")

        def wait_continuar(ctx, timeout_ms=120_000):
            deadline = time.time()+timeout_ms/1000
            while time.time()<deadline:
                if self._cancel_requested: raise BPMUserCancelled()
                ok = ctx.evaluate("""(function(){
                    var el=document.querySelector('#continuar');
                    if(!el||el.disabled) return false;
                    var s=window.getComputedStyle(el);
                    if(s.display==='none'||s.visibility==='hidden') return false;
                    return true;
                })()""")
                if ok: return
                time.sleep(0.25)
            raise TimeoutError("Timeout: #continuar não ficou clicável.")

        def click_continuar(ctx, timeout_ms=120_000):
            loc = ctx.locator('input#continuar[type="submit"]').first
            for action in [
                lambda: loc.click(timeout=12_000),
                lambda: loc.click(force=True, timeout=12_000),
                lambda: ctx.evaluate("document.querySelector('#continuar').click()"),
            ]:
                try: action(); return
                except: pass
            raise RuntimeError("Falha ao clicar #continuar.")

        def wait_nova(ctx, timeout_ms=120_000):
            deadline = time.time()+timeout_ms/1000
            while time.time()<deadline:
                if self._cancel_requested: raise BPMUserCancelled()
                ok = ctx.evaluate("""(function(){
                    var el=document.querySelector('#nova-solicitacao');
                    if(!el||el.disabled) return false;
                    var s=window.getComputedStyle(el);
                    return !(s.display==='none'||s.visibility==='hidden');
                })()""")
                if ok: return
                time.sleep(0.4)
            raise TimeoutError("Timeout: #nova-solicitacao.")

        def click_nova(ctx, timeout_ms=120_000):
            wait_nova(ctx, timeout_ms)
            def done():
                try: ctx.locator("#tipooper").wait_for(state="visible",timeout=5_000); return True
                except: return False
            for action in [
                lambda: ctx.locator("#nova-solicitacao").first.click(timeout=8_000),
                lambda: ctx.evaluate("document.querySelector('#nova-solicitacao').click()"),
                lambda: ctx.locator("#nova-solicitacao").first.click(force=True,timeout=8_000),
            ]:
                try: action()
                except: pass
                for _ in range(12):
                    if done(): return
                    time.sleep(0.5)
            raise RuntimeError("Falha ao clicar #nova-solicitacao.")

        def ensure_painel():
            if self._cancel_requested: raise BPMUserCancelled()
            page.goto(PAINEL, wait_until="domcontentloaded")
            if "/Home" in page.url: page.goto(PAINEL, wait_until="domcontentloaded")
            if page.locator("#username").count()>0 or "login.itau/oauth" in page.url:
                page.locator("#username").fill(funcional)
                page.locator("#password").fill(senha)
                page.locator("#btLogin").click()
                page.wait_for_load_state("domcontentloaded")
                page.goto(PAINEL, wait_until="domcontentloaded")
            ctx = resolve_frame(timeout_ms=120_000)
            wait_nova(ctx, timeout_ms=120_000)
            return ctx

        STEP = 15_000
        with sync_playwright() as p:
            browser = pick_browser(p)
            self._browser = browser
            page = browser.new_page()
            page.set_default_timeout(120_000)
            try:
                for idx, item in enumerate(self._selected):
                    if self._cancel_requested: raise BPMUserCancelled()
                    client_name = item.get("cliente","")
                    raw_amount  = item.get("valor","")
                    if self._mode == "nova_plataforma":
                        info = {"CNPJ": item.get("CNPJ",""), "AG": item.get("AG",""),
                                 "CONTA": item.get("CONTA",""), "PLATAFORMA": item.get("PLATAFORMA","")}
                        if not all(info.values()):
                            self._log_line(f"[{client_name}] Dados incompletos.", "err"); continue
                    else:
                        info = BPM_CLIENT_DATA.get(client_name)
                        if not info:
                            self._log_line(f"[{client_name}] Sem mapeamento interno.", "err"); continue
                    amount_web = _fmt_brl_plain_web(raw_amount)
                    if not amount_web:
                        self._log_line(f"[{client_name}] Valor inválido: {raw_amount}", "err"); continue

                    self._log_line(f"━━ {client_name} ━━", "heading")
                    self._log_line("  Razão social: Indefinido", "step")
                    tentativa = 0
                    while True:
                        if self._cancel_requested: raise BPMUserCancelled()
                        tentativa += 1
                        try:
                            ctx = ensure_painel()
                            self._log_line(f"  Painel pronto.", "step")
                            click_nova(ctx)
                            pace()
                            self._log_line("  Nova solicitação aberta.", "step")

                            ctx.locator("#tipooper").wait_for(state="visible",timeout=STEP)
                            ctx.locator("#tipooper").select_option(value="AT")
                            pace()
                            ctx.locator("#CPNJ").fill(info["CNPJ"]); pace()
                            ctx.locator('input[name="agenciafiltro"]').fill(info["AG"]); pace()
                            ctx.locator('input[name="contadacfiltro"]').fill(info["CONTA"]); pace()
                            self._log_line("  Dados do cliente preenchidos.", "step")

                            loc_proc = ctx.locator("#processar-nova")
                            wait_enabled(loc_proc, STEP); loc_proc.click(); pace()

                            loc_plat = ctx.locator('input[name="plataforma"]')
                            wait_enabled(loc_plat,STEP); loc_plat.fill(info["PLATAFORMA"]); pace()

                            loc_fn = ctx.locator("#funcionalac")
                            wait_enabled(loc_fn,STEP); loc_fn.fill(funcional)
                            loc_fn.press("Tab")
                            try: loc_fn.evaluate("el=>{el.dispatchEvent(new Event('input',{bubbles:true}));el.dispatchEvent(new Event('change',{bubbles:true}));el.blur();}")
                            except: pass
                            time.sleep(0.35*PACE)
                            wait_continuar(ctx,STEP); click_continuar(ctx,STEP); pace()
                            self._log_line("  Filtro AplicAut preenchido.", "step")

                            produto_val = "20" if self._mode == "nova_plataforma" else "34"
                            tpop_val = "0" if self._mode == "nova_plataforma" else "1"
                            ctx.locator('select[name="produto"]').wait_for(state="visible",timeout=STEP)
                            ctx.locator('select[name="produto"]').select_option(value=produto_val); pace()
                            ctx.locator('select[name="tpOperacao"]').wait_for(state="visible",timeout=STEP)
                            ctx.locator('select[name="tpOperacao"]').select_option(value=tpop_val); pace()

                            loc_buscar = ctx.locator('input[type="submit"][data-ng-click="BuscarListaPN()"]')
                            if loc_buscar.count()==0: loc_buscar=ctx.locator('input[type="submit"][value="Processar"]')
                            wait_enabled(loc_buscar,STEP); loc_buscar.click(); pace()

                            checkbox = ctx.locator('input[type="checkbox"]')
                            checkbox.first.wait_for(state="visible",timeout=STEP)
                            checkbox.first.click(); pace()

                            loc_val = ctx.locator('input[name="ValordaOperacao0"]')
                            self._log_line(f"  Preenchendo valor {raw_amount}…", "step")
                            fill_currency(ctx, loc_val, amount_web, STEP); pace()
                            self._log_line("  Valor preenchido.", "step")

                            loc_final = ctx.locator('input[type="submit"][ng-click="vm.finalizarSol()"]')
                            if loc_final.count()==0: loc_final=ctx.locator('input[type="submit"][value="Continuar"]')
                            wait_enabled(loc_final,STEP); loc_final.click(); pace()

                            loc_inc = ctx.locator('input[type="submit"][value="Incluir"]')
                            if loc_inc.count()==0: loc_inc=ctx.locator('input[type="submit"][ng-click="vm.IncluirConta()"]')
                            wait_enabled(loc_inc,STEP); loc_inc.click(); pace(1.4,2.1)

                            def try_grey():
                                loc_g=ctx.locator('input[type="button"].grey[ng-click="vm.continuar()"]')
                                if loc_g.count()==0: return False
                                wait_enabled(loc_g.first,STEP); loc_g.first.click(); return True

                            try_grey(); pace()
                            try: try_grey()
                            except: pass

                            loc_v = ctx.locator('input[type="button"].grey[ng-click="vm.voltarAplicAut()"]')
                            if loc_v.count()==0: loc_v=ctx.locator('input.grey[ng-click="vm.voltarAplicAut()"]')
                            if loc_v.count()>0:
                                wait_enabled(loc_v.first,max(STEP,25_000)); loc_v.first.click()

                            ctx_after = resolve_frame(120_000)
                            wait_nova(ctx_after,120_000)
                            self._log_line(f"  {client_name} → CONCLUÍDO ✓", "ok")
                            self._ui(lambda i=idx: self._set_card_state(i,"done",show=True))
                            break

                        except BPMUserCancelled: raise
                        except Exception as e:
                            self._log_line(f"  Tentativa {tentativa} falhou: {e}", "warn")
                            if tentativa >= 20:
                                self._log_line(f"  {client_name} → FALHA PERMANENTE ✗", "err")
                                self._ui(lambda i=idx: self._set_card_state(i,"error",show=True))
                                raise RuntimeError(f"Falha recorrente {client_name}: {e}") from e
                            try: page.goto(PAINEL, wait_until="domcontentloaded")
                            except: pass

            except BPMUserCancelled: self._log_line("Operações canceladas pelo usuário.", "warn")
            except Exception as e:
                if not self._cancel_requested:
                    em = str(e)
                    self._log_line(f"ERRO GERAL: {em}", "err")
                    self._ui(lambda: messagebox.showerror("Erro na rotina", em))
            finally:
                self._browser = None
                try: browser.close()
                except: pass
                self._worker_running = False
                self._started = False
                if not self._cancel_requested:
                    self.controller.bpm_run_selection = []

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mesa Itaú — Risco Sacado")
        self.geometry("1060x720")
        self.minsize(860, 580)
        self.configure(bg=C["bg"])
        self._setup_ttk_styles()
        ico_path = _ensure_ico_path()
        if ico_path:
            try:
                self.iconbitmap(default=ico_path)
            except Exception:
                pass
        self.overrideredirect(True)
        self.bpm_run_selection = []
        self.bpm_funcional     = ""
        self.invertido_limites_cache = {}
        self._limites_listeners = []
        self.bpm_password      = ""
        self.rotina_em_execucao= None
        self._active_frame     = "Home"

        self._shell = tk.Frame(self, bg=C["bg"])
        self._shell.pack(fill="both", expand=True)

        self._titlebar = AppTitleBar(self._shell, self)
        self._titlebar.pack(side="top", fill="x")

        self._main = tk.Frame(self._shell, bg=C["bg"])
        self._main.pack(fill="both", expand=True)

        self._sidebar = Sidebar(self._main, self)
        self._sidebar.pack(side="left", fill="y")

        make_hairline(self._main, orient="v", bg=C["hair"]).pack(side="left", fill="y")

        self._content = tk.Frame(self._main, bg=C["bg"])
        self._content.pack(side="left", fill="both", expand=True)
        self._content.rowconfigure(0, weight=1)
        self._content.columnconfigure(0, weight=1)

        self.frames = {}
        for Cls, name in [
            (HomeFrame,              "Home"),
            (RotinasFrame,           "Rotinas"),
            (ShareFrame,             "Share"),
            (BPMHubFrame,            "BPM_HUB"),
            (BPMConfigFrame,         "BPM_CONFIG"),
            (BPMNovaConfigFrame,     "BPM_CONFIG_NOVA"),
            (BPMFrame,               "BPM"),
            (OperacoesInvertidoFrame,"OperacoesInvertido"),
            (AnalisarOperacoesFrame, "AnalisarOperacoes"),
            (LimitesInvertidoFrame,  "LimitesInvertido"),
            (TaxasInvertidoFrame,    "TaxasInvertido"),
        ]:
            f = Cls(self._content, self)
            self.frames[name] = f
            f.grid(row=0, column=0, sticky="nsew")

        self._statusbar = AppStatusBar(self._shell, self)
        self._statusbar.pack(side="bottom", fill="x")

        self.show_frame("Home")
        self.after(120, self._apply_window_chrome)

    def _apply_window_chrome(self):
        apply_modern_window_chrome(self)
        apply_frameless_resize(self)
        apply_windows_shell(self)

    def _setup_ttk_styles(self):
        s = ttk.Style(self)
        try: s.theme_use("clam")
        except: pass
        s.configure("TCombobox",
                    fieldbackground=C["bg"], background=C["surface"],
                    foreground=C["ink"], selectbackground=C["surface2"],
                    selectforeground=C["ink"], borderwidth=1,
                    lightcolor=C["hair"], darkcolor=C["hair"])
        s.map("TCombobox", fieldbackground=[("readonly",C["bg"])],
              selectbackground=[("!focus",C["surface2"])])

    def _refresh_frame_scroll(self, frame):
        if isinstance(frame, ScrollableFrame):
            frame.refresh_bindings()
        for child in frame.winfo_children():
            self._refresh_frame_scroll(child)

    def register_limites_listener(self, fn):
        if fn not in self._limites_listeners:
            self._limites_listeners.append(fn)

    def unregister_limites_listener(self, fn):
        try:
            self._limites_listeners.remove(fn)
        except ValueError:
            pass

    def publish_limite_result(self, cnpj_digits, data):
        self.invertido_limites_cache[cnpj_digits] = data
        for fn in list(self._limites_listeners):
            try:
                fn(cnpj_digits)
            except Exception:
                pass

    def show_frame(self, name):
        if name == "BPM" and not getattr(self,"bpm_run_selection",[]):
            name = "BPM_HUB"
        f = self.frames.get(name)
        if f is None: return
        self._active_frame = name
        sidebar_name = name
        if name in ("BPM", "BPM_CONFIG", "BPM_CONFIG_NOVA", "BPM_HUB"):
            sidebar_name = "BPM"
        elif name in ("OperacoesInvertido", "LimitesInvertido", "AnalisarOperacoes",
                      "TaxasInvertido"):
            sidebar_name = "OperacoesInvertido"
        self._sidebar.set_active(sidebar_name)
        self._titlebar.set_module(name)
        self._statusbar.set_module(name)
        f.tkraise()
        self._refresh_frame_scroll(f)
        if hasattr(f,"on_show"):
            try: f.on_show()
            except: pass

if __name__ == "__main__":
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
        except Exception:
            pass
    App().mainloop()
